"""
通用接口路由
"""
import os
import json
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, List, Optional
from fastapi import APIRouter, Request, HTTPException, status, Body
from fastapi.responses import JSONResponse
from pydantic import BaseModel

from services.api.shared.config import logger, project_root, LMS_BASE_URL, MOCK_USER
from services.api.shared.operation_log import (
    log_operation,
    get_recent_operations,
    get_all_operations,
    OPERATION_LOGS_DIR,
    LOG_TYPES,
)

router = APIRouter(prefix="/api", tags=["common"])


# 管理员联系方式配置
ADMIN_CONTACTS = {
    "phone": os.getenv("ADMIN_PHONE", "13800138000"),
    "email": os.getenv("ADMIN_EMAIL", "admin@example.com"),
    "wechat": os.getenv("ADMIN_WECHAT", "admin_wechat"),
}


class OperationLogCreate(BaseModel):
    """创建操作记录请求模型"""
    id: Optional[str] = None
    timestamp: Optional[str] = None
    operation_type: str = "other"
    user_id: Optional[str] = None
    user_name: Optional[str] = None
    action: str = ""
    target: Optional[str] = None
    status: str = "success"
    details: Dict[str, Any] = {}
    ip_address: Optional[str] = None
    metadata: Dict[str, Any] = {}


@router.get("/adminContact")
async def get_admin_contact():
    """获取管理员联系方式"""
    return JSONResponse(
        status_code=status.HTTP_200_OK,
        content={
            "code": 200,
            "message": "获取管理员联系方式成功",
            "data": ADMIN_CONTACTS
        }
    )


@router.get("/operationLogs")
async def get_operation_logs(days: int = 180, limit: Optional[int] = None):
    """获取操作记录"""
    try:
        operations = get_all_operations(days=days)
        if limit:
            operations = operations[:limit]

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取操作记录成功",
                "data": {
                    "logs": operations,
                    "total": len(operations)
                }
            }
        )
    except Exception as e:
        logger.error(f"获取操作记录失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取操作记录失败: {str(e)}"
        )


@router.post("/operationLogs")
async def create_operation_log(request: Request):
    """创建操作记录"""
    try:
        data = await request.json()

        # 如果前端没有传递 user_name，尝试从 auth token 获取
        user_name = data.get("user_name")
        user_id = data.get("user_id")

        if not user_name:
            auth_token = request.headers.get("authToken")
            if auth_token:
                # 直接调用 LMS 接口获取用户信息
                try:
                    import requests
                    lms_auth_url = f"{LMS_BASE_URL}/auth/token?token={auth_token}"
                    response = requests.get(lms_auth_url, timeout=5)
                    if response.status_code == 200:
                        user_info = response.json().get("data", {})
                        user_name = user_info.get("userName")
                        user_id = user_id or user_info.get("userId")
                        logger.info(f"从 LMS 获取用户信息: userName={user_name}")
                    else:
                        # LMS 不可用，使用模拟用户
                        user_name = MOCK_USER.get("userName", "管理员账号")
                        user_id = user_id or MOCK_USER.get("userId", "1000000")
                        logger.info(f"LMS 不可用，使用模拟用户: {user_name}")
                except Exception as e:
                    # 任何错误都使用模拟用户
                    user_name = MOCK_USER.get("userName", "管理员账号")
                    user_id = user_id or MOCK_USER.get("userId", "1000000")
                    logger.warning(f"获取用户信息失败，使用模拟用户: {str(e)}")

        # 如果还是没有用户名，使用默认值
        if not user_name:
            user_name = "未知操作员"

        log_success = log_operation(
            operation_type=data.get("operation_type", "other"),
            action=data.get("action", ""),
            user_id=user_id,
            user_name=user_name,
            target=data.get("target"),
            status=data.get("status", "success"),
            details=data.get("details", {}),
            ip_address=data.get("ip_address"),
            metadata=data.get("metadata", {})
        )

        if log_success:
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "操作记录已保存",
                    "data": {"success": True}
                }
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="保存操作记录失败"
            )

    except Exception as e:
        logger.error(f"创建操作记录失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"创建操作记录失败: {str(e)}"
        )


@router.delete("/operationLogs")
async def delete_operation_logs(request: Request):
    """删除操作记录"""
    try:
        user_level = request.headers.get("X-User-Level")
        if user_level != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="权限不足，仅管理员可删除操作记录"
            )

        data = await request.json()
        log_ids = data.get("logIds", [])

        if not log_ids:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="请提供要删除的记录ID列表"
            )

        deleted_count = 0
        for log_id in log_ids:
            for log_type in LOG_TYPES:
                log_file = OPERATION_LOGS_DIR / log_type / f"{log_id}.json"
                if log_file.exists():
                    log_file.unlink()
                    deleted_count += 1
                    break

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": f"已删除 {deleted_count} 条记录",
                "data": {"deletedCount": deleted_count}
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除操作记录失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"删除操作记录失败: {str(e)}"
        )


@router.post("/operationLogs/autoCleanup")
async def auto_cleanup_logs(request: Request):
    """自动清理过期的操作记录"""
    try:
        user_level = request.headers.get("X-User-Level")
        if user_level != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="权限不足"
            )

        data = await request.json()
        days = data.get("days", 180)

        deleted_count = 0
        cutoff_date = datetime.now() - timedelta(days=days)

        for log_type in LOG_TYPES:
            log_dir = OPERATION_LOGS_DIR / log_type
            if not log_dir.exists():
                continue

            for log_file in log_dir.glob("*.json"):
                try:
                    with open(log_file, 'r', encoding='utf-8') as f:
                        log_data = json.load(f)
                    log_time_str = log_data.get("timestamp", "")
                    if log_time_str:
                        try:
                            log_time = datetime.fromisoformat(log_time_str.replace('Z', '+00:00'))
                            if log_time < cutoff_date:
                                log_file.unlink()
                                deleted_count += 1
                        except:
                            pass
                except Exception as e:
                    logger.warning(f"处理日志文件失败 {log_file}: {str(e)}")
                    continue

        log_operation(
            operation_type="system_cleanup",
            action="自动清理操作记录",
            status="success",
            details={"days": days, "deleted_count": deleted_count}
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": f"已清理 {deleted_count} 条过期记录",
                "data": {"deletedCount": deleted_count}
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"自动清理操作记录失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"自动清理操作记录失败: {str(e)}"
        )


@router.get("/dashboard/stats")
async def get_dashboard_stats():
    """获取仪表盘统计数据 - 储位数和品类数"""
    try:
        import openpyxl

        # 原始文件路径
        bins_file = project_root / "services" / "sim" / "lms" / "bins_data.xlsx"
        category_file = project_root / "shared" / "data" / "烟箱信息汇总完整版.xlsx"

        bin_count = 0
        category_count = 0

        # 读取储位数
        if bins_file.exists():
            try:
                wb = openpyxl.load_workbook(bins_file, read_only=True)
                ws = wb.active
                bin_count = max(0, ws.max_row - 1)  # 总行数 - 1 (减去表头)
                wb.close()
            except Exception as e:
                logger.error(f"读取储位文件失败: {str(e)}")

        # 读取品类数
        if category_file.exists():
            try:
                wb = openpyxl.load_workbook(category_file, read_only=True)
                ws = wb.active
                category_count = max(0, ws.max_row - 1)  # 总行数 - 1 (减去表头)
                wb.close()
            except Exception as e:
                logger.error(f"读取品类文件失败: {str(e)}")

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取统计数据成功",
                "data": {
                    "bin_count": bin_count,
                    "category_count": category_count
                }
            }
        )

    except Exception as e:
        logger.error(f"获取统计数据失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取统计数据失败: {str(e)}"
        )
