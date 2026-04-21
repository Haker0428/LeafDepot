"""
盘点任务路由
"""
import os
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List, Set

from fastapi import APIRouter, Request, BackgroundTasks, HTTPException, status, Body
from fastapi.responses import JSONResponse, Response
import pandas as pd

from services.api.shared.config import (
    logger,
    project_root,
    ENABLE_BARCODE,
    DETECT_MODULE_AVAILABLE,
    BARCODE_MODULE_AVAILABLE,
    IS_SIM,
    ENABLE_DEBUG,
    ENABLE_VISUALIZATION,
)
from services.api.shared.models import (
    TaskStatus,
    BinLocationStatus,
    InventoryTaskProgress,
    ScanAndRecognizeRequest,
)
from services.api.shared.operation_log import log_operation
from services.api.shared.tobacco_resolver import get_tobacco_case_resolver
from services.api.shared.excel_writer import build_excel_data, write_excel

# 从 service.py 导入核心函数和状态存储
from services.api.inventory.service import (
    execute_inventory_workflow,
    get_task_state_storage,
    inventory_tasks,
    inventory_task_bins,
    inventory_task_details,
    abort_inventory_task,  # TODO: RCS cancel API 就位后，取消接口中将调用此函数
    _get_next_task_no,
)
from services.api.shared.websocket_manager import ws_manager
from services.api.inventory.task_state import on_server_startup, clear_task

from services.api.auth.router import get_user_info_from_token

# 条形码识别模块（条件导入）
if ENABLE_BARCODE and BARCODE_MODULE_AVAILABLE:
    from core.vision.barcode_recognizer import BarcodeRecognizer
else:
    BarcodeRecognizer = None

# 检测模块（条件导入）
if DETECT_MODULE_AVAILABLE:
    from core.detection import count_boxes
else:
    count_boxes = None

router = APIRouter(prefix="/api/inventory", tags=["inventory"])


# ==================== 用户进行中任务查询 ====================

@router.get("/running-task")
async def get_running_task(request: Request):
    """查询当前登录用户是否有正在进行的盘点任务，有则返回任务信息"""
    auth_token = request.headers.get("authToken")
    user_info = await get_user_info_from_token(auth_token) if auth_token else {}
    user_id = user_info.get("userId", "")
    if not user_id:
        return JSONResponse(status_code=200, content={"code": 200, "data": None})

    for running_task_no, task_status in inventory_tasks.items():
        if task_status.status == "running":
            task_user_info = inventory_task_details.get(running_task_no, {}).get("userInfo", {})
            if task_user_info.get("userId") == user_id:
                return JSONResponse(
                    status_code=200,
                    content={
                        "code": 200,
                        "data": {
                            "taskNo": running_task_no,
                            "operatorName": task_user_info.get("userName", ""),
                            "startTime": task_status.start_time or "",
                            "totalBins": task_status.total_bins or 0,
                            "completedBins": task_status.completed_bins or 0,
                        }
                    }
                )
    return JSONResponse(status_code=200, content={"code": 200, "data": None})


# ==================== 任务进度接口 ====================

@router.post("/start-inventory")
async def start_inventory(request: Request, background_tasks: BackgroundTasks):
    """启动盘点任务，接收任务编号和储位名称列表"""
    try:
        data = await request.json()
        task_no = data.get("taskNo") or _get_next_task_no()
        bin_locations = data.get("binLocations", [])
        is_sim = IS_SIM  # 使用配置文件中的值
        inventory_items = data.get("inventoryItems", [])
        # 重新盘点时使用独立的临时任务号，避免覆盖原任务状态
        recount_task_id = data.get("recountTaskId")
        storage_key = recount_task_id if recount_task_id else task_no

        if not task_no or not bin_locations:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="任务编号和储位名称列表不能为空"
            )

        logger.info(f"启动盘点任务: {task_no}, 包含 {len(bin_locations)} 个储位, 模拟模式: {is_sim}, 存储键: {storage_key}")

        # 重新盘点的独立任务号不受"任务已在执行中"检查影响
        if not recount_task_id:
            # 已取消的任务，清除内存状态，允许重新下发
            if task_no in inventory_tasks and inventory_tasks[task_no].status == "cancelled":
                inventory_tasks.pop(task_no, None)
                inventory_task_details.pop(task_no, None)
                inventory_task_bins.pop(task_no, None)

        # 保存原始盘点项信息（使用存储键）
        if storage_key not in inventory_task_details:
            inventory_task_details[storage_key] = {}
        inventory_task_details[storage_key]["inventoryItems"] = inventory_items

        # 记录盘点任务启动
        auth_token = request.headers.get("authToken")
        logger.info(f"收到 authToken: {auth_token}")
        user_info = await get_user_info_from_token(auth_token) if auth_token else {}
        logger.info(f"获取到的 user_info: {user_info}")

        # ===== 全局并发检查：不允许两个盘点任务同时运行 =====
        for running_task_no, task_status in inventory_tasks.items():
            if task_status.status == "running":
                # 获取正在运行任务的操作人信息
                running_user_info = {}
                if running_task_no in inventory_task_details:
                    running_user_info = inventory_task_details[running_task_no].get("userInfo", {})
                operator_name = running_user_info.get("userName", "未知")
                operator_id = running_user_info.get("userId", "")
                start_time = task_status.start_time or ""
                conflict_msg = (
                    f"有其他盘点任务正在进行中（任务号: {running_task_no}，"
                    f"操作人: {operator_name}{f'（{operator_id}）' if operator_id else ''}，"
                    f"开始时间: {start_time}），请等待该任务结束后再下发新任务。"
                )
                logger.warning(f"[并发冲突] 用户 {user_info.get('userName','未知')} 试图下发任务 {task_no}，但任务 {running_task_no} 正在运行")
                return JSONResponse(
                    status_code=status.HTTP_409_CONFLICT,
                    content={
                        "code": 409,
                        "message": conflict_msg,
                        "data": {
                            "runningTaskNo": running_task_no,
                            "operatorName": operator_name,
                            "operatorId": operator_id,
                            "startTime": start_time,
                        }
                    }
                )

        # 保存用户信息到任务详情，供后台任务使用
        inventory_task_details[storage_key]["userInfo"] = user_info
        # 同时记录原始任务号（用于合并保存）
        inventory_task_details[storage_key]["originalTaskNo"] = task_no

        log_operation(
            operation_type="inventory",
            action="启动盘点任务",
            user_id=user_info.get("userId"),
            user_name=user_info.get("userName"),
            target=task_no,
            status="running",
            details={
                "task_no": task_no,
                "bin_locations": bin_locations,
                "bin_count": len(bin_locations),
                "is_sim": is_sim
            }
        )

        # 在后台异步执行盘点任务
        background_tasks.add_task(
            execute_inventory_workflow,
            task_no=storage_key,  # 使用存储键，避免与原任务冲突
            bin_locations=bin_locations,
            is_sim=is_sim
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "盘点任务已启动",
                "data": {
                    "taskNo": task_no,
                    "actualTaskNo": storage_key,  # 返回实际用于轮询的任务号
                    "bin_locations": bin_locations,
                    "is_sim": is_sim
                }
            }
        )

    except Exception as e:
        logger.error(f"启动盘点任务失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"启动盘点任务失败: {str(e)}"
        )


@router.get("/progress")
async def get_inventory_progress(taskNo: str):
    """获取盘点任务进度"""
    try:
        if taskNo not in inventory_tasks:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "任务不存在",
                    "data": None
                }
            )

        task_status = inventory_tasks[taskNo]
        bin_statuses = inventory_task_bins.get(taskNo, [])

        completed_count = sum(1 for bin in bin_statuses if bin.status == "completed")
        progress_percentage = (completed_count / task_status.total_steps * 100) if task_status.total_steps > 0 else 0

        progress_data = InventoryTaskProgress(
            task_no=task_status.task_no,
            status=task_status.status,
            current_step=task_status.current_step,
            total_steps=task_status.total_steps,
            progress_percentage=round(progress_percentage, 2),
            bin_locations=bin_statuses,
            start_time=task_status.start_time,
            end_time=task_status.end_time
        )

        # 任务失败时附带错误类型，便于前端展示错误弹窗
        response_data = progress_data.dict()
        if task_status.status == "failed":
            response_data["error_type"] = task_status.error_type or "other"
            response_data["message"] = task_status.error_message or "任务失败"

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取进度成功",
                "data": response_data
            }
        )
    except Exception as e:
        logger.error(f"获取任务进度失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取任务进度失败: {str(e)}"
        )


@router.get("/results")
async def get_inventory_results(taskNo: str):
    """获取盘点任务的完整结果"""
    try:
        if taskNo not in inventory_tasks:
            return JSONResponse(
                    status_code=status.HTTP_404_NOT_FOUND,
                    content={
                        "code": 404,
                        "message": "任务不存在",
                        "data": None
                    }
                )

        task_status = inventory_tasks[taskNo]

        if task_status.status == "failed":
            # 全部库位失败时，携带错误信息返回
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": task_status.error_message or "任务失败",
                    "errorType": task_status.error_type or "other",
                    "data": {
                        "taskNo": taskNo,
                        "status": task_status.status,
                        "currentStep": task_status.current_step,
                        "totalSteps": task_status.total_steps,
                        "inventoryResults": []
                    }
                }
            )

        # completed 或 partial：返回已收集到的结果
        inventory_results = []
        if taskNo in inventory_task_details and "inventoryResults" in inventory_task_details[taskNo]:
            inventory_results = inventory_task_details[taskNo]["inventoryResults"]

        message = "获取盘点结果成功" if task_status.status == "completed" else "部分库位盘点完成"
        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": message,
                "data": {
                    "taskNo": taskNo,
                    "status": task_status.status,
                    "inventoryResults": inventory_results
                }
            }
        )
    except Exception as e:
        logger.error(f"获取盘点结果失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取盘点结果失败: {str(e)}"
        )


@router.get("/image")
async def get_inventory_image(
    taskNo: str,
    binLocation: str,
    cameraType: str,
    filename: str,
    source: str = "output"
):
    """获取盘点任务中的图片"""
    try:
        if source == "capture_img":
            image_path = project_root / "capture_img" / taskNo / binLocation / cameraType / filename
        else:
            image_path = project_root / "output" / taskNo / binLocation / cameraType / filename

        if not image_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"图片不存在: {filename} (路径: {image_path})"
            )

        with open(image_path, "rb") as f:
            image_data = f.read()

        media_type = "image/jpeg"
        if filename.endswith(".png"):
            media_type = "image/png"
        elif filename.endswith(".bmp"):
            media_type = "image/bmp"

        return Response(content=image_data, media_type=media_type)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取图片失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取图片失败: {str(e)}"
        )


@router.get("/task-detail")
async def get_task_detail(taskNo: str, binLocation: str):
    """获取任务的详细信息"""
    try:
        storage = get_task_state_storage()
        task_details = storage["details"]

        if taskNo in task_details and binLocation in task_details[taskNo]:
            detail = task_details[taskNo][binLocation]
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "获取任务详情成功",
                    "data": detail
                }
            )
        else:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "任务详情不存在",
                    "data": None
                }
            )

    except Exception as e:
        logger.error(f"获取任务详情失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取任务详情失败: {str(e)}"
        )


@router.post("/scan-and-recognize")
async def scan_and_recognize(request: ScanAndRecognizeRequest = Body(...)):
    """扫码+识别接口"""
    try:
        image_path = f"{request.taskNo}/{request.binLocation}/3d_camera/"
        image_dir = project_root / "capture_img" / image_path

        recognition_time = datetime.now().isoformat()
        results = {
            "taskNo": request.taskNo,
            "binLocation": request.binLocation,
            "recognition_time": recognition_time,
            "detect_result": None,
            "barcode_result": None,
            "photos": []
        }

        # 查找照片
        try:
            photos = []
            image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']

            if image_dir.exists():
                for ext in image_extensions:
                    for common_name in ['main', 'MAIN', 'raw', 'RAW', 'image', 'IMAGE']:
                        main_file = image_dir / f"{common_name}{ext}"
                        if main_file.exists():
                            relative_path = f"{request.taskNo}/{request.binLocation}/3d_camera/{common_name.upper()}{ext}"
                            photos.append(f"/{relative_path}")
                            break
                    if photos:
                        break

            results["photos"] = photos
        except Exception as e:
            logger.error(f"查找照片路径失败: {str(e)}")
            results["photos"] = []

        # Barcode 模块（处理 scan_camera_1 和 scan_camera_2）
        detected_pile_id = request.pile_id
        scan_dir_1 = image_dir.parent / "scan_camera_1"
        scan_dir_2 = image_dir.parent / "scan_camera_2"
        if ENABLE_BARCODE and BARCODE_MODULE_AVAILABLE and BarcodeRecognizer:
            try:
                recognizer = BarcodeRecognizer(code_type=request.code_type)
                all_barcode_results = []
                for scan_dir in [scan_dir_1, scan_dir_2]:
                    if scan_dir.exists():
                        barcode_results = recognizer.process_folder(input_dir=str(scan_dir))
                        all_barcode_results.extend(barcode_results)
                resolver = get_tobacco_case_resolver()

                resolved_info = None
                for br in all_barcode_results:
                    # 从 recognizer 结果中提取真正的条码字符串
                    raw = br.get('output')
                    barcode_text = None
                    if raw:
                        try:
                            data = json.loads(raw)
                            for session in data.get('sessions', []):
                                for bc in session.get('barcodes', []):
                                    if bc.get('text'):
                                        barcode_text = bc['text']
                                        break
                                if barcode_text:
                                    break
                        except Exception:
                            pass
                    if not barcode_text:
                        barcode_text = br.get('text')
                    if barcode_text:
                        resolved_info = resolver.resolve(barcode_text)
                        if resolved_info['success']:
                            break

                if resolved_info and resolved_info['success']:
                    detected_pile_id = resolved_info['pile_id']
                    results["barcode_result"] = {
                        "image_path": str(scan_dir_1),
                        "code_type": request.code_type,
                        "six_digit_code": resolved_info['six_digit_code'],
                        "stack_type_1": resolved_info['stack_type_1'],
                        "product_name": resolved_info['product_name'],
                        "tobacco_code": resolved_info['tobacco_code'],
                        "mapped_pile_id": detected_pile_id,
                        "total_images": len(all_barcode_results),
                        "status": "success"
                    }
                else:
                    results["barcode_result"] = {
                        "image_path": str(scan_dir_1),
                        "status": "no_match",
                        "message": "条码识别成功但未匹配到烟箱信息"
                    }
            except Exception as e:
                logger.error(f"Barcode模块识别失败: {str(e)}")
                results["barcode_result"] = {"status": "failed", "error": str(e)}
        else:
            results["barcode_result"] = {"status": "disabled"}

        # Detect 模块
        if DETECT_MODULE_AVAILABLE and count_boxes:
            try:
                if not image_dir.exists() or not image_dir.is_dir():
                    return JSONResponse(
                        status_code=status.HTTP_200_OK,
                        content={"code": 200, "message": "图片目录不存在", "data": results}
                    )

                image_files = []
                image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
                for name in ['main', 'raw', 'image']:
                    for ext in image_extensions:
                        common_file = image_dir / f"{name}{ext}"
                        if common_file.exists():
                            image_files.append(common_file)
                            break
                    if image_files:
                        break

                if not image_files:
                    for ext in image_extensions:
                        image_files.extend(list(image_dir.glob(f"*{ext}")))
                        if image_files:
                            break

                if image_files:
                    debug_output_dir = project_root / "debug" / request.taskNo / request.binLocation
                    debug_output_dir.mkdir(parents=True, exist_ok=True)

                    depth_path = image_dir / "depth.jpg"
                    total_count = count_boxes(
                        image_path=str(image_files[0]),
                        pile_id=detected_pile_id,
                        depth_image_path=str(depth_path) if depth_path.exists() else None,
                        enable_debug=ENABLE_DEBUG,
                        enable_visualization=ENABLE_VISUALIZATION,
                        output_dir=str(debug_output_dir)
                    )

                    results["detect_result"] = {
                        "image_path": str(image_files[0]),
                        "pile_id": detected_pile_id,
                        "total_count": total_count,
                        "status": "success"
                    }
            except Exception as e:
                logger.error(f"Detect模块识别失败: {str(e)}")
                results["detect_result"] = {"status": "failed", "error": str(e)}

        # 更新任务状态
        storage = get_task_state_storage()
        task_bins = storage["bins"]
        task_details = storage["details"]

        if request.taskNo in task_bins:
            for bin_status in task_bins[request.taskNo]:
                if bin_status.bin_location == request.binLocation:
                    bin_status.detect_result = results["detect_result"]
                    bin_status.barcode_result = results["barcode_result"]
                    bin_status.recognition_time = recognition_time
                    break

        if request.taskNo not in task_details:
            task_details[request.taskNo] = {}
        if request.binLocation not in task_details[request.taskNo]:
            task_details[request.taskNo][request.binLocation] = {}
        task_details[request.taskNo][request.binLocation]["recognition"] = results

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": "扫码+识别执行完成", "data": results}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"扫码+识别失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"扫码+识别失败: {str(e)}"
        )


@router.get("/recognition-result")
async def get_recognition_result(taskNo: str, binLocation: str):
    """读取识别结果接口"""
    try:
        storage = get_task_state_storage()
        task_bins = storage["bins"]
        task_details = storage["details"]

        result_data = {
            "taskNo": taskNo,
            "binLocation": binLocation,
            "detect_result": None,
            "barcode_result": None,
            "recognition_time": None
        }

        if taskNo in task_bins:
            for bin_status in task_bins[taskNo]:
                if bin_status.bin_location == binLocation:
                    result_data["detect_result"] = bin_status.detect_result
                    result_data["barcode_result"] = bin_status.barcode_result
                    result_data["recognition_time"] = bin_status.recognition_time
                    break

        if not result_data["detect_result"] and taskNo in task_details:
            if binLocation in task_details[taskNo]:
                recognition_data = task_details[taskNo][binLocation].get("recognition")
                if recognition_data:
                    result_data["detect_result"] = recognition_data.get("detect_result")
                    result_data["barcode_result"] = recognition_data.get("barcode_result")
                    result_data["recognition_time"] = recognition_data.get("recognition_time")

        if not result_data["detect_result"] and not result_data["barcode_result"]:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={"code": 404, "message": "识别结果不存在", "data": None}
            )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": "获取识别结果成功", "data": result_data}
        )

    except Exception as e:
        logger.error(f"获取识别结果失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取识别结果失败: {str(e)}"
        )


@router.post("/save-results")
async def save_inventory_results(request: Request):
    """保存盘点结果并生成Excel文件"""
    try:
        data = await request.json()
        task_no = data.get("taskNo")
        inventory_results = data.get("inventoryResults", [])

        if not task_no:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="任务编号不能为空")

        if not inventory_results:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="盘点结果不能为空")

        logger.info(f"保存盘点结果: {task_no}, 共 {len(inventory_results)} 个储位")

        # 获取操作员信息
        user_info = data.get("userInfo", {})
        operator_name = user_info.get("userName", "")

        # 获取人工校准过的储位列表（点击对勾才记录，与品规检测结果无关）
        manual_calibrated: Set[str] = set(data.get("manualCalibratedLocations", []))
        # 兼容旧版 calibrationRecords（保留向后兼容）
        calibration_records: Dict[str, Dict[str, Any]] = data.get("calibrationRecords", {})

        # 定义输出目录（merge 和非 merge 路径都需要）
        output_dir = project_root / "output" / "history_data"
        output_dir.mkdir(parents=True, exist_ok=True)
        xlsx_file = output_dir / f"{task_no}.xlsx"

        # 是否为合并模式：merge=true 时读取现有文件，只更新传入的储位结果
        merge_mode = data.get("merge", False)

        if merge_mode and xlsx_file.exists():
            # 合并模式：读取现有数据，只更新传入的储位
            existing_df = pd.read_excel(xlsx_file, sheet_name='盘点结果')
            # 建立新结果映射：key=储位名称
            new_results_map = {}
            for result in inventory_results:
                bin_loc = result.get("binLocation", "")
                if bin_loc:
                    new_results_map[bin_loc] = result

            # 更新对应的行
            updated_rows = 0
            for idx, row in existing_df.iterrows():
                bin_loc = str(row.get("储位名称", ""))
                if bin_loc in new_results_map:
                    result = new_results_map[bin_loc]
                    is_manually_calibrated = (
                        bin_loc in manual_calibrated or
                        bool(calibration_records.get(bin_loc, {}).get("specModified")) or
                        bool(calibration_records.get(bin_loc, {}).get("quantityModified"))
                    )
                    mod_record = "人工修改" if is_manually_calibrated else ""
                    spec_name = result.get("specName", "")
                    actual_spec = result.get("actualSpec", "")
                    quantity_diff = result.get("difference", 0)
                    if actual_spec and actual_spec != spec_name and actual_spec != "未识别":
                        diff_desc = "品规不一致"
                    elif actual_spec == "未识别":
                        diff_desc = "品规不一致"
                    elif quantity_diff != 0:
                        diff_desc = quantity_diff
                    else:
                        diff_desc = "一致"
                    existing_df.at[idx, "实际品规"] = result.get("actualSpec", "")
                    existing_df.at[idx, "库存数量"] = result.get("systemQuantity", 0)
                    existing_df.at[idx, "实际数量"] = result.get("actualQuantity", 1)
                    existing_df.at[idx, "差异"] = diff_desc
                    existing_df.at[idx, "修改记录"] = mod_record
                    # 手动保存时将有效状态提升为"有效"
                    existing_df.at[idx, "有效状态"] = "有效"
                    existing_df.at[idx, "照片1路径"] = result.get("photo3dPath", "")
                    existing_df.at[idx, "照片2路径"] = result.get("photoDepthPath", "")
                    existing_df.at[idx, "照片3路径"] = result.get("photoScan1Path", "")
                    existing_df.at[idx, "照片4路径"] = result.get("photoScan2Path", "")
                    updated_rows += 1
            # 手动合并确认：将所有行的有效状态提升为"有效"
            if "有效状态" in existing_df.columns:
                existing_df["有效状态"] = "有效"
            df = existing_df
            logger.info(f"合并保存：更新了 {updated_rows} 个储位")
        else:
            # 全新保存模式：使用共享工具函数构建 DataFrame
            df = build_excel_data(
                task_no=task_no,
                inventory_results=inventory_results,
                operator_name=operator_name,
                manual_calibrated=manual_calibrated,
                calibration_records=calibration_records,
            )

        # 写入 Excel 文件
        write_excel(task_no, df, output_dir)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "保存成功",
                "data": {
                    "taskNo": task_no,
                    "xlsxFile": str(xlsx_file),
                    "xlsxUrl": f"/api/inventory/download-xlsx?taskNo={task_no}",
                    "count": len(inventory_results)
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"保存盘点结果失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"保存盘点结果失败: {str(e)}"
        )


@router.get("/download-xlsx")
async def download_xlsx(taskNo: str):
    """下载Excel文件"""
    try:
        output_dir = project_root / "output" / "history_data"
        xlsx_file = output_dir / f"{taskNo}.xlsx"

        if not xlsx_file.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"Excel文件不存在: {taskNo}.xlsx")

        with open(xlsx_file, "rb") as f:
            excel_data = f.read()

        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={taskNo}.xlsx"}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载Excel文件失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"下载Excel文件失败: {str(e)}")


@router.post("/update-bins-data")
async def update_bins_data(request: Request):
    """更新 bins_data.xlsx 中的数量和品规"""
    try:
        data = await request.json()
        inventory_items = data.get("inventoryItems", [])

        if not inventory_items:
            return {"code": 400, "message": "没有盘点数据"}

        bins_data_path = project_root / "services" / "sim" / "lms" / "bins_data.xlsx"
        category_file = project_root / "shared" / "data" / "烟箱信息汇总完整版.xlsx"

        if not bins_data_path.exists():
            return {"code": 500, "message": "bins_data.xlsx 文件不存在"}

        # 建立 品名→品规代码 的映射表（从品类汇总文件读取）
        name_to_code = {}
        if category_file.exists():
            try:
                import openpyxl
                cat_wb = openpyxl.load_workbook(category_file, read_only=True)
                cat_ws = cat_wb.active
                for row in cat_ws.iter_rows(min_row=2, values_only=True):
                    if row and row[0] and row[1]:
                        name_to_code[str(row[0]).strip()] = str(row[1]).strip()
                cat_wb.close()
            except Exception as ex:
                logger.warning(f"读取品类汇总文件失败，使用现有品规代码: {ex}")

        from openpyxl import load_workbook
        wb = load_workbook(bins_data_path)
        ws = wb.active

        # 准备待更新数据：key=储位名称, value={数量, 品规名称, 品规代码}
        location_to_data = {}
        for item in inventory_items:
            location_name = item.get("locationName") or item.get("binDesc")
            actual_quantity = item.get("actualQuantity")
            actual_spec = item.get("actualSpec") or item.get("tobaccoName") or item.get("productName")
            if location_name and actual_quantity is not None:
                # 查找对应品规代码
                spec_code = name_to_code.get(actual_spec, "") if actual_spec else ""
                location_to_data[location_name] = {
                    "quantity": actual_quantity,
                    "specName": actual_spec,
                    "specCode": spec_code,
                }

        updated_count = 0
        for row in range(2, ws.max_row + 1):
            location_name = ws.cell(row, 5).value  # Column E: 储位名称
            if location_name and location_name in location_to_data:
                item_data = location_to_data[location_name]
                ws.cell(row, 8).value = item_data["quantity"]  # Column H: 数量(万支)
                if item_data["specName"]:
                    ws.cell(row, 10).value = item_data["specName"]  # Column J: 品规名称
                    if item_data["specCode"]:
                        ws.cell(row, 9).value = item_data["specCode"]  # Column I: 品规代码
                updated_count += 1

        wb.save(bins_data_path)
        wb.close()

        return {"code": 200, "message": f"成功更新 {updated_count} 个储位的数量和品规", "data": {"updatedCount": updated_count}}

    except Exception as e:
        logger.error(f"更新 bins_data.xlsx 失败: {str(e)}")
        return {"code": 500, "message": f"更新失败: {str(e)}"}


@router.get("/get-local-bins")
async def get_local_bins():
    """从本地 bins_data.xlsx 读取储位数据（用于模拟模式）"""
    try:
        bins_data_path = project_root / "services" / "sim" / "lms" / "bins_data.xlsx"

        if not bins_data_path.exists():
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={"code": 404, "message": "bins_data.xlsx 文件不存在", "data": []}
            )

        from openpyxl import load_workbook
        wb = load_workbook(bins_data_path, read_only=True)
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        bins_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 跳过空行
                continue

            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            # 构建与 LMS 接口兼容的数据格式
            bin_item = {
                "whCode": row_dict.get("仓库编码", ""),
                "areaCode": row_dict.get("区域编码", ""),
                "areaName": row_dict.get("区域名称", ""),
                "binCode": row_dict.get("储位编码", ""),
                "binDesc": row_dict.get("储位名称", ""),
                "maxQty": row_dict.get("最大数量", 0),
                "binStatus": row_dict.get("储位状态", ""),
                "tobaccoQty": row_dict.get("数量(万支)", 0) or 0,
                "tobaccoCode": row_dict.get("烟草编码", ""),
                "tobaccoName": row_dict.get("品规名称", ""),
                "rcsCode": row_dict.get("RCS编码", ""),
            }
            bins_list.append(bin_item)

        wb.close()

        logger.info(f"从本地文件读取 {len(bins_list)} 条储位数据")
        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": "获取本地储位数据成功", "data": bins_list}
        )

    except Exception as e:
        logger.error(f"读取本地储位数据失败: {str(e)}")
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"code": 500, "message": f"读取失败: {str(e)}", "data": []}
        )


# ==================== 取消任务接口 ====================

@router.post("/cancel-inventory")
async def cancel_inventory(taskNo: str):
    """取消正在运行的盘点任务"""
    robot_task_code = ""
    if taskNo in inventory_tasks:
        robot_task_code = inventory_tasks[taskNo].robot_task_code or ""
        inventory_tasks[taskNo].status = "cancelled"
        clear_task(taskNo)
        logger.info(f"任务已取消: {taskNo}, robotTaskCode={robot_task_code}")

        # 广播任务取消通知
        user_info = inventory_task_details.get(taskNo, {}).get("userInfo", {})
        await ws_manager.broadcast_task_event(
            "task_cancelled",
            taskNo,
            {
                "taskNo": taskNo,
                "operatorName": user_info.get("userName", ""),
            }
        )

    # TODO: RCS 提供 cancel API 后，在此调用 abort_inventory_task
    # 等 RCS 接口到位后，取消时会主动通知 RCS 停止所有进行中的任务
    # await abort_inventory_task(robot_task_code, is_sim=False)
    if robot_task_code:
        logger.info(f"[cancel] 等待 RCS cancel API 就位，robotTaskCode={robot_task_code}")

    return JSONResponse(
        status_code=status.HTTP_200_OK,
        content={"code": 200, "message": "任务已取消", "data": None}
    )


# ==================== Gateway 启动时处理中断任务 ====================

@router.on_event("startup")
async def startup_event():
    on_server_startup()
