"""
历史记录路由
"""
import re
import json
import logging
import shutil
from pathlib import Path
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional, Set
from fastapi import APIRouter, Request, HTTPException, status
from fastapi.responses import JSONResponse, Response

from services.api.shared.config import logger, project_root
from services.api.shared.operation_log import log_operation

router = APIRouter(prefix="/api/history", tags=["history"])

# 历史数据输出根目录
OUTPUT_ROOT = project_root / "output"
# 任务索引缓存文件
_TASK_INDEX_FILE = OUTPUT_ROOT / "task_index.json"
# 内存中的任务索引缓存（首次请求时加载，之后复用）
_TASK_INDEX_CACHE: Dict[str, Dict] = {}
_INDEX_LOADED = False


def _load_task_index() -> Dict[str, Dict]:
    """从磁盘加载任务索引缓存"""
    if _TASK_INDEX_FILE.exists():
        try:
            with open(_TASK_INDEX_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _save_task_index(index: Dict[str, Dict]):
    """保存任务索引缓存到磁盘"""
    OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
    with open(_TASK_INDEX_FILE, "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)


def _ensure_task_index():
    """确保索引已加载，构建或从磁盘读取"""
    global _INDEX_LOADED, _TASK_INDEX_CACHE
    if _INDEX_LOADED:
        return
    _TASK_INDEX_CACHE = _load_task_index()
    _INDEX_LOADED = True


def invalidate_task_index():
    """使索引缓存失效，下次请求时重建（任务新增/修改时调用）"""
    global _INDEX_LOADED, _TASK_INDEX_CACHE
    _INDEX_LOADED = False
    _TASK_INDEX_CACHE = {}


def _rebuild_task_index(history_dir: Path) -> Dict[str, Dict]:
    """扫描所有 Excel 文件，重建索引缓存"""
    import openpyxl
    index = {}
    for xlsx_file in history_dir.glob("*.xlsx"):
        try:
            task_id = xlsx_file.stem
            dispatch_time = None
            task_operator = ""
            task_has_modified = False
            task_is_valid = True

            wb = openpyxl.load_workbook(str(xlsx_file), data_only=True)
            ws = wb.active
            headers = [str(cell.value) for cell in ws[1] if cell.value]

            # 解析下发时间
            if "下发时间" in headers:
                col = headers.index("下发时间") + 1
                val = ws.cell(row=2, column=col).value
                if val:
                    dispatch_time = val.isoformat() if isinstance(val, datetime) else None
            if dispatch_time is None:
                dt = parse_task_date_from_filename(task_id)
                if dt:
                    dispatch_time = dt.isoformat()

            # 解析操作员
            if "操作员" in headers and ws.max_row >= 2:
                val = ws.cell(row=2, column=headers.index("操作员") + 1).value
                if val:
                    task_operator = str(val)

            # 解析修改记录
            if "修改记录" in headers and ws.max_row >= 2:
                col = headers.index("修改记录") + 1
                for row in range(2, ws.max_row + 1):
                    val = str(ws.cell(row=row, column=col).value or "")
                    if val and val != "None":
                        task_has_modified = True
                        break

            # 解析有效状态
            if "有效状态" in headers and ws.max_row >= 2:
                val = str(ws.cell(row=2, column=headers.index("有效状态") + 1).value or "").strip()
                task_is_valid = (val == "有效")

            wb.close()
            index[task_id] = {
                "taskId": task_id,
                "taskDate": dispatch_time,
                "fileName": xlsx_file.name,
                "filePath": str(xlsx_file.relative_to(OUTPUT_ROOT)),
                "operator": task_operator,
                "hasModified": task_has_modified,
                "isValid": task_is_valid,
            }
        except Exception as ex:
            logger.warning(f"索引重建失败 {xlsx_file.name}: {ex}")
            continue
    return index


def _get_task_meta(task_id: str) -> Optional[Dict]:
    """获取单个任务的索引元数据（自动懒加载索引）"""
    _ensure_task_index()
    if task_id in _TASK_INDEX_CACHE:
        return _TASK_INDEX_CACHE[task_id]

    # 不在缓存中，扫描目录增量补充
    history_dir = OUTPUT_ROOT / "history_data"
    if not history_dir.exists():
        return None
    xlsx_file = history_dir / f"{task_id}.xlsx"
    if not xlsx_file.exists():
        return None

    # 增量读取该文件
    import openpyxl
    dispatch_time = None
    task_operator = ""
    task_has_modified = False
    task_is_valid = True
    try:
        wb = openpyxl.load_workbook(str(xlsx_file), data_only=True)
        ws = wb.active
        headers = [str(cell.value) for cell in ws[1] if cell.value]
        if "下发时间" in headers:
            col = headers.index("下发时间") + 1
            val = ws.cell(row=2, column=col).value
            if val:
                dispatch_time = val.isoformat() if isinstance(val, datetime) else None
        if dispatch_time is None:
            dt = parse_task_date_from_filename(task_id)
            if dt:
                dispatch_time = dt.isoformat()
        if "操作员" in headers and ws.max_row >= 2:
            val = ws.cell(row=2, column=headers.index("操作员") + 1).value
            if val:
                task_operator = str(val)
        if "修改记录" in headers and ws.max_row >= 2:
            col = headers.index("修改记录") + 1
            for row in range(2, ws.max_row + 1):
                val = str(ws.cell(row=row, column=col).value or "")
                if val and val != "None":
                    task_has_modified = True
                    break
        if "有效状态" in headers and ws.max_row >= 2:
            val = str(ws.cell(row=2, column=headers.index("有效状态") + 1).value or "").strip()
            task_is_valid = (val == "有效")
        wb.close()
    except Exception:
        pass

    meta = {
        "taskId": task_id,
        "taskDate": dispatch_time,
        "fileName": xlsx_file.name,
        "filePath": str(xlsx_file.relative_to(OUTPUT_ROOT)),
        "operator": task_operator,
        "hasModified": task_has_modified,
        "isValid": task_is_valid,
    }
    _TASK_INDEX_CACHE[task_id] = meta
    return meta


def parse_task_date_from_filename(task_id: str) -> Optional[datetime]:
    """从任务ID解析日期"""
    patterns = [
        (r'(\d{4})(\d{2})(\d{2})', '%Y%m%d'),
        (r'(\d{4})-(\d{2})-(\d{2})', '%Y-%m-%d'),
        (r'(\d{2})(\d{2})(\d{4})', '%d%m%Y'),
    ]

    for pattern, date_format in patterns:
        match = re.search(pattern, task_id)
        if match:
            try:
                date_str = ''.join(match.groups())
                return datetime.strptime(date_str, '%Y%m%d')
            except ValueError:
                continue
    return None


def is_task_expired(task_date: datetime) -> bool:
    """检查任务是否过期（超过6个月）"""
    if not task_date:
        return False
    now = datetime.now()
    six_months_ago = now - timedelta(days=180)
    return task_date < six_months_ago


@router.get("/available-dates")
async def get_available_dates():
    """获取所有有盘点数据的日期列表"""
    try:
        _ensure_task_index()
        history_dir = OUTPUT_ROOT / "history_data"
        history_dir.mkdir(parents=True, exist_ok=True)

        # 先从缓存获取已有日期
        dates: Set[str] = set()
        for meta in _TASK_INDEX_CACHE.values():
            td = meta.get("taskDate")
            if td:
                try:
                    dates.add(datetime.fromisoformat(td).strftime("%Y-%m-%d"))
                except Exception:
                    pass

        # 缓存不够？扫描目录增量补充
        xlsx_files = list(history_dir.glob("*.xlsx"))
        if len(xlsx_files) > len(_TASK_INDEX_CACHE):
            for xlsx_file in xlsx_files:
                task_id = xlsx_file.stem
                if task_id not in _TASK_INDEX_CACHE:
                    _get_task_meta(task_id)

        # 再次从缓存收集
        dates = set()
        for meta in _TASK_INDEX_CACHE.values():
            td = meta.get("taskDate")
            if td:
                try:
                    dates.add(datetime.fromisoformat(td).strftime("%Y-%m-%d"))
                except Exception:
                    pass

        sorted_dates = sorted(dates, reverse=True)
        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": "获取可用日期成功", "data": {"dates": sorted_dates}}
        )

    except Exception as e:
        logger.error(f"获取可用日期失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"获取可用日期失败: {str(e)}")


@router.get("/tasks")
async def get_history_tasks(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """获取历史任务列表（使用索引缓存加速）"""
    try:
        history_dir = OUTPUT_ROOT / "history_data"
        history_dir.mkdir(parents=True, exist_ok=True)

        # 首次请求：构建索引缓存
        if not _INDEX_LOADED:
            _ensure_task_index()
            # 如果缓存为空则重建
            if not _TASK_INDEX_CACHE:
                _TASK_INDEX_CACHE.update(_rebuild_task_index(history_dir))
                _save_task_index(_TASK_INDEX_CACHE)

        # 增量补充：目录中有但缓存中没有的文件
        for xlsx_file in history_dir.glob("*.xlsx"):
            if xlsx_file.stem not in _TASK_INDEX_CACHE:
                _get_task_meta(xlsx_file.stem)

        tasks = []
        for meta in _TASK_INDEX_CACHE.values():
            td = meta.get("taskDate")
            dispatch_time = None
            if td:
                try:
                    dispatch_time = datetime.fromisoformat(td)
                except Exception:
                    pass
            is_expired = is_task_expired(dispatch_time) if dispatch_time else False

            tasks.append({
                "taskId": meta["taskId"],
                "taskDate": td,
                "fileName": meta.get("fileName", meta["taskId"] + ".xlsx"),
                "isExpired": is_expired,
                "filePath": meta.get("filePath", ""),
                "operator": meta.get("operator", ""),
                "hasModified": meta.get("hasModified", False),
                "isValid": meta.get("isValid", True),
            })

        # 按下发时间范围过滤
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, "%Y-%m-%d")
                tasks = [t for t in tasks if t.get("taskDate") and datetime.fromisoformat(t["taskDate"]) >= start_dt]
            except ValueError:
                pass

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
                tasks = [t for t in tasks if t.get("taskDate") and datetime.fromisoformat(t["taskDate"]) <= end_dt]
            except ValueError:
                pass

        tasks.sort(key=lambda x: x.get("taskDate") or "", reverse=True)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": "获取历史任务成功", "data": {"tasks": tasks, "total": len(tasks)}}
        )

    except Exception as e:
        logger.error(f"获取历史任务列表失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"获取历史任务列表失败: {str(e)}")


@router.get("/task/{task_id}")
async def get_history_task_details(task_id: str):
    """获取历史任务详情"""
    try:
        history_tasks_dir = OUTPUT_ROOT / "history_data"
        xlsx_file = history_tasks_dir / f"{task_id}.xlsx"

        if not xlsx_file.exists():
            possible_files = list(history_tasks_dir.glob(f"{task_id}.*"))
            if possible_files:
                xlsx_file = possible_files[0]
            else:
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"历史任务文件不存在: {task_id}.xlsx")

        logger.info(f"读取Excel文件: {xlsx_file}")

        import openpyxl
        workbook = openpyxl.load_workbook(str(xlsx_file), data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]

        headers = [str(cell.value) for cell in worksheet[1] if cell.value]
        logger.info(f"Excel表头: {headers}")

        # 预先从第一行数据读取操作员和有效状态（直接按列索引读取，不走 row_dict）
        task_is_valid = True  # 默认为有效
        if worksheet.max_row >= 2:
            if "操作员" in headers:
                op_val = worksheet.cell(row=2, column=headers.index("操作员") + 1).value
                if op_val:
                    operator = str(op_val)
            if "有效状态" in headers:
                valid_col_idx = headers.index("有效状态") + 1
                valid_str = str(worksheet.cell(row=2, column=valid_col_idx).value or "").strip()
                task_is_valid = (valid_str == "有效")
                logger.info(f"详情接口读取 isValid: {task_is_valid}, valid_str: '{valid_str}'")

        details = []
        has_modified = False
        modified_bins: List[str] = []
        for row_index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 1):
            if all(cell is None for cell in row):
                continue

            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            mod_record = str(row_dict.get("修改记录", "") or "")
            # 排除空字符串和字符串 "None"（旧版 Excel 遗留值）
            if mod_record and mod_record != "None":
                has_modified = True
                bin_name = str(row_dict.get("储位名称", ""))
                if bin_name not in modified_bins:
                    modified_bins.append(bin_name)
            detail = {
                "序号": row_dict.get("序号") or row_index,
                "品规名称": str(row_dict.get("品规名称", "")),
                "储位名称": str(row_dict.get("储位名称", "")),
                "实际品规": str(row_dict.get("实际品规", row_dict.get("品规名称", ""))),
                "库存数量": int(row_dict.get("库存数量", 0) or 0),
                "实际数量": int(row_dict.get("实际数量", 0) or 0),
                "差异": str(row_dict.get("差异", "一致")),
                "修改记录": mod_record,
                "照片1路径": str(row_dict.get("照片1路径") or ""),
                "照片2路径": str(row_dict.get("照片2路径") or ""),
                "照片3路径": str(row_dict.get("照片3路径") or ""),
                "照片4路径": str(row_dict.get("照片4路径") or ""),
            }
            details.append(detail)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": "获取历史任务详情成功",
                     "data": {"taskId": task_id, "details": details, "total": len(details),
                              "operator": operator, "hasModified": has_modified, "modifiedBins": modified_bins,
                              "isValid": task_is_valid}}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取历史任务详情失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"获取历史任务详情失败: {str(e)}")


@router.delete("/task/{task_id}")
async def delete_history_task(task_id: str, request: Request):
    """按任务ID删除历史盘点文件"""
    try:
        user_level = request.headers.get("X-User-Level")
        if user_level != "admin":
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="权限不足，仅管理员可删除历史数据")

        if not re.match(r'^[A-Za-z0-9_-]+$', task_id):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="任务ID包含非法字符")

        history_tasks_dir = OUTPUT_ROOT / "history_data"
        xlsx_file = history_tasks_dir / f"{task_id}.xlsx"

        if not xlsx_file.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"任务 {task_id} 的历史文件不存在")

        xlsx_file.unlink()
        invalidate_task_index()  # 刷新缓存
        logger.info(f"已删除历史任务文件: {xlsx_file.name}")

        client_host = request.client.host if request.client else "unknown"
        log_operation(
            operation_type="system_cleanup",
            action="删除历史任务",
            user_name="前端传递",
            target=task_id,
            status="success",
            ip_address=client_host,
            details={"task_id": task_id, "filename": xlsx_file.name}
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": f"任务 {task_id} 已成功删除", "data": {"task_id": task_id}}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除历史任务失败 {task_id}: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"删除历史任务失败: {str(e)}")


@router.get("/task/{task_id}/bin/{bin_location}")
async def get_history_task_bin_detail(task_id: str, bin_location: str):
    """获取历史任务中某个储位的详细信息"""
    try:
        history_tasks_dir = OUTPUT_ROOT / "history_data"
        xlsx_file = history_tasks_dir / f"{task_id}.xlsx"

        if not xlsx_file.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"历史任务文件不存在: {task_id}.xlsx")

        import openpyxl
        workbook = openpyxl.load_workbook(str(xlsx_file), data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]

        headers = [str(cell.value) for cell in worksheet[1] if cell.value]

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
            if row_dict.get("储位名称") == bin_location:
                return JSONResponse(
                    status_code=status.HTTP_200_OK,
                    content={"code": 200, "message": "获取储位详情成功", "data": row_dict}
                )

        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"储位 {bin_location} 不存在")

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取储位详情失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"获取储位详情失败: {str(e)}")


@router.get("/image")
async def get_history_image(taskNo: str, binLocation: str, cameraType: str, filename: str, source: str = "output"):
    """获取历史图片"""
    try:
        # 根据 source 参数确定根目录
        if source == "capture_img":
            base_dir = project_root / "capture_img"
        else:
            base_dir = OUTPUT_ROOT

        image_dir = base_dir / taskNo / binLocation / cameraType

        # 尝试查找匹配的图片文件（支持带扩展名和不带扩展名的 filename）
        image_path = None
        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.JPG', '.JPEG', '.PNG', '.BMP']

        # 先尝试直接使用 filename 作为完整路径
        direct_path = image_dir / filename
        if direct_path.exists():
            image_path = direct_path
        else:
            # 尝试添加扩展名
            for ext in image_extensions:
                test_path = image_dir / f"{filename}{ext}"
                if test_path.exists():
                    image_path = test_path
                    break

        if not image_path or not image_path.exists():
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=f"图片不存在: {filename}")

        with open(image_path, "rb") as f:
            image_data = f.read()

        media_type = "image/jpeg"
        filename_lower = image_path.name.lower()
        if filename_lower.endswith(".png"):
            media_type = "image/png"
        elif filename_lower.endswith(".bmp"):
            media_type = "image/bmp"

        return Response(content=image_data, media_type=media_type)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取图片失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"获取图片失败: {str(e)}")


@router.post("/cleanup")
async def cleanup_history(request: Request):
    """清理过期的历史数据"""
    try:
        user_level = request.headers.get("X-User-Level")
        if user_level != "admin":
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="权限不足")

        data = await request.json()
        days = data.get("days", 180)

        history_tasks_dir = OUTPUT_ROOT / "history_data"
        deleted_count = 0

        for xlsx_file in history_tasks_dir.glob("*.xlsx"):
            task_id = xlsx_file.stem
            task_date = parse_task_date_from_filename(task_id)
            if task_date and is_task_expired(task_date):
                xlsx_file.unlink()
                deleted_count += 1
                logger.info(f"已清理过期文件: {xlsx_file.name}")

        if deleted_count > 0:
            invalidate_task_index()  # 刷新缓存

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": f"已清理 {deleted_count} 个过期文件", "data": {"deletedCount": deleted_count}}
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"清理历史数据失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"清理历史数据失败: {str(e)}")


@router.get("/monthly-count")
async def get_monthly_count():
    """获取本月盘点次数和准确率"""
    try:
        import openpyxl

        # 历史数据目录
        history_tasks_dir = OUTPUT_ROOT / "history_data"

        if not history_tasks_dir.exists():
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "历史数据目录不存在",
                    "data": {
                        "count": 0,
                        "current_month": datetime.now().strftime("%Y-%m"),
                        "accuracy": None,
                        "files": []
                    }
                }
            )

        # 获取当前年月
        current_year = datetime.now().year
        current_month = datetime.now().month
        current_month_str = f"{current_year}-{current_month:02d}"

        # 查找所有历史数据文件
        all_files = list(history_tasks_dir.glob("*.xlsx"))
        all_files.extend(history_tasks_dir.glob("*.xls"))

        monthly_files = []

        for file_path in all_files:
            try:
                # 从文件名解析日期
                filename = file_path.stem
                if len(filename) >= 10 and filename[:2] == "HS":
                    date_str = filename[2:10]
                    try:
                        file_date = datetime.strptime(date_str, "%Y%m%d")
                        file_year = file_date.year
                        file_month = file_date.month

                        if file_year == current_year and file_month == current_month:
                            monthly_files.append({
                                "filename": file_path.name,
                                "filepath": str(file_path),
                                "date": file_date.strftime("%Y-%m-%d"),
                                "task_no": filename
                            })
                    except ValueError:
                        continue
            except Exception as e:
                logger.error(f"处理文件失败 {file_path.name}: {str(e)}")
                continue

        count = len(monthly_files)

        # 计算准确率
        accuracy = None
        if count > 0:
            total_count = 0
            match_count = 0

            for file_info in monthly_files:
                try:
                    wb = openpyxl.load_workbook(file_info["filepath"], read_only=True)
                    ws = wb.active

                    for row_idx in range(2, ws.max_row + 1):
                        cell_h = ws.cell(row=row_idx, column=8)  # 第H列（差异）
                        if cell_h.value is not None:
                            total_count += 1
                            if str(cell_h.value).strip() == "一致":
                                match_count += 1

                    wb.close()
                except Exception as e:
                    logger.error(f"读取Excel失败 {file_info['filename']}: {str(e)}")
                    continue

            if total_count > 0:
                accuracy = round((match_count / total_count) * 100, 1)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取月度统计成功",
                "data": {
                    "count": count,
                    "current_month": current_month_str,
                    "accuracy": accuracy,
                    "files": [f["filename"] for f in monthly_files]
                }
            }
        )

    except Exception as e:
        logger.error(f"获取月度统计失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=f"获取月度统计失败: {str(e)}")
