"""
Gateway服务 - 向后兼容入口
"""
import warnings
import asyncio
import os
import time
import json
import logging
import glob
import uuid
import shutil
import re
import pandas as pd
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta

from fastapi import FastAPI, Request, BackgroundTasks, HTTPException, status, Body
from fastapi.responses import JSONResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import requests
import uvicorn
import sys

import custom_utils

# 配置日志
# 创建 debug 日志目录（在文件开头就定义，供后续使用）
_project_root = Path(__file__).parent.parent.parent
_debug_log_dir = _project_root / "debug"
_debug_log_dir.mkdir(parents=True, exist_ok=True)

# 创建日志文件路径（按日期命名）
_log_filename = _debug_log_dir / \
    f"gateway_{datetime.now().strftime('%Y%m%d')}.log"

# 配置根日志记录器
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    handlers=[
        logging.StreamHandler(),  # 控制台输出
        logging.FileHandler(str(_log_filename), encoding='utf-8')  # 文件输出
    ]
)
logger = logging.getLogger(__name__)
logger.info(f"日志文件保存路径: {_log_filename}")

# 导出供其他函数使用
debug_log_dir = _debug_log_dir

# 操作记录存储根目录
OPERATION_LOGS_DIR = debug_log_dir / "operation_logs"
OPERATION_LOGS_DIR.mkdir(parents=True, exist_ok=True)

# 创建各类型子目录
LOG_TYPES = ["inventory", "user_login",
             "user_management", "system_cleanup", "other"]
for log_type in LOG_TYPES:
    (OPERATION_LOGS_DIR / log_type).mkdir(parents=True, exist_ok=True)


class OperationLog(BaseModel):
    """操作记录模型"""
    id: str  # 格式: YYYYMMDD_HHMMSS_xxx
    timestamp: str  # ISO格式时间戳
    operation_type: str  # 操作类型
    user_id: Optional[str] = None  # 操作用户ID
    user_name: Optional[str] = None  # 操作用户名
    action: str  # 具体动作
    target: Optional[str] = None  # 操作目标（如任务ID、用户名）
    status: str  # 成功/失败/进行中
    details: Dict[str, Any] = {}  # 详细数据
    ip_address: Optional[str] = None  # 操作IP
    metadata: Dict[str, Any] = {}  # 元数据


def generate_operation_id() -> str:
    """生成操作记录ID"""
    return datetime.now().strftime("%Y%m%d_%H%M%S_") + str(int(time.time() * 1000))[-3:]


def save_operation_log(operation_log: OperationLog):
    """保存操作记录到文件"""
    try:
        # 根据类型选择目录
        log_type = operation_log.operation_type
        if log_type not in LOG_TYPES:
            log_type = "other"

        log_dir = OPERATION_LOGS_DIR / log_type

        # 创建文件名
        filename = f"{operation_log.id}.json"
        filepath = log_dir / filename

        # 保存为JSON
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(operation_log.dict(), f, ensure_ascii=False, indent=2)

        logger.info(f"操作记录已保存: {filepath}")
        return True
    except Exception as e:
        logger.error(f"保存操作记录失败: {str(e)}")
        return False


def log_operation(
    operation_type: str,
    action: str,
    user_id: Optional[str] = None,
    user_name: Optional[str] = None,
    target: Optional[str] = None,
    status: str = "success",
    details: Optional[Dict[str, Any]] = None,
    ip_address: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None
):
    """记录操作的辅助函数"""
    operation_log = OperationLog(
        id=generate_operation_id(),
        timestamp=datetime.now().isoformat(),
        operation_type=operation_type,
        user_id=user_id,
        user_name=user_name,
        action=action,
        target=target,
        status=status,
        details=details or {},
        ip_address=ip_address,
        metadata=metadata or {}
    )

    return save_operation_log(operation_log)


def get_recent_operations(limit: int = 5, days: int = 180) -> List[Dict[str, Any]]:
    """
    获取最近的操作记录

    Args:
        limit: 返回数量限制
        days: 查询天数（默认180天，约6个月）
    """
    try:
        all_operations = []
        cutoff_date = datetime.now() - timedelta(days=days)

        # 遍历所有类型目录
        for log_type in LOG_TYPES:
            log_dir = OPERATION_LOGS_DIR / log_type
            if not log_dir.exists():
                continue

            # 查找所有JSON文件
            json_files = list(log_dir.glob("*.json"))

            for json_file in json_files:
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)

                    # 检查时间是否在有效期内
                    log_time_str = data.get("timestamp", "")
                    if log_time_str:
                        try:
                            log_time = datetime.fromisoformat(
                                log_time_str.replace('Z', '+00:00'))
                            if log_time < cutoff_date:
                                continue  # 跳过过期记录
                        except:
                            pass  # 如果时间解析失败，仍然包含

                    all_operations.append(data)
                except Exception as e:
                    logger.warning(f"读取操作记录文件失败 {json_file}: {str(e)}")
                    continue

        # 按时间倒序排序
        all_operations.sort(key=lambda x: x.get("timestamp", ""), reverse=True)

        # 限制返回数量
        return all_operations[:limit]

    except Exception as e:
        logger.error(f"获取操作记录失败: {str(e)}")
        return []


def get_all_operations(days: int = 180) -> List[Dict[str, Any]]:
    """获取指定天数内的所有操作记录"""
    try:
        all_operations = []
        cutoff_date = datetime.now() - timedelta(days=days)

        # 遍历所有类型目录
        for log_type in LOG_TYPES:
            log_dir = OPERATION_LOGS_DIR / log_type
            if not log_dir.exists():
                continue

            # 查找所有JSON文件
            json_files = list(log_dir.glob("*.json"))

            for json_file in json_files:
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)

                    # 检查时间是否在有效期内
                    log_time_str = data.get("timestamp", "")
                    if log_time_str:
                        try:
                            log_time = datetime.fromisoformat(
                                log_time_str.replace('Z', '+00:00'))
                            if log_time < cutoff_date:
                                continue  # 跳过过期记录
                        except:
                            pass  # 如果时间解析失败，仍然包含

                    all_operations.append(data)
                except Exception as e:
                    logger.warning(f"读取操作记录文件失败 {json_file}: {str(e)}")
                    continue

        # 按时间倒序排序
        all_operations.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
        return all_operations

    except Exception as e:
        logger.error(f"获取所有操作记录失败: {str(e)}")
        return []


# 数据模型定义
class TaskStatus(BaseModel):
    """任务状态模型"""
    task_no: str
    status: str  # init, running, completed, failed
    current_step: int = 0
    total_steps: int = 0
    start_time: Optional[str] = None
    end_time: Optional[str] = None


class BinLocationStatus(BaseModel):
    """储位状态模型"""
    bin_location: str
    status: str  # pending, running, completed, failed
    sequence: int = 0
    image_data: Optional[Dict[str, Any]] = None
    compute_result: Optional[Dict[str, Any]] = None
    capture_time: Optional[str] = None
    compute_time: Optional[str] = None
    detect_result: Optional[Dict[str, Any]] = None  # Detect模块识别结果
    barcode_result: Optional[Dict[str, Any]] = None  # Barcode模块识别结果
    recognition_time: Optional[str] = None  # 识别时间


class InventoryTaskProgress(BaseModel):
    """盘点任务进度模型"""
    task_no: str
    status: str
    current_step: int
    total_steps: int
    progress_percentage: float
    bin_locations: List[BinLocationStatus]
    start_time: Optional[str] = None
    end_time: Optional[str] = None


# 请求模型
class ScanAndRecognizeRequest(BaseModel):
    """扫码+识别请求模型"""
    taskNo: str  # 任务编号
    binLocation: str  # 库位号
    pile_id: int = 1  # 堆垛ID，默认为1
    code_type: str = "ucc128"  # 条码类型，默认ucc128


class FrontendLogRequest(BaseModel):
    """前端日志请求模型"""
    level: str  # log, info, warn, error
    message: str
    timestamp: Optional[str] = None
    source: Optional[str] = None  # 前端来源标识
    extra: Optional[Dict[str, Any]] = None  # 额外信息


# 创建 FastAPI 应用实例
app = FastAPI(
    title="LeafDepot API Gateway",
    description="LeafDepot 系统 API 网关服务",
    version="1.0.0"
)

# 配置 CORS
# 从环境变量读取允许的源，如果没有设置则使用默认列表
cors_origins_env = os.getenv("CORS_ORIGINS", "")
if cors_origins_env:
    # 如果设置了环境变量，使用环境变量的值（逗号分隔）
    origins = [origin.strip() for origin in cors_origins_env.split(",")]
else:
    # 默认允许的源列表
    origins = [
        "http://localhost",
        "http://localhost:8000",
        "http://localhost:8001",
        "http://localhost:8080",  # 测试页面服务器端口
        "http://localhost:3000",
        "http://localhost:5000",
        "http://127.0.0.1:8000",
        "http://127.0.0.1:8001",
        "http://127.0.0.1:8080",  # 测试页面服务器端口
        "http://127.0.0.1:3000",
        "http://127.0.0.1:5000",
    ]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],  # 暴露所有头部
    max_age=3600,  # 预检请求缓存时间
)

# Barcode功能开关（默认开启）
ENABLE_BARCODE = os.getenv(
    "ENABLE_BARCODE", "true").lower() in ("true", "1", "yes")

# 导入条形码路由（如果存在且开关开启）
if ENABLE_BARCODE:
    try:
        from services.api.routers.barcode import router as barcode_router
        app.include_router(barcode_router)
        logger.info("条形码路由已启用")
    except ImportError:
        logger.warning("条形码路由模块未找到，跳过注册")
else:
    logger.info("条形码功能已禁用（ENABLE_BARCODE=false）")

# 导入工具模块

# 导入检测模块
try:
    from core.detection import count_boxes
    DETECT_MODULE_AVAILABLE = True
except ImportError as e:
    logger.warning(f"检测模块导入失败: {e}")
    DETECT_MODULE_AVAILABLE = False

# 导入条码识别模块（仅在开关开启时导入）
BARCODE_MODULE_AVAILABLE = False
if ENABLE_BARCODE:
    try:
        from core.vision.barcode_recognizer import BarcodeRecognizer
        BARCODE_MODULE_AVAILABLE = True
        logger.info("条形码识别模块已启用")
    except ImportError as e:
        logger.warning(f"条形码识别模块导入失败: {e}")
        BARCODE_MODULE_AVAILABLE = False
else:
    logger.info("条形码识别模块已禁用（ENABLE_BARCODE=false）")


# ============================================================
# 烟草烟箱信息查询模块
# ============================================================

# 垛型字符串 → 垛型编码 映射
STACK_TYPE_TO_CODE = {
    "5*8": 0,
    "5*6": 1,
    "5*5": 2,
    "3*10": 3,
    "4*7+2": 4,
}

# 垛型编码 → pile_id 映射（对应 pile_config.json）
STACK_TYPE_CODE_TO_PILE_ID = {
    0: 2,   # "5*8"   → pile_id=2
    1: 3,   # "5*6"   → pile_id=3
    2: 5,   # "5*5"   → pile_id=5
    3: 1,   # "3*10"  → pile_id=1
    4: 4,   # "4*7+2" → pile_id=4
    -1: 1,  # unknown → default
}


class TobaccoCaseInfoResolver:
    """
    烟草烟箱信息解析器

    功能：根据条码识别结果，查询烟箱信息并返回垛型编码和映射后的 pile_id
    """

    _instance = None
    _code_mapping = None

    def __new__(cls):
        """单例模式，避免重复加载Excel数据"""
        if cls._instance is None:
            cls._instance = super().__new__(cls)
        return cls._instance

    def __init__(self):
        """初始化：加载Excel数据"""
        if self._code_mapping is None:
            self._load_excel_data()

    def _load_excel_data(self):
        """加载烟箱信息Excel数据"""
        try:
            excel_path = _project_root / "shared" / "data" / "烟箱信息汇总完整版.xlsx"

            if not excel_path.exists():
                logger.warning(f"烟箱信息Excel文件不存在: {excel_path}")
                self._code_mapping = {}
                return

            df = pd.read_excel(excel_path, dtype=str)
            self._code_mapping = {}

            for _, row in df.iterrows():
                # 获取6位数字代码
                code = str(row.get('提取的6位数字', '')).strip()
                if code and code not in ['nan', '']:
                    self._code_mapping[code] = {
                        'product_name': str(row.get('品名', '')),
                        'tobacco_code': str(row.get('烟草内部品规代号', '')),
                        'stack_type_1': str(row.get('垛型_1', '')),
                        'stack_type_2': str(row.get('垛型_2', '')),
                    }

            logger.info(f"成功加载 {len(self._code_mapping)} 条烟箱信息")

        except Exception as e:
            logger.error(f"加载烟箱信息Excel失败: {e}")
            self._code_mapping = {}

    def _extract_six_digits(self, barcode: str) -> Optional[str]:
        """从条码中提取6位数字（忽略91前缀）"""
        if not barcode:
            return None

        # 移除 (91) 或 91 前缀
        if barcode.startswith('(91)'):
            barcode = barcode[4:]
        elif barcode.startswith('91'):
            barcode = barcode[2:]

        # 提取前6位连续数字
        match = re.search(r'(\d{6})', barcode)
        return match.group(1) if match else None

    def _parse_stack_type(self, stack_type_str: str) -> int:
        """将垛型字符串转换为编码"""
        if not stack_type_str or str(stack_type_str).strip().lower() in ['nan', 'none', '']:
            return -1
        return STACK_TYPE_TO_CODE.get(str(stack_type_str).strip(), -1)

    def resolve(self, barcode: str) -> Dict[str, Any]:
        """
        根据条码解析烟箱信息

        Args:
            barcode: 识别到的条码字符串

        Returns:
            {
                'success': bool,
                'six_digit_code': str,
                'stack_type_1': int,  # 垛型编码
                'pile_id': int,       # 映射后的 pile_id
                'product_name': str,
                'tobacco_code': str,
            }
        """
        result = {
            'success': False,
            'six_digit_code': None,
            'stack_type_1': -1,
            'pile_id': None,
            'product_name': '',
            'tobacco_code': '',
        }

        if not barcode:
            return result

        # 提取6位数字
        six_digits = self._extract_six_digits(barcode)
        if not six_digits:
            logger.warning(f"无法从条码提取6位数字: {barcode}")
            return result

        result['six_digit_code'] = six_digits

        # 查找匹配的烟箱信息
        match_data = self._code_mapping.get(six_digits)
        if not match_data:
            # 尝试模糊匹配
            for code in self._code_mapping:
                if six_digits in code or code in six_digits:
                    match_data = self._code_mapping[code]
                    logger.info(f"使用模糊匹配: {six_digits} -> {code}")
                    break

        if not match_data:
            logger.warning(f"未找到匹配的烟箱信息: {six_digits}")
            return result

        # 解析垛型
        stack_type_code = self._parse_stack_type(match_data['stack_type_1'])

        # 映射 pile_id
        pile_id = STACK_TYPE_CODE_TO_PILE_ID.get(stack_type_code, 1)

        result.update({
            'success': True,
            'stack_type_1': stack_type_code,
            'pile_id': pile_id,
            'product_name': match_data['product_name'],
            'tobacco_code': match_data['tobacco_code'],
        })

        logger.info(f"烟箱信息解析成功: code={six_digits}, stack_type={stack_type_code}, pile_id={pile_id}")
        return result


def get_tobacco_case_resolver() -> TobaccoCaseInfoResolver:
    """获取烟箱信息解析器单例"""
    return TobaccoCaseInfoResolver()


# 配置常量（这些应该从配置文件或环境变量中读取）
# 默认使用模拟服务端口（6000），如果环境变量设置了则使用环境变量的值
LMS_BASE_URL = os.getenv("LMS_BASE_URL", "http://localhost:6000")
RCS_BASE_URL = os.getenv("RCS_BASE_URL", "http://localhost:4001")
RCS_PREFIX = os.getenv("RCS_PREFIX", "")

# 抓图脚本路径（应该从配置文件读取）
CAPTURE_SCRIPTS = [
    os.path.join(os.path.dirname(__file__), "..", "..",
                 "hardware", "cam_sys", "scan_1_capture.py"),
    os.path.join(os.path.dirname(__file__), "..", "..",
                 "hardware", "cam_sys", "scan_2_capture.py"),
]

# 延迟导入服务（避免循环导入）


def get_inventory_service():
    """获取盘点服务实例"""
    from services.vision.box_count_service import get_box_count_service
    return get_box_count_service()


# 任务状态存储（如果任务状态管理器不存在，使用简单的内存存储）
_inventory_tasks: Dict[str, TaskStatus] = {}
_inventory_task_bins: Dict[str, List[BinLocationStatus]] = {}
_inventory_task_details: Dict[str, Dict[str, Dict[str, Any]]] = {}

# 获取任务状态存储的辅助函数


def get_task_state_storage():
    """获取任务状态存储（优先使用任务状态管理器，否则使用简单内存存储）"""
    try:
        from services.api.state.task_state import get_task_state_manager
        manager = get_task_state_manager()
        return {
            "tasks": manager._inventory_tasks,
            "bins": manager._inventory_task_bins,
            "details": manager._inventory_task_details
        }
    except ImportError:
        # 如果任务状态管理器不存在，使用简单的内存存储
        return {
            "tasks": _inventory_tasks,
            "bins": _inventory_task_bins,
            "details": _inventory_task_details
        }


# 初始化模块级别的变量（供函数内部直接使用）
# 这些变量指向实际存储（通过 get_task_state_storage() 获取）
_storage = get_task_state_storage()
inventory_tasks = _storage["tasks"]
inventory_task_bins = _storage["bins"]
inventory_task_details = _storage["details"]

# 为了向后兼容，保留一些全局变量引用（通过延迟导入避免循环）


def __getattr__(name):
    """延迟导入以支持向后兼容"""
    if name == "inventory_tasks":
        try:
            from services.api.state.task_state import get_task_state_manager
            return get_task_state_manager()._inventory_tasks
        except ImportError:
            logger.warning("任务状态管理器模块不存在，使用简单内存存储")
            return _inventory_tasks
    elif name == "inventory_task_bins":
        try:
            from services.api.state.task_state import get_task_state_manager
            return get_task_state_manager()._inventory_task_bins
        except ImportError:
            logger.warning("任务状态管理器模块不存在，使用简单内存存储")
            return _inventory_task_bins
    elif name == "inventory_task_details":
        try:
            from services.api.state.task_state import get_task_state_manager
            return get_task_state_manager()._inventory_task_details
        except ImportError:
            logger.warning("任务状态管理器模块不存在，使用简单内存存储")
            return _inventory_task_details
    elif name in ["STATUS_KEY", "status_event", "robot_status_store"]:
        try:
            from services.api.state.robot_state import get_robot_state_manager
            manager = get_robot_state_manager()
            if name == "STATUS_KEY":
                from services.api.config import ROBOT_STATUS_KEY
                return ROBOT_STATUS_KEY
            elif name == "status_event":
                return manager._status_event
            elif name == "robot_status_store":
                return manager._robot_status_store
        except ImportError:
            logger.warning(f"机器人状态管理器模块不存在，无法访问 {name}")
            return None
    raise AttributeError(f"module '{__name__}' has no attribute '{name}'")


# ========== 在关键操作处添加记录 ==========
# 辅助函数：根据token获取用户信息
async def get_user_info_from_token(auth_token: str) -> Dict[str, Any]:
    """根据authToken获取用户信息"""
    try:
        # 调用LMS的auth/token接口
        lms_auth_url = f"{LMS_BASE_URL}/auth/token?token={auth_token}"
        response = requests.get(lms_auth_url)

        if response.status_code == 200:
            return response.json().get("data", {})
        else:
            return {}
    except Exception:
        return {}


# 1. 在登录接口添加记录
@app.post("/login")
async def login(request: Request):
    """处理前端登录请求，调用LMS的login接口"""
    try:
        # 从前端获取用户名和密码
        data = await request.json()
        username = data.get("username")
        password = data.get("password")

        # 验证输入
        if not username or not password:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="用户名和密码不能为空"
            )

        # 调用LMS的login接口
        lms_login_url = f"{LMS_BASE_URL}/login"
        headers = {
            "userCode": username,
            "password": password
        }
        logger.info(f"尝试连接LMS服务: {lms_login_url}")
        try:
            response = requests.get(lms_login_url, headers=headers, timeout=5)
        except requests.exceptions.ConnectionError as e:
            logger.error(f"无法连接到LMS服务 {lms_login_url}: {str(e)}")
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail=f"无法连接到LMS服务，请确保LMS服务正在运行（{LMS_BASE_URL}）"
            )
        except requests.exceptions.Timeout:
            logger.error(f"连接LMS服务超时: {lms_login_url}")
            raise HTTPException(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                detail="LMS服务响应超时"
            )

        if response.status_code == 200:
            # 获取LMS返回的token
            lms_response = response.json()
            token = lms_response.get("authToken")

            if not token:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="登录成功但未返回authToken"
                )

            # 记录登录操作
            client_host = request.client.host if request.client else "unknown"
            log_operation(
                operation_type="user_login",
                action="用户登录",
                user_id=lms_response.get("userId"),
                user_name=lms_response.get("userName"),
                status="success",
                ip_address=client_host,
                details={
                    "login_method": "password",
                    "user_level": lms_response.get("userLevel")
                }
            )

            # 返回给前端的响应
            return {
                "success": True,
                "data": {
                    "userId": lms_response.get("userId"),
                    "userCode": lms_response.get("userCode"),
                    "userName": lms_response.get("userName"),
                    "authToken": token,
                    "userLevel": lms_response.get("userLevel"),
                }
            }
        else:
            # 记录登录失败
            client_host = request.client.host if request.client else "unknown"
            log_operation(
                operation_type="user_login",
                action="用户登录",
                user_id=username,
                status="failed",
                ip_address=client_host,
                details={
                    "error": response.text[:200],
                    "status_code": response.status_code
                }
            )

            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS登录失败: {response.text}"
            )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"登录请求失败: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"登录请求处理失败: {str(e)}"
        )


async def submit_inventory_task(task_no: str, bin_locations: List[str]):
    """下发盘点任务，接收任务编号和储位名称列表"""
    try:

        logger.info(f"下发盘点任务: {task_no}, 储位: {bin_locations}")

        url = f"{RCS_BASE_URL}{RCS_PREFIX}/api/robot/controller/task/submit"
        headers = {
            "X-lr-request-id": "ldui",
            "Content-Type": "application/json"
        }

        # 构建targetRoute数组
        target_route = []
        for index, location in enumerate(bin_locations):
            route_item = {
                "seq": index,
                "type": "ZONE",
                "code": location,  # 使用储位名称作为目标区域
            }
            target_route.append(route_item)

        # 构建请求体 - 单个任务对象
        request_body = {
            "taskType": "PF-CTU-COMMON-TEST",
            "targetRoute": target_route
        }

        response = requests.post(
            url, json=request_body, headers=headers, timeout=30)

        if response.status_code == 200:
            response_data = response.json()

            if response_data.get("code") == "SUCCESS":
                logger.info(f"储位 {bin_locations} 已发送到机器人系统")
                return {"success": True, "message": "盘点任务已下发"}
        else:
            return {"success": False, "message": "盘点任务下发失败"}

    except Exception as e:
        logger.error(f"下发盘点任务失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"下发盘点任务失败: {str(e)}"
        )


async def continue_inventory_task():
    """继续盘点任务"""
    try:
        logger.info(f"继续执行盘点任务")

        url = f"{RCS_BASE_URL}{RCS_PREFIX}/api/robot/controller/task/extend/continue"
        headers = {
            "X-lr-request-id": "ldui",
            "Content-Type": "application/json"
        }

        # 构建请求体
        request_body = {
            "triggerType": "TASK",
            "triggerCode": "001"
        }

        response = requests.post(
            url, json=request_body, headers=headers, timeout=30)

        if response.status_code == 200:
            response_data = response.json()

            if response_data.get("code") == "SUCCESS":
                logger.info(f"继续执行盘点任务命令已发送到机器人系统")
                return {"success": True, "message": "盘点任务已继续"}
        else:
            return {"success": False, "message": "盘点任务下发失败"}

    except Exception as e:
        logger.error(f"继续盘点任务失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"继续盘点任务失败: {str(e)}"
        )


async def update_robot_status(method: str, data: Optional[Dict] = None):
    """更新机器人状态并触发事件"""
    try:
        from services.api.state.robot_state import get_robot_state_manager
        manager = get_robot_state_manager()
        from services.api.config import ROBOT_STATUS_KEY

        # 保存状态信息
        manager._robot_status_store[ROBOT_STATUS_KEY] = {
            "method": method,
            "timestamp": time.time(),
            "data": data or {}
        }

        logger.info(f"更新机器人状态: {method}")

        # 设置事件，通知等待的进程
        manager._status_event.set()
    except ImportError:
        logger.warning("机器人状态管理器模块不存在，无法更新状态")


async def wait_for_robot_status(expected_method: str, timeout: int = 300):
    """
    等待特定机器人状态的同步函数

    这个函数会阻塞直到收到期望的状态或超时
    """
    try:
        from services.api.state.robot_state import get_robot_state_manager
        from services.api.config import ROBOT_STATUS_KEY
        manager = get_robot_state_manager()

        logger.info(f"开始等待机器人状态: {expected_method}, 超时: {timeout}秒")

        start_time = time.time()

        # 清除事件，确保我们等待的是新的事件
        manager._status_event.clear()

        # 检查是否已经有期望的状态
        if ROBOT_STATUS_KEY in manager._robot_status_store:
            current_status = manager._robot_status_store[ROBOT_STATUS_KEY]
            if current_status.get("method") == expected_method:
                logger.info(f"已存在期望状态: {expected_method}")
                return current_status

        while True:
            try:
                # 等待事件被设置
                await asyncio.wait_for(manager._status_event.wait(), timeout=1.0)

                # 检查状态
                if ROBOT_STATUS_KEY in manager._robot_status_store:
                    current_status = manager._robot_status_store[ROBOT_STATUS_KEY]
                    logger.info(f"收到机器人状态: {current_status.get('method')}")

                    if current_status.get("method") == expected_method:
                        logger.info(f"收到期望状态: {expected_method}")
                        return current_status

                # 重置事件，准备下一次等待
                manager._status_event.clear()

            except asyncio.TimeoutError:
                # 检查是否总时间超时
                elapsed_time = time.time() - start_time
                if elapsed_time >= timeout:
                    logger.error(f"等待机器人状态超时: {expected_method}")
                    raise asyncio.TimeoutError(f"等待 {expected_method} 状态超时")

                # 继续等待
                continue
    except ImportError:
        logger.error("机器人状态管理器模块不存在，无法等待状态")
        raise


async def check_image_exists(task_no: str, bin_location: str, camera_type: str) -> bool:
    """
    检查指定相机类型的图片是否存在

    Args:
        task_no: 任务编号
        bin_location: 储位名称
        camera_type: 相机类型 (3d_camera, scan_camera_1, scan_camera_2)

    Returns:
        True 如果图片存在，否则 False
    """
    try:
        project_root = Path(__file__).parent.parent.parent
        image_dir = project_root / "capture_img" / task_no / bin_location / camera_type

        if not image_dir.exists():
            logger.warning(f"图片目录不存在: {image_dir}")
            return False

        # 查找图片文件
        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
        for ext in image_extensions:
            # 查找常见的文件名
            for common_name in ['main', 'MAIN', 'raw', 'RAW', 'image', 'IMAGE']:
                image_file = image_dir / f"{common_name}{ext}"
                if image_file.exists():
                    logger.info(f"找到图片文件: {image_file}")
                    return True

        # 如果没有找到常见文件名，查找第一张图片
        for ext in image_extensions:
            image_files = list(image_dir.glob(f"*{ext}"))
            if image_files:
                logger.info(f"找到图片文件: {image_files[0]}")
                return True

        logger.warning(f"未找到图片文件: {image_dir}")
        return False

    except Exception as e:
        logger.error(f"检查图片存在性失败: {str(e)}")
        return False


async def execute_capture_script(script_path: str, task_no: str, bin_location: str) -> Dict[str, Any]:
    """
    在指定 Conda 环境中执行单个抓图脚本

    Args:
        script_path: 脚本路径
        task_no: 任务编号
        bin_location: 储位名称
        conda_env: Conda 环境名称，默认为 'your_env_name'

    Returns:
        脚本执行结果
    """
    conda_env = "tobacco_env"
    try:
        logger.info(f"在 Conda 环境 '{conda_env}' 中执行抓图脚本: {script_path}")

        # 方法1: 使用 conda run 命令
        # 构建命令行参数
        cmd = ["conda", "run", "-n", conda_env, "python", script_path,
               "--task-no", task_no, "--bin-location", bin_location]

        # 方法2: 直接使用 conda 环境中的 python 路径（如果知道路径）
        # 假设你的 conda 环境路径是已知的
        # conda_python_path = f"/home/user/anaconda3/envs/{conda_env}/bin/python"
        # cmd = [conda_python_path, script_path, "--task-no", task_no, "--bin-location", bin_location]

        # 执行脚本
        process = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.PIPE
        )

        # 等待脚本完成
        stdout, stderr = await process.communicate()

        # 解析结果
        result = {
            "script": os.path.basename(script_path),
            "conda_env": conda_env,
            "returncode": process.returncode,
            "stdout": stdout.decode('utf-8') if stdout else "",
            "stderr": stderr.decode('utf-8') if stderr else "",
            "success": process.returncode == 0
        }

        if process.returncode == 0:
            logger.info(f"脚本执行成功: {script_path} (环境: {conda_env})")
        else:
            logger.error(
                f"脚本执行失败: {script_path}, 错误: {stderr.decode('utf-8')}")

        return result

    except FileNotFoundError as e:
        logger.error(f"conda 命令未找到或 Conda 环境 '{conda_env}' 不存在: {str(e)}")
        return {
            "script": os.path.basename(script_path),
            "conda_env": conda_env,
            "returncode": -1,
            "stdout": "",
            "stderr": f"Conda 环境 '{conda_env}' 未找到或 conda 命令不可用",
            "success": False
        }
    except Exception as e:
        logger.error(f"执行脚本失败 {script_path}: {str(e)}")
        return {
            "script": os.path.basename(script_path),
            "conda_env": conda_env,
            "returncode": -1,
            "stdout": "",
            "stderr": str(e),
            "success": False
        }


async def capture_images_with_scripts(task_no: str, bin_location: str) -> Dict[str, Any]:
    """
    按顺序执行抓图脚本，并检查图片是否存在

    Args:
        task_no: 任务编号
        bin_location: 储位名称

    Returns:
        抓图结果，包含每个相机的抓图状态和图片路径
    """
    results = {
        "success": True,
        "camera_results": [],
        "photo3dPath": None,
        "photoDepthPath": None
    }

    # 定义相机类型与脚本的对应关系
    # CAPTURE_SCRIPTS 只有两个脚本，对应 scan_camera_1 和 scan_camera_2
    # 还需要添加 3d_camera 的处理
    camera_types = ["3d_camera", "scan_camera_1", "scan_camera_2"]

    # 处理 3d_camera（如果有对应的脚本，则使用脚本，否则直接检查图片）
    camera_type = "3d_camera"
    logger.info(f"检查 {camera_type} 图片是否存在")

    # 3d_camera 可能没有对应的抓图脚本，或者使用其他方式抓图
    # 这里我们假设 3d_camera 的图片是由其他机制生成的
    # 如果 3d_camera 有对应的脚本，可以在这里添加
    # 否则，我们假设图片会在抓图脚本执行后生成

    for i, script_path in enumerate(CAPTURE_SCRIPTS, 1):
        camera_type = camera_types[i]  # 跳过 3d_camera，从 scan_camera_1 开始
        logger.info(f"开始执行第 {i} 个抓图脚本: {script_path} (相机类型: {camera_type})")

        capture_success = False
        last_error = None

        # 重试机制：最多重试 3 次，每次超时 5 秒
        for retry in range(3):
            try:
                # 检查脚本文件是否存在
                if not os.path.exists(script_path):
                    error_msg = f"脚本文件不存在: {script_path}"
                    logger.error(error_msg)
                    last_error = error_msg
                    break

                # 执行脚本
                result = await execute_capture_script(script_path, task_no, bin_location)

                if result.get("success"):
                    # 脚本执行成功，等待一段时间让图片保存
                    await asyncio.sleep(1)

                    # 检查图片是否存在
                    image_exists = await check_image_exists(task_no, bin_location, camera_type)

                    if image_exists:
                        logger.info(f"{camera_type} 图片抓取成功")
                        capture_success = True
                        results["camera_results"].append({
                            "camera_type": camera_type,
                            "success": True,
                            "retry_count": retry + 1
                        })
                        break
                    else:
                        warning_msg = f"脚本执行成功，但 {camera_type} 图片不存在"
                        logger.warning(warning_msg)
                        last_error = warning_msg
                        if retry < 2:
                            logger.info(f"重试第 {retry + 2} 次...")
                            await asyncio.sleep(1)
                        continue
                else:
                    error_msg = f"脚本执行失败: {result.get('stderr', '未知错误')}"
                    logger.error(error_msg)
                    last_error = error_msg
                    if retry < 2:
                        logger.info(f"重试第 {retry + 2} 次...")
                        await asyncio.sleep(1)
                    continue

            except Exception as e:
                error_msg = f"执行脚本时发生异常: {str(e)}"
                logger.error(error_msg)
                last_error = error_msg
                if retry < 2:
                    logger.info(f"重试第 {retry + 2} 次...")
                    await asyncio.sleep(1)
                continue

        if not capture_success:
            logger.error(f"{camera_type} 图片抓取失败（重试 3 次后仍失败）")
            results["camera_results"].append({
                "camera_type": camera_type,
                "success": False,
                "error": last_error
            })
            results["success"] = False

    # 检查 3d_camera 图片（假设 3d_camera 图片会在其他相机抓图后生成）
    if await check_image_exists(task_no, bin_location, "3d_camera"):
        results["photo3dPath"] = f"/{task_no}/{bin_location}/3d_camera/MAIN.JPG"
        results["camera_results"].append({
            "camera_type": "3d_camera",
            "success": True,
            "retry_count": 1
        })
        logger.info("3d_camera 图片存在")
    else:
        logger.warning("3d_camera 图片不存在")
        results["camera_results"].append({
            "camera_type": "3d_camera",
            "success": False,
            "error": "3d_camera 图片不存在"
        })
        results["success"] = False

    # 检查 scan_camera_1 图片，作为 depth 图片
    if await check_image_exists(task_no, bin_location, "scan_camera_1"):
        results["photoDepthPath"] = f"/{task_no}/{bin_location}/scan_camera_1/MAIN.JPG"
        logger.info("scan_camera_1 图片存在（作为 depth 图片）")
    else:
        logger.warning("scan_camera_1 图片不存在")

    return results


async def run_barcode_and_detect(
    task_no: str,
    bin_location: str,
    image_dir: Path,
    pile_id: int = 1,
    code_type: str = "ucc128"
) -> dict:
    """
    执行条码识别和数量检测的公共函数

    Args:
        task_no: 任务编号
        bin_location: 储位名称
        image_dir: 图片目录路径 (Path对象)
        pile_id: 初始垛型ID，默认为1
        code_type: 条码类型，默认为ucc128

    Returns:
        dict: 包含 barcode_result, detect_result, photos 的结果字典
    """
    project_root = Path(__file__).parent.parent.parent
    recognition_time = datetime.now().isoformat()

    results = {
        "taskNo": task_no,
        "binLocation": bin_location,
        "recognition_time": recognition_time,
        "barcode_result": None,
        "detect_result": None,
        "photos": []
    }

    # ========== Step 0: 查找照片文件 ==========
    try:
        photos = []
        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']

        # 查找主图 (3d_camera 目录)
        if image_dir.exists():
            for ext in image_extensions:
                for common_name in ['main', 'MAIN', 'raw', 'RAW', 'image', 'IMAGE']:
                    main_file = image_dir / f"{common_name}{ext}"
                    if main_file.exists():
                        relative_path = f"{task_no}/{bin_location}/3d_camera/{common_name.upper()}{ext}"
                        photos.append(f"/{relative_path}")
                        logger.info(f"找到主图: {relative_path}")
                        break
                if photos:
                    break

        # 查找深度图 (3d_camera 目录，与主图在同一目录)
        depth_dir = project_root / "capture_img" / f"{task_no}/{bin_location}/3d_camera"
        if depth_dir.exists():
            for ext in image_extensions:
                depth_file = depth_dir / f"depth{ext}"
                if depth_file.exists():
                    relative_path = f"{task_no}/{bin_location}/3d_camera/depth{ext}"
                    photos.append(f"/{relative_path}")
                    logger.info(f"找到深度图: {relative_path}")
                    break

        results["photos"] = photos
        logger.info(f"共找到 {len(photos)} 张图片")

    except Exception as e:
        logger.error(f"查找照片路径失败: {str(e)}", exc_info=True)
        results["photos"] = []

    # ========== Step 1: 执行 Barcode 模块获取垛型信息 ==========
    detected_pile_id = pile_id  # 默认使用传入的值

    if ENABLE_BARCODE and BARCODE_MODULE_AVAILABLE:
        try:
            logger.info(f"开始执行Barcode模块识别: {task_no}/{bin_location}")

            recognizer = BarcodeRecognizer(code_type=code_type)

            # Barcode 只处理 scan_camera_1 和 scan_camera_2 的图片
            parent_dir = image_dir.parent  # 从 3d_camera 的父目录开始
            scan_dirs = [
                parent_dir / "scan_camera_1",
                parent_dir / "scan_camera_2"
            ]

            all_barcode_results = []
            for scan_dir in scan_dirs:
                if scan_dir.exists():
                    logger.info(f"[Barcode] 处理扫描目录: {scan_dir}")
                    results_part = recognizer.process_folder(input_dir=str(scan_dir))
                    # 添加目录信息到结果
                    for r in results_part:
                        r['source_dir'] = scan_dir.name
                    all_barcode_results.extend(results_part)
                else:
                    logger.warning(f"[Barcode] 扫描目录不存在: {scan_dir}")

            barcode_results = all_barcode_results

            # 记录所有识别结果
            logger.info(f"[Barcode] 识别完成，共处理 {len(barcode_results)} 张图片")
            for idx, result in enumerate(barcode_results):
                filename = result.get('filename', f'unknown_{idx}')
                source_dir = result.get('source_dir', 'unknown')
                output = result.get('output', '')
                error = result.get('error', '')
                if output:
                    logger.info(f"[Barcode] 结果 #{idx+1}: 目录={source_dir}, 图片={filename}, 识别内容='{output}'")
                elif error:
                    logger.warning(f"[Barcode] 结果 #{idx+1}: 目录={source_dir}, 图片={filename}, 错误={error}")
                else:
                    logger.warning(f"[Barcode] 结果 #{idx+1}: 目录={source_dir}, 图片={filename}, 未识别到条码")

            resolver = get_tobacco_case_resolver()

            def extract_barcodes_from_output(output_str):
                """从嵌套 JSON 中提取所有条码 text"""
                if not output_str:
                    return []
                try:
                    data = json.loads(output_str)
                    barcodes = []
                    for session in data.get('sessions', []):
                        for barcode in session.get('barcodes', []):
                            text = barcode.get('text')
                            if text:
                                barcodes.append(text)
                    if not barcodes:
                        logger.warning(f"[Barcode] JSON解析成功但sessions/barcodes结构为空，data keys={list(data.keys()) if isinstance(data, dict) else type(data)}")
                    return barcodes
                except (json.JSONDecodeError, AttributeError):
                    # 如果不是 JSON 格式，直接返回原字符串作为单个条码
                    return [output_str] if output_str else []

            def _extract_six_digits(barcode: str) -> str:
                """从条码中提取6位数字（忽略91前缀）"""
                if not barcode:
                    return ""
                # 移除 (91) 或 91 前缀
                processed = barcode
                if barcode.startswith('(91)'):
                    processed = barcode[4:]
                elif barcode.startswith('91'):
                    processed = barcode[2:]
                # 提取前6位连续数字
                import re
                match = re.search(r'(\d{6})', processed)
                return match.group(1) if match else ""

            resolved_info = None
            tried_barcodes = []  # 记录尝试过的条码
            for bar_result in barcode_results:
                output = bar_result.get('output', '')
                filename = bar_result.get('filename', 'unknown')
                source_dir = bar_result.get('source_dir', 'unknown')

                # 解析嵌套 JSON 提取所有条码
                barcode_texts = extract_barcodes_from_output(output)

                # 诊断日志：条码提取结果
                if not barcode_texts:
                    logger.warning(f"[Barcode] 解析条码失败: 图片={filename}, output长度={len(output)}, output前100字符='{output[:100]}'")

                for barcode_text in barcode_texts:
                    # 输出条码处理详细日志
                    logger.info(f"成功识别条码: {barcode_text}")
                    logger.info(f"原始条码数据: {barcode_text}")
                    processed_barcode = barcode_text
                    if barcode_text.startswith('(91)'):
                        processed_barcode = barcode_text[4:]
                    elif barcode_text.startswith('91'):
                        processed_barcode = barcode_text[2:]
                    logger.info(f"移除91前缀后: {processed_barcode}")
                    six_digits = _extract_six_digits(barcode_text)
                    logger.info(f"提取的6位数字: {six_digits}")

                    logger.info(f"[Barcode] 尝试匹配烟箱信息: 目录={source_dir}, 图片={filename}")
                    tried_barcodes.append({'filename': filename, 'source_dir': source_dir, 'barcode': barcode_text, 'six_digits': six_digits})
                    resolved_info = resolver.resolve(barcode_text)
                    if resolved_info['success']:
                        product_name = resolved_info.get('product_name', '')
                        logger.info(f"成功匹配烟箱信息: {product_name}")
                        break
                    else:
                        logger.warning(f"[Barcode] 条码匹配失败: '{barcode_text}', 原因: {resolved_info.get('message', '未找到匹配记录')}")
                if resolved_info and resolved_info['success']:
                    break

            if resolved_info and resolved_info['success']:
                detected_pile_id = resolved_info['pile_id']
                logger.info(f"映射后的 pile_id: {detected_pile_id} (原请求: {pile_id})")

                results["barcode_result"] = {
                    "image_path": str(image_dir),
                    "code_type": code_type,
                    "six_digit_code": resolved_info.get('six_digit_code'),
                    "stack_type_1": resolved_info.get('stack_type_1'),
                    "product_name": resolved_info.get('product_name'),
                    "tobacco_code": resolved_info.get('tobacco_code'),
                    "mapped_pile_id": detected_pile_id,
                    "total_images": len(barcode_results),
                    "status": "success"
                }
            else:
                if len(tried_barcodes) == 0:
                    logger.warning(f"[Barcode] 未检测到任何条码（目录可能不存在或图片为空）")
                else:
                    logger.warning(f"条码识别成功但未匹配到烟箱信息，共尝试 {len(tried_barcodes)} 个条码")
                for tried in tried_barcodes:
                    logger.warning(f"  - 目录: {tried['source_dir']}, 图片: {tried['filename']}, 条码: '{tried['barcode']}'")
                results["barcode_result"] = {
                    "image_path": str(image_dir),
                    "code_type": code_type,
                    "results": barcode_results,
                    "tried_barcodes": tried_barcodes,
                    "total_images": len(barcode_results),
                    "status": "no_match",
                    "message": "条码识别成功但未匹配到烟箱信息"
                }

        except Exception as e:
            logger.error(f"Barcode模块识别失败: {str(e)}", exc_info=True)
            results["barcode_result"] = {
                "status": "failed",
                "error": str(e)
            }
    else:
        logger.info("Barcode模块已禁用，跳过识别")
        results["barcode_result"] = {
            "status": "disabled",
            "message": "条形码功能已禁用（ENABLE_BARCODE=false）"
        }

    # ========== Step 2: 执行 Detect 模块 ==========
    if DETECT_MODULE_AVAILABLE:
        if not image_dir.exists() or not image_dir.is_dir():
            logger.error(f"图片目录不存在: {image_dir}")
            results["detect_result"] = {
                "status": "failed",
                "error": "图片目录不存在"
            }
            return results

        try:
            logger.info(f"开始执行Detect模块识别: {task_no}/{bin_location}, pile_id={detected_pile_id}")

            # 查找主图片文件
            image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
            image_files = []

            common_names = ['main', 'raw', 'image', 'img', 'photo']
            for name in common_names:
                for ext in image_extensions:
                    common_file = image_dir / f"{name}{ext}"
                    if common_file.exists():
                        image_files.append(common_file)
                        break
                if image_files:
                    break

            if not image_files:
                for ext in image_extensions:
                    image_files.extend(list(image_dir.glob(f"*{ext}")))
                    if image_files:
                        break

            if not image_files:
                results["detect_result"] = {
                    "status": "failed",
                    "error": f"在路径 {image_dir} 中未找到图片文件"
                }
                return results

            image_path_for_detect = str(image_files[0])

            # 查找深度图（与主图在同一目录 3d_camera）
            depth_dir = project_root / "capture_img" / task_no / bin_location / "3d_camera"
            depth_image_path_for_detect = None
            for ext in image_extensions:
                depth_file = depth_dir / f"depth{ext}"
                if depth_file.exists():
                    depth_image_path_for_detect = str(depth_file)
                    logger.info(f"找到深度图文件: {depth_image_path_for_detect}")
                    break

            if not depth_image_path_for_detect:
                logger.warning(f"深度图文件不存在: {depth_dir}/depth.*")

            # 构建debug输出目录
            debug_output_dir = project_root / "debug" / task_no / bin_location
            debug_output_dir.mkdir(parents=True, exist_ok=True)

            # 创建日志文件处理器
            log_file_path = debug_output_dir / f"debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
            file_handler = logging.FileHandler(str(log_file_path), encoding='utf-8')
            file_handler.setLevel(logging.DEBUG)
            file_formatter = logging.Formatter(
                '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            file_handler.setFormatter(file_formatter)
            logger.addHandler(file_handler)

            try:
                logger.info(f"Debug输出目录: {debug_output_dir}")
                logger.info(f"使用主图片: {image_path_for_detect}")
                if depth_image_path_for_detect:
                    logger.info(f"使用深度图: {depth_image_path_for_detect}")

                # 执行检测
                total_count = count_boxes(
                    image_path=image_path_for_detect,
                    pile_id=detected_pile_id,
                    depth_image_path=depth_image_path_for_detect,
                    enable_debug=True,
                    enable_visualization=True,
                    output_dir=str(debug_output_dir)
                )

                logger.info(f"Detect模块识别完成，箱数: {total_count}")

                results["detect_result"] = {
                    "image_path": image_path_for_detect,
                    "pile_id": detected_pile_id,
                    "total_count": total_count,
                    "status": "success"
                }

            except Exception as e:
                logger.error(f"Detect模块识别失败: {str(e)}", exc_info=True)
                results["detect_result"] = {
                    "status": "failed",
                    "error": str(e)
                }
            finally:
                logger.removeHandler(file_handler)
                file_handler.close()

        except Exception as e:
            logger.error(f"Detect模块初始化失败: {str(e)}", exc_info=True)
            results["detect_result"] = {
                "status": "failed",
                "error": str(e)
            }
    else:
        logger.info("Detect模块不可用，跳过识别")
        results["detect_result"] = {
            "status": "unavailable",
            "message": "检测模块不可用"
        }

    # ========== Step 3: 更新任务状态 ==========
    storage = get_task_state_storage()
    inventory_task_bins = storage["bins"]
    inventory_task_details = storage["details"]

    if task_no in inventory_task_bins:
        for bin_status in inventory_task_bins[task_no]:
            if bin_status.bin_location == bin_location:
                bin_status.detect_result = results["detect_result"]
                bin_status.barcode_result = results["barcode_result"]
                bin_status.recognition_time = recognition_time
                break

    # 存储到任务详情
    if task_no not in inventory_task_details:
        inventory_task_details[task_no] = {}

    if bin_location not in inventory_task_details[task_no]:
        inventory_task_details[task_no][bin_location] = {}

    inventory_task_details[task_no][bin_location]["recognition"] = results

    # 保存识别结果为ini文件
    try:
        import configparser
        config = configparser.ConfigParser()

        config.add_section('BasicInfo')
        config.set('BasicInfo', 'taskNo', task_no)
        config.set('BasicInfo', 'binLocation', bin_location)
        config.set('BasicInfo', 'recognition_time', recognition_time)

        config.add_section('ImagePaths')
        config.set('ImagePaths', 'original_image_dir', str(image_dir))
        if results.get("detect_result") and results["detect_result"].get("image_path"):
            config.set('ImagePaths', 'detect_image_path', results["detect_result"]["image_path"])

        if results.get("detect_result"):
            config.add_section('DetectResult')
            detect_result = results["detect_result"]
            if detect_result.get("status"):
                config.set('DetectResult', 'status', detect_result["status"])
            if detect_result.get("total_count") is not None:
                config.set('DetectResult', 'total_count', str(detect_result["total_count"]))
            if detect_result.get("pile_id") is not None:
                config.set('DetectResult', 'pile_id', str(detect_result["pile_id"]))
            if detect_result.get("error"):
                config.set('DetectResult', 'error', detect_result["error"])

        if results.get("barcode_result"):
            config.add_section('BarcodeResult')
            barcode_result = results["barcode_result"]
            if barcode_result.get("status"):
                config.set('BarcodeResult', 'status', barcode_result["status"])
            if barcode_result.get("mapped_pile_id") is not None:
                config.set('BarcodeResult', 'mapped_pile_id', str(barcode_result["mapped_pile_id"]))
            if barcode_result.get("product_name"):
                config.set('BarcodeResult', 'product_name', barcode_result["product_name"])

        ini_file_path = image_dir / "recognition_result.ini"
        with open(ini_file_path, 'w', encoding='utf-8') as f:
            config.write(f)
        logger.info(f"识别结果已保存到: {ini_file_path}")

    except Exception as e:
        logger.error(f"保存识别结果ini文件失败: {str(e)}", exc_info=True)

    return results


async def process_single_bin_location(task_no: str, bin_location: str, index: int, total: int, is_sim: bool = False):
    """处理单个储位的完整流程"""
    result = {
        "binLocation": bin_location,
        "sequence": index + 1,
        "startTime": datetime.now().isoformat(),
        "endTime": None,
        "status": None,
        "actualQuantity": None,
        "actualSpec": None,
        "photo3dPath": None,
        "photoDepthPath": None,
        "error": None
    }

    try:
        # 更新任务状态（已在execute_inventory_workflow中处理）

        # 模拟模式处理
        if is_sim:
            logger.info(f"模拟模式：处理储位 {bin_location}")
            await asyncio.sleep(2)  # 模拟RCS移动到储位

            # 在模拟模式下，跳过抓图，直接使用模拟图片
            project_root = Path(__file__).parent.parent.parent
            capture_img_dir = project_root / "capture_img" / task_no / bin_location

            # 如果目录已存在，跳过拷贝
            if capture_img_dir.exists():
                logger.info(f"模拟模式：目录已存在，跳过拷贝 {capture_img_dir}")
            else:
                # 创建目录
                capture_img_dir.mkdir(parents=True, exist_ok=True)

                # 复制模拟图片到 capture_img 目录
                # 4张照片的映射关系：
                # 1.jpg -> 3d_camera/main.jpg
                # 2.jpg -> 3d_camera/depth.jpg (用于算法计算)
                #       -> 3d_camera/depth_color.jpg (用于UI显示)
                # 3.jpg -> scan_camera_1/main.jpg
                # 4.jpg -> scan_camera_2/main.jpg
                public_dir = project_root / "web" / "src" / "public"
                image_mapping = [
                    ("1.jpg", "3d_camera", "main.jpg"),
                    ("2.jpg", "3d_camera", "depth.jpg"),         # 用于算法
                    ("2.jpg", "3d_camera", "depth_color.jpg"),   # 用于UI显示
                    ("3.jpg", "scan_camera_1", "main.jpg"),
                    ("4.jpg", "scan_camera_2", "main.jpg")
                ]

                for img_file, camera_dir, dest_filename in image_mapping:
                    src_file = public_dir / img_file
                    dest_dir = capture_img_dir / camera_dir
                    dest_dir.mkdir(parents=True, exist_ok=True)
                    dest_file = dest_dir / dest_filename

                    if src_file.exists():
                        shutil.copy(src_file, dest_file)
                        logger.info(f"模拟模式：复制图片 {src_file} -> {dest_file}")
                    else:
                        logger.warning(f"模拟图片不存在: {src_file}")

            # 设置模拟图片路径（UI显示用旋转后的图片）
            result["photo3dPath"] = f"/{task_no}/{bin_location}/3d_camera/main_rotated.jpg"
            result["photoDepthPath"] = f"/{task_no}/{bin_location}/3d_camera/depth_color.jpg"
            result["photoScan1Path"] = f"/{task_no}/{bin_location}/scan_camera_1/main.jpg"
            result["photoScan2Path"] = f"/{task_no}/{bin_location}/scan_camera_2/main.jpg"

            logger.info(f"模拟模式：抓图成功: {bin_location}")

            # ========== 调用公共识别函数 ==========
            # 构建3d_camera目录路径
            image_dir = capture_img_dir / "3d_camera"

            # 调用公共识别函数（Barcode + Detect）
            recognition_result = await run_barcode_and_detect(
                task_no=task_no,
                bin_location=bin_location,
                image_dir=image_dir,
                pile_id=1,
                code_type="ucc128"
            )

            # 从识别结果中提取数量和品规
            detect_result = recognition_result.get("detect_result", {})
            barcode_result = recognition_result.get("barcode_result", {})

            # 优先使用识别结果，如果识别失败则使用系统库存作为备用
            if detect_result.get("status") == "success" and detect_result.get("total_count") is not None:
                result["actualQuantity"] = detect_result["total_count"]
                logger.info(f"模拟模式：使用Detect识别数量: {result['actualQuantity']}")
            else:
                # 识别失败，从inventoryItems中获取系统数量作为备用
                inventory_item = None
                if task_no in inventory_task_details and "inventoryItems" in inventory_task_details[task_no]:
                    for item in inventory_task_details[task_no]["inventoryItems"]:
                        if item.get("locationName") == bin_location:
                            inventory_item = item
                            break
                result["actualQuantity"] = inventory_item.get("systemQuantity", 0) if inventory_item else 0
                logger.info(f"模拟模式：Detect识别失败，使用系统数量: {result['actualQuantity']}")

            # 优先使用条码识别的品规，识别失败则显示"未识别"
            if barcode_result.get("status") == "success" and barcode_result.get("product_name"):
                result["actualSpec"] = barcode_result["product_name"]
                logger.info(f"模拟模式：使用Barcode识别品规: {result['actualSpec']}")
            else:
                result["actualSpec"] = "未识别"
                logger.info(f"模拟模式：Barcode识别失败，品规=未识别")

            # 更新结果中的识别详情
            result["barcodeResult"] = barcode_result
            result["detectResult"] = detect_result
            result["photos"] = recognition_result.get("photos", [])

            logger.info(
                f"模拟模式：识别完成，实际数量: {result['actualQuantity']}, 品规: {result['actualSpec']}")

            # 更新任务状态中的图片和计算结果（供前端获取）
            if task_no in inventory_task_bins:
                for bin_status in inventory_task_bins[task_no]:
                    if bin_status.bin_location == bin_location:
                        bin_status.image_data = {
                            "success": True,
                            "photo3dPath": result["photo3dPath"],
                            "photoDepthPath": result["photoDepthPath"]
                        }
                        bin_status.compute_result = {
                            "actualQuantity": result["actualQuantity"],
                            "actualSpec": result["actualSpec"]
                        }
                        bin_status.capture_time = result["startTime"]
                        bin_status.compute_time = datetime.now().isoformat()
                        # 新增识别结果
                        bin_status.detect_result = detect_result
                        bin_status.barcode_result = barcode_result
                        break

            # 存储详细结果
            if task_no not in inventory_task_details:
                inventory_task_details[task_no] = {}

            inventory_task_details[task_no][bin_location] = {
                "image_data": {
                    "success": True,
                    "photo3dPath": result["photo3dPath"],
                    "photoDepthPath": result["photoDepthPath"]
                },
                "compute_result": {
                    "actualQuantity": result["actualQuantity"],
                    "actualSpec": result["actualSpec"]
                },
                "capture_time": result["startTime"],
                "compute_time": datetime.now().isoformat(),
                "updated_at": datetime.now().isoformat(),
                # 新增识别结果
                "recognition": recognition_result
            }

            result["status"] = "成功"
            result["endTime"] = datetime.now().isoformat()

            # 模拟继续任务接口
            if ((index + 1) < total):
                logger.info(f"模拟模式：继续下一个储位")
                await asyncio.sleep(1)

        else:
            # 真实模式处理
            # 等待机器人就位
            logger.info(f"============等待机器人就位信息: {bin_location}")
            try:
                ctu_status = await wait_for_robot_status("end", timeout=300)

                # 这个判断一定会执行，因为wait_for_robot_status会阻塞直到收到end状态或超时
                if ctu_status and ctu_status.get("method") == "end":

                    # 执行抓图脚本（带超时重试机制）
                    capture_results = await capture_images_with_scripts(task_no, bin_location)
                    result["captureResults"] = capture_results

                    # 检查抓图结果
                    if not capture_results.get("success"):
                        logger.error(f"抓图失败，跳过储位: {bin_location}")
                        result["status"] = "异常"
                        result["error"] = "抓图失败（重试3次后仍失败）"
                        result["endTime"] = datetime.now().isoformat()

                        # 即使失败，也要调用继续任务接口，跳到下一个储位
                        if ((index + 1) < total):
                            logger.info(f"跳过储位 {bin_location}，继续下一个储位")
                            continue_result = await continue_inventory_task()
                            logger.info(f"继续任务接口调用结果: {continue_result}")
                            result["continueResult"] = continue_result

                        # 返回异常结果，不执行后续流程
                        return result

                    logger.info(f"抓图成功: {bin_location}")

                    # 设置照片路径
                    result["photo3dPath"] = capture_results.get("photo3dPath")
                    result["photoDepthPath"] = capture_results.get(
                        "photoDepthPath")

                    # ========== 调用公共识别函数 ==========
                    # 构建3d_camera目录路径
                    project_root = Path(__file__).parent.parent.parent
                    capture_img_dir = project_root / "capture_img" / task_no / bin_location
                    image_dir = capture_img_dir / "3d_camera"

                    # 调用公共识别函数（Barcode + Detect）
                    recognition_result = await run_barcode_and_detect(
                        task_no=task_no,
                        bin_location=bin_location,
                        image_dir=image_dir,
                        pile_id=1,
                        code_type="ucc128"
                    )

                    detect_result = recognition_result.get("detect_result", {})
                    barcode_result = recognition_result.get("barcode_result", {})

                    # 优先使用识别结果，如果识别失败则使用系统库存作为备用
                    if detect_result.get("status") == "success" and detect_result.get("total_count") is not None:
                        result["actualQuantity"] = detect_result["total_count"]
                        logger.info(f"真实模式：使用Detect识别数量: {result['actualQuantity']}")
                    else:
                        # 识别失败，从inventoryItems中获取系统数量作为备用
                        inventory_item = None
                        if task_no in inventory_task_details and "inventoryItems" in inventory_task_details[task_no]:
                            for item in inventory_task_details[task_no]["inventoryItems"]:
                                if item.get("locationName") == bin_location:
                                    inventory_item = item
                                    break
                        result["actualQuantity"] = inventory_item.get("systemQuantity", 0) if inventory_item else 0
                        logger.info(f"真实模式：Detect识别失败，使用系统数量: {result['actualQuantity']}")

                    # 优先使用条码识别的品规，识别失败则显示"未识别"
                    if barcode_result.get("status") == "success" and barcode_result.get("product_name"):
                        result["actualSpec"] = barcode_result["product_name"]
                        logger.info(f"真实模式：使用Barcode识别品规: {result['actualSpec']}")
                    else:
                        result["actualSpec"] = "未识别"
                        logger.info(f"真实模式：Barcode识别失败，品规=未识别")

                    # 更新结果中的识别详情
                    result["barcodeResult"] = barcode_result
                    result["detectResult"] = detect_result
                    result["photos"] = recognition_result.get("photos", [])

                    # 更新任务状态中的图片和计算结果（供前端获取）
                    if task_no in inventory_task_bins:
                        for bin_status in inventory_task_bins[task_no]:
                            if bin_status.bin_location == bin_location:
                                bin_status.image_data = capture_results
                                bin_status.compute_result = {
                                    "actualQuantity": result["actualQuantity"],
                                    "actualSpec": result["actualSpec"]
                                }
                                bin_status.capture_time = result["startTime"]
                                bin_status.compute_time = datetime.now().isoformat()
                                # 新增识别结果
                                bin_status.detect_result = detect_result
                                bin_status.barcode_result = barcode_result
                                break

                    # 存储详细结果
                    if task_no not in inventory_task_details:
                        inventory_task_details[task_no] = {}

                    inventory_task_details[task_no][bin_location] = {
                        "image_data": capture_results,
                        "compute_result": {
                            "actualQuantity": result["actualQuantity"],
                            "actualSpec": result["actualSpec"]
                        },
                        "capture_time": result["startTime"],
                        "compute_time": datetime.now().isoformat(),
                        "updated_at": datetime.now().isoformat(),
                        # 新增识别结果
                        "recognition": recognition_result
                    }

                    result["status"] = "成功"
                    result["endTime"] = datetime.now().isoformat()

                    if ((index + 1) < total):
                        logger.info(f"收到机器人结束状态: {bin_location}")

                        # 只有在收到end状态后才调用继续任务接口
                        continue_result = await continue_inventory_task()
                        logger.info(f"继续任务接口调用结果: {continue_result}")
                        result["continueResult"] = continue_result

                else:
                    # 正常情况下不会执行到这里，除非wait_for_robot_status返回了非end状态
                    logger.warning(f"未收到预期的结束状态，当前状态: {ctu_status}")

            except asyncio.TimeoutError as e:
                logger.error(f"等待机器人结束状态超时: {str(e)}")
                result["status"] = "异常"
                result["error"] = "等待机器人结束状态超时"
                result["endTime"] = datetime.now().isoformat()
                return result

    except Exception as e:
        result["status"] = "异常"
        result["error"] = str(e)
        result["endTime"] = datetime.now().isoformat()
        logger.error(f"处理储位失败 {bin_location}: {str(e)}")

        # 记录错误但继续处理下一个储位（根据业务需求决定是否中断）
        # 可以发送错误通知到前端
        # try:
        #     async with APIClient(SERVICE_CONFIG["notification_service"]) as client:
        #         error_payload = {
        #             "taskNo": task_no,
        #             "binLocation": bin_location,
        #             "error": str(e),
        #             "timestamp": datetime.now().isoformat(),
        #             "messageType": "ERROR"
        #         }
        #         await client.post("/api/notification/error", json=error_payload)
        # except:
        #     pass

    return result


async def execute_inventory_workflow(task_no: str, bin_locations: List[str], is_sim: bool = False):
    """执行完整的盘点工作流"""
    logger.info(
        f"开始执行盘点工作流: {task_no}, 共 {len(bin_locations)} 个储位, 模拟模式: {is_sim}")

    # 初始化任务状态
    task_status = TaskStatus(
        task_no=task_no,
        status="init",
        current_step=0,
        total_steps=len(bin_locations),
        start_time=datetime.now().isoformat()
    )

    # 按任务编号存储任务状态
    inventory_tasks[task_no] = task_status

    # 初始化每个储位的状态
    bin_statuses = [
        BinLocationStatus(
            bin_location=location,
            status="pending",
            sequence=index + 1
        )
        for index, location in enumerate(bin_locations)
    ]
    inventory_task_bins[task_no] = bin_statuses

    # 更新任务状态为运行中
    inventory_tasks[task_no].status = "running"
    inventory_tasks[task_no].current_step = 0

    # 整体下发盘点任务
    method = "start"
    await update_robot_status(method)

    submit_result = await submit_inventory_task(task_no, bin_locations)

    # 存储所有储位的盘点结果
    inventory_results = []

    try:
        # 并行处理所有储位
        logger.info(f"开始并行处理 {len(bin_locations)} 个储位")

        # 先将所有储位状态设为运行中
        if task_no in inventory_task_bins:
            for bin_status in inventory_task_bins[task_no]:
                bin_status.status = "running"

        # 并行执行所有 bin 的处理
        async def process_and_wrap(i: int, bin_location: str):
            result = await process_single_bin_location(
                task_no=task_no,
                bin_location=bin_location,
                index=i,
                total=len(bin_locations),
                is_sim=is_sim
            )
            return bin_location, result

        tasks = [
            process_and_wrap(i, bin_location)
            for i, bin_location in enumerate(bin_locations)
        ]
        bin_results = await asyncio.gather(*tasks)

        # 更新每个 bin 的状态并收集结果
        for bin_location, result in bin_results:
            # 更新 bin 状态
            if task_no in inventory_task_bins:
                for bin_status in inventory_task_bins[task_no]:
                    if bin_status.bin_location == bin_location:
                        bin_status.status = "completed" if result["status"] == "成功" else "failed"
                        break

            # 从原始盘点项中查找匹配的储位信息
            inventory_item = None
            if task_no in inventory_task_details and "inventoryItems" in inventory_task_details[task_no]:
                for item in inventory_task_details[task_no]["inventoryItems"]:
                    if item.get("locationName") == bin_location:
                        inventory_item = item
                        break

            # 计算差异
            actual_qty = result.get("actualQuantity", 0)
            system_qty = inventory_item.get("systemQuantity", 0) if inventory_item else 0
            difference = actual_qty - system_qty

            # 收集盘点结果
            inventory_results.append({
                "binLocation": bin_location,
                "status": result["status"],
                "actualQuantity": result.get("actualQuantity"),
                "actualSpec": result.get("actualSpec"),
                "photo3dPath": result.get("photo3dPath"),
                "photoDepthPath": result.get("photoDepthPath"),
                "photoScan1Path": result.get("photoScan1Path", ""),
                "photoScan2Path": result.get("photoScan2Path", ""),
                "error": result.get("error"),
                "specName": inventory_item.get("productName", "") if inventory_item else "",
                "systemQuantity": system_qty,
                "difference": difference,
            })

        logger.info(f"所有 {len(bin_locations)} 个储位处理完成")

        # 更新任务状态为完成
        inventory_tasks[task_no].status = "completed"
        inventory_tasks[task_no].current_step = len(bin_locations)
        inventory_tasks[task_no].end_time = datetime.now().isoformat()

        # 保存盘点结果到任务详情
        if task_no not in inventory_task_details:
            inventory_task_details[task_no] = {}

        inventory_task_details[task_no]["inventoryResults"] = inventory_results

        logger.info(f"盘点任务完成: {task_no}, 成功处理 {len(bin_locations)} 个储位")
        logger.info(f"盘点结果: {inventory_results}")

        # 记录任务完成
        log_operation(
            operation_type="inventory",
            action="盘点任务完成",
            target=task_no,
            status="completed",
            details={
                "task_no": task_no,
                "bin_locations": bin_locations,
                "completed_count": len(bin_locations),
                "xlsx_file": f"./output/history_data/{task_no}.xlsx"
            }
        )

        # 发送任务完成通知
        # try:
        #     async with APIClient(SERVICE_CONFIG["notification_service"]) as client:
        #         completion_payload = {
        #             "taskNo": task_no,
        #             "status": "COMPLETED",
        #             "totalBins": len(bin_locations),
        #             "successfulBins": sum(1 for r in inventory_tasks[task_no].results
        #                                   if r.get("status") == "completed"),
        #             "failedBins": sum(1 for r in inventory_tasks[task_no].results
        #                               if r.get("status") == "failed"),
        #             "completionTime": datetime.now().isoformat(),
        #             "messageType": "TASK_COMPLETED"
        #         }
        #         await client.post("/api/notification/task-complete", json=completion_payload)
        # except Exception as e:
        #     logger.warning(f"发送任务完成通知失败: {str(e)}")

    except Exception as e:
        # 任务执行过程中出现异常
        if task_no in inventory_tasks:
            inventory_tasks[task_no].status = "failed"
            inventory_tasks[task_no].end_time = datetime.now().isoformat()

        # 记录任务失败
        log_operation(
            operation_type="inventory",
            action="盘点任务失败",
            target=task_no,
            status="failed",
            details={
                "task_no": task_no,
                "error": str(e),
                "failed_at_step": inventory_tasks[task_no].current_step if task_no in inventory_tasks else 0
            }
        )

        logger.error(f"盘点任务失败: {task_no}, 错误: {str(e)}")

        # 发送任务失败通知
        # try:
        #     async with APIClient(SERVICE_CONFIG["notification_service"]) as client:
        #         error_payload = {
        #             "taskNo": task_no,
        #             "status": "FAILED",
        #             "error": str(e),
        #             "failedAtBin": inventory_tasks[task_no].current_bin,
        #             "completedBins": len(inventory_tasks[task_no].results),
        #             "timestamp": datetime.now().isoformat(),
        #             "messageType": "TASK_FAILED"
        #         }
        #         await client.post("/api/notification/task-error", json=error_payload)
        # except Exception as e2:
        #     logger.error(f"发送任务失败通知失败: {str(e2)}")


######################################### 盘点任务接口 #########################################


@app.post("/api/inventory/start-inventory")
async def start_inventory(request: Request, background_tasks: BackgroundTasks):
    """启动盘点任务，接收任务编号和储位名称列表"""
    try:
        data = await request.json()
        task_no = data.get("taskNo")
        bin_locations = data.get("binLocations", [])
        is_sim = data.get("is_sim", False)  # 新增模拟模式参数
        inventory_items = data.get("inventoryItems", [])  # 接收完整的盘点项列表

        if not task_no or not bin_locations:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="任务编号和储位名称列表不能为空"
            )

        logger.info(
            f"启动盘点任务: {task_no}, 包含 {len(bin_locations)} 个储位, 模拟模式: {is_sim}")

        # 检查任务是否已存在
        if task_no in inventory_tasks:
            existing_task = inventory_tasks[task_no]
            if existing_task.status in ["running", "init"]:
                return JSONResponse(
                    status_code=status.HTTP_200_OK,
                    content={
                        "code": 200,
                        "message": "任务已在执行中",
                        "data": {
                            "taskNo": existing_task.task_no,
                            "status": existing_task.status,
                        }
                    }
                )

        # 保存原始盘点项信息，用于生成Excel
        if task_no not in inventory_task_details:
            inventory_task_details[task_no] = {}
        inventory_task_details[task_no]["inventoryItems"] = inventory_items

        # 记录盘点任务启动
        auth_token = request.headers.get("authToken")
        user_info = await get_user_info_from_token(auth_token) if auth_token else {}

        log_operation(
            operation_type="inventory",
            action="启动盘点任务",
            user_id=user_info.get("userId"),
            user_name=user_info.get("userName"),
            target=task_no,
            status="running",
            details={
                "task_no": task_no,
                "bin_locations": bin_locations,
                "bin_count": len(bin_locations),
                "is_sim": is_sim
            }
        )

        # 在后台异步执行盘点任务
        background_tasks.add_task(
            execute_inventory_workflow,
            task_no=task_no,
            bin_locations=bin_locations,
            is_sim=is_sim
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "盘点任务已启动",
                "data": {
                    "taskNo": task_no,
                    "bin_locations": bin_locations,
                    "is_sim": is_sim
                }
            }
        )

    except Exception as e:
        logger.error(f"启动盘点任务失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"启动盘点任务失败: {str(e)}"
        )


@app.get("/api/inventory/progress")
async def get_inventory_progress(taskNo: str):
    """获取盘点任务进度"""
    try:
        if taskNo not in inventory_tasks:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "任务不存在",
                    "data": None
                }
            )

        task_status = inventory_tasks[taskNo]
        bin_statuses = inventory_task_bins.get(taskNo, [])

        # 计算进度百分比
        completed_count = sum(
            1 for bin in bin_statuses if bin.status == "completed")
        progress_percentage = (
            completed_count / task_status.total_steps * 100) if task_status.total_steps > 0 else 0

        progress_data = InventoryTaskProgress(
            task_no=task_status.task_no,
            status=task_status.status,
            current_step=task_status.current_step,
            total_steps=task_status.total_steps,
            progress_percentage=round(progress_percentage, 2),
            bin_locations=bin_statuses,
            start_time=task_status.start_time,
            end_time=task_status.end_time
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取进度成功",
                "data": progress_data.dict()
            }
        )
    except Exception as e:
        logger.error(f"获取任务进度失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取任务进度失败: {str(e)}"
        )


@app.get("/api/inventory/results")
async def get_inventory_results(taskNo: str):
    """获取盘点任务的完整结果"""
    try:
        if taskNo not in inventory_tasks:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "任务不存在",
                    "data": None
                }
            )

        task_status = inventory_tasks[taskNo]

        # 检查任务是否完成
        if task_status.status != "completed":
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "任务尚未完成",
                    "data": {
                        "taskNo": taskNo,
                        "status": task_status.status,
                        "currentStep": task_status.current_step,
                        "totalSteps": task_status.total_steps,
                        "inventoryResults": []
                    }
                }
            )

        # 获取盘点结果
        inventory_results = []
        if taskNo in inventory_task_details and "inventoryResults" in inventory_task_details[taskNo]:
            inventory_results = inventory_task_details[taskNo]["inventoryResults"]

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取盘点结果成功",
                "data": {
                    "taskNo": taskNo,
                    "status": task_status.status,
                    "inventoryResults": inventory_results
                }
            }
        )
    except Exception as e:
        logger.error(f"获取盘点结果失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取盘点结果失败: {str(e)}"
        )


@app.options("/api/inventory/save-results")
async def options_save_inventory_results():
    """OPTIONS 预检请求"""
    return Response(
        status_code=200,
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "POST, OPTIONS",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Max-Age": "3600"
        }
    )


@app.get("/api/inventory/image")
async def get_inventory_image(
    taskNo: str,
    binLocation: str,
    cameraType: str,
    filename: str,
    source: str = "output"  # 新增参数：source可以是"output"或"capture_img"
):
    """
    获取盘点任务中的图片

    Args:
        taskNo: 任务编号
        binLocation: 储位名称
        cameraType: 相机类型
        filename: 文件名
        source: 图片源目录，可以是"output"或"capture_img"，默认为"output"
    """
    try:
        project_root = Path(__file__).parent.parent.parent

        # 根据source参数选择不同的基础目录
        if source == "capture_img":
            # 从capture_img目录获取图片
            image_path = project_root / "capture_img" / \
                taskNo / binLocation / cameraType / filename
        else:
            # 从output目录获取图片（默认）
            image_path = project_root / "output" / \
                taskNo / binLocation / cameraType / filename

        if not image_path.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"图片不存在: {filename} (路径: {image_path})"
            )

        # 读取图片文件
        with open(image_path, "rb") as f:
            image_data = f.read()

        # 根据文件扩展名确定媒体类型
        media_type = "image/jpeg"
        if filename.endswith(".png"):
            media_type = "image/png"
        elif filename.endswith(".bmp"):
            media_type = "image/bmp"

        return Response(content=image_data, media_type=media_type)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取图片失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取图片失败: {str(e)}"
        )


@app.get("/api/inventory/task-detail")
async def get_task_detail(taskNo: str, binLocation: str):
    """
    获取任务的详细信息，包括图片和计算结果

    Args:
        taskNo: 任务编号
        binLocation: 储位名称
    """
    try:
        # 从任务详情中获取（使用存储函数）
        storage = get_task_state_storage()
        task_details = storage["details"]

        if taskNo in task_details and binLocation in task_details[taskNo]:
            detail = task_details[taskNo][binLocation]
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "获取任务详情成功",
                    "data": detail
                }
            )
        else:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "任务详情不存在",
                    "data": None
                }
            )

    except Exception as e:
        logger.error(f"获取任务详情失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取任务详情失败: {str(e)}"
        )


@app.post("/api/inventory/scan-and-recognize")
async def scan_and_recognize(request: ScanAndRecognizeRequest = Body(...)):
    """
    扫码+识别接口
    同时执行Detect模块和Barcode模块的识别

    Args:
        request: ScanAndRecognizeRequest对象，包含taskNo、binLocation、pile_id和code_type
    """
    try:
        project_root = Path(__file__).parent.parent.parent

        # 构建图片路径: taskNo/binLocation/3d_camera/
        image_path = f"{request.taskNo}/{request.binLocation}/3d_camera/"
        image_dir = project_root / "capture_img" / image_path

        recognition_time = datetime.now().isoformat()
        results = {
            "taskNo": request.taskNo,
            "binLocation": request.binLocation,
            "recognition_time": recognition_time,
            "detect_result": None,
            "barcode_result": None,
            "photos": []  # 添加照片路径数组
        }

        # 查找并返回照片路径（无论检测模块是否可用）
        try:
            photos = []
            image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']

            # 优先查找主图 (3d_camera 目录)
            main_images = []
            if image_dir.exists():
                for ext in image_extensions:
                    # 查找常见的文件名
                    for common_name in ['main', 'MAIN', 'raw', 'RAW', 'image', 'IMAGE']:
                        main_file = image_dir / f"{common_name}{ext}"
                        if main_file.exists():
                            # 构建相对于 capture_img 的路径
                            relative_path = f"{request.taskNo}/{request.binLocation}/3d_camera/{common_name.upper()}{ext}"
                            main_images.append(f"/{relative_path}")
                            logger.info(f"找到主图: {relative_path}")
                            break
                    if main_images:
                        break

                # 如果没有找到常见文件名，查找第一张图片
                if not main_images:
                    for ext in image_extensions:
                        image_files = list(image_dir.glob(f"*{ext}"))
                        if image_files:
                            main_file = image_files[0]
                            relative_path = f"{request.taskNo}/{request.binLocation}/3d_camera/{main_file.name}"
                            photos.append(f"/{relative_path}")
                            logger.info(f"找到主图: {relative_path}")
                            break

                photos.extend(main_images)

                # 查找深度图 (3d_camera 目录，与主图在同一目录)
                depth_dir = project_root / "capture_img" / \
                    f"{request.taskNo}/{request.binLocation}/3d_camera"
                if depth_dir.exists():
                    for ext in image_extensions:
                        depth_file = depth_dir / f"depth{ext}"
                        if depth_file.exists():
                            relative_path = f"{request.taskNo}/{request.binLocation}/3d_camera/depth{ext}"
                            photos.append(f"/{relative_path}")
                            logger.info(f"找到深度图: {relative_path}")
                            break

            results["photos"] = photos
            logger.info(f"共找到 {len(photos)} 张图片")

        except Exception as e:
            logger.error(f"查找照片路径失败: {str(e)}", exc_info=True)
            results["photos"] = []  # 失败时返回空数组，不中断流程

        # ========== Step 1: 先执行 Barcode 模块获取垛型信息 ==========
        # 用于存储从条码识别中获取的 pile_id
        detected_pile_id = request.pile_id  # 默认使用前端传来的值

        if ENABLE_BARCODE and BARCODE_MODULE_AVAILABLE:
            try:
                logger.info(
                    f"开始执行Barcode模块识别（获取垛型）: {request.taskNo}/{request.binLocation}")

                # 使用 BarcodeRecognizer 识别条码
                recognizer = BarcodeRecognizer(code_type=request.code_type)
                barcode_results = recognizer.process_folder(input_dir=str(image_dir))

                # 使用 TobaccoCaseInfoResolver 解析条码获取垛型
                resolver = get_tobacco_case_resolver()

                # 遍历识别结果，找到第一个有效的烟箱信息
                resolved_info = None
                for result in barcode_results:
                    barcode_text = result.get('output') or result.get('text')
                    if barcode_text:
                        resolved_info = resolver.resolve(barcode_text)
                        if resolved_info['success']:
                            logger.info(f"成功解析烟箱信息: {resolved_info}")
                            break

                if resolved_info and resolved_info['success']:
                    # 使用解析出的 pile_id
                    detected_pile_id = resolved_info['pile_id']
                    logger.info(f"映射后的 pile_id: {detected_pile_id} (原请求: {request.pile_id})")

                    results["barcode_result"] = {
                        "image_path": str(image_dir),
                        "code_type": request.code_type,
                        "six_digit_code": resolved_info['six_digit_code'],
                        "stack_type_1": resolved_info['stack_type_1'],
                        "product_name": resolved_info['product_name'],
                        "tobacco_code": resolved_info['tobacco_code'],
                        "mapped_pile_id": detected_pile_id,
                        "total_images": len(barcode_results),
                        "status": "success"
                    }
                else:
                    # 条码识别成功但未匹配到烟箱信息
                    logger.warning("条码识别成功但未匹配到烟箱信息")
                    results["barcode_result"] = {
                        "image_path": str(image_dir),
                        "code_type": request.code_type,
                        "results": barcode_results,
                        "total_images": len(barcode_results),
                        "successful": sum(1 for r in barcode_results if r.get("output")),
                        "failed": sum(1 for r in barcode_results if r.get("error") and not r.get("output")),
                        "status": "no_match",
                        "message": "条码识别成功但未匹配到烟箱信息"
                    }

            except Exception as e:
                logger.error(f"Barcode模块识别失败: {str(e)}", exc_info=True)
                results["barcode_result"] = {
                    "status": "failed",
                    "error": str(e)
                }
        else:
            logger.info("Barcode模块已禁用，跳过识别")
            results["barcode_result"] = {
                "status": "disabled",
                "message": "条形码功能已禁用（ENABLE_BARCODE=false）"
            }

        # ========== Step 2: 执行 Detect 模块（使用映射后的 pile_id） ==========
        # 如果检测模块可用，执行识别
        if DETECT_MODULE_AVAILABLE:
            # 严格检查 capture_img 下的路径是否存在
            if not image_dir.exists() or not image_dir.is_dir():
                logger.error(f"图片目录不存在: {image_dir}")
                # 即使目录不存在，也返回照片路径（为空）
                return JSONResponse(
                    status_code=status.HTTP_200_OK,
                    content={
                        "code": 200,
                        "message": "图片目录不存在，但返回了照片路径信息",
                        "data": results
                    }
                )

            try:
                logger.info(
                    f"开始执行Detect模块识别: {request.taskNo}/{request.binLocation}, pile_id={detected_pile_id}")

                # 查找目录中的图片文件
                image_extensions = ['.jpg', '.jpeg', '.png', '.bmp']
                image_files = []

                # 优先查找常见文件名
                common_names = ['main', 'raw', 'image', 'img', 'photo']
                for name in common_names:
                    for ext in image_extensions:
                        common_file = image_dir / f"{name}{ext}"
                        if common_file.exists():
                            image_files.append(common_file)
                            break
                    if image_files:
                        break

                # 如果没找到，查找所有图片
                if not image_files:
                    for ext in image_extensions:
                        image_files.extend(list(image_dir.glob(f"*{ext}")))
                        if image_files:
                            break

                if not image_files:
                    raise HTTPException(
                        status_code=status.HTTP_404_NOT_FOUND,
                        detail=f"在路径 {image_dir} 中未找到图片文件"
                    )

                image_path_for_detect = str(image_files[0])

                # 从输入路径加载深度图 depth.jpg
                depth_image_path = image_dir / "depth.jpg"
                depth_image_path_for_detect = None
                if depth_image_path.exists():
                    depth_image_path_for_detect = str(depth_image_path)
                    logger.info(f"找到深度图文件: {depth_image_path_for_detect}")
                else:
                    logger.warning(f"深度图文件不存在: {depth_image_path}，将不使用深度图")

                # 构建debug输出目录: debug/{taskNo}/{binLocation}/
                debug_output_dir = project_root / "debug" / request.taskNo / request.binLocation
                debug_output_dir.mkdir(parents=True, exist_ok=True)

                # 创建日志文件处理器，将算法日志保存到debug目录
                log_file_path = debug_output_dir / \
                    f"debug_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
                file_handler = logging.FileHandler(
                    str(log_file_path), encoding='utf-8')
                file_handler.setLevel(logging.DEBUG)
                file_formatter = logging.Formatter(
                    '%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S'
                )
                file_handler.setFormatter(file_formatter)

                # 添加文件处理器到logger（临时添加，执行完后移除）
                logger.addHandler(file_handler)

                try:
                    logger.info(f"Debug输出目录: {debug_output_dir}")
                    logger.info(f"Debug日志文件: {log_file_path}")
                    logger.info(f"使用主图片: {image_path_for_detect}")
                    if depth_image_path_for_detect:
                        logger.info(f"使用深度图: {depth_image_path_for_detect}")

                    # 重定向stdout和stderr到日志文件，捕获count_boxes的print输出
                    import io
                    from contextlib import redirect_stdout, redirect_stderr

                    # 创建一个类，将写入操作转换为日志记录
                    class LogWriter:
                        def __init__(self, logger_instance, level=logging.INFO):
                            self.logger = logger_instance
                            self.level = level
                            self.buffer = ""

                        def write(self, message):
                            if message.strip():  # 忽略空行
                                self.buffer += message
                                # 检查是否有完整的行
                                while '\n' in self.buffer:
                                    line, self.buffer = self.buffer.split(
                                        '\n', 1)
                                    if line.strip():
                                        self.logger.log(
                                            self.level, line.strip())

                        def flush(self):
                            if self.buffer.strip():
                                self.logger.log(
                                    self.level, self.buffer.strip())
                                self.buffer = ""

                    log_writer = LogWriter(logger, logging.INFO)

                    # 执行count_boxes，使用从条码识别映射后的 pile_id
                    with redirect_stdout(log_writer), redirect_stderr(log_writer):
                        total_count = count_boxes(
                            image_path=image_path_for_detect,
                            pile_id=detected_pile_id,  # 使用映射后的 pile_id
                            depth_image_path=depth_image_path_for_detect,
                            enable_debug=False,
                            enable_visualization=False,
                            output_dir=str(debug_output_dir)
                        )

                    logger.info(f"Detect模块识别完成，箱数: {total_count}")

                    results["detect_result"] = {
                        "image_path": image_path_for_detect,
                        "pile_id": detected_pile_id,  # 记录实际使用的 pile_id
                        "total_count": total_count,
                        "status": "success"
                    }

                except Exception as e:
                    logger.error(f"Detect模块识别失败: {str(e)}", exc_info=True)
                    results["detect_result"] = {
                        "status": "failed",
                        "error": str(e)
                    }
                finally:
                    # 移除文件处理器，避免日志重复输出
                    logger.removeHandler(file_handler)
                    file_handler.close()

            except HTTPException:
                raise
            except Exception as e:
                logger.error(f"Detect模块初始化失败: {str(e)}", exc_info=True)
                results["detect_result"] = {
                    "status": "failed",
                    "error": str(e)
                }

        # 3. 更新任务状态中的识别结果
        storage = get_task_state_storage()
        inventory_task_bins = storage["bins"]
        inventory_task_details = storage["details"]

        if request.taskNo in inventory_task_bins:
            for bin_status in inventory_task_bins[request.taskNo]:
                if bin_status.bin_location == request.binLocation:
                    bin_status.detect_result = results["detect_result"]
                    bin_status.barcode_result = results["barcode_result"]
                    bin_status.recognition_time = recognition_time
                    break

        # 4. 存储到任务详情中
        if request.taskNo not in inventory_task_details:
            inventory_task_details[request.taskNo] = {}

        if request.binLocation not in inventory_task_details[request.taskNo]:
            inventory_task_details[request.taskNo][request.binLocation] = {}

        inventory_task_details[request.taskNo][request.binLocation]["recognition"] = results

        # 5. 将识别结果保存为ini格式文件到原始图路径
        try:
            import configparser

            # 创建ConfigParser对象
            config = configparser.ConfigParser()

            # 添加基本信息节
            config.add_section('BasicInfo')
            config.set('BasicInfo', 'taskNo', request.taskNo)
            config.set('BasicInfo', 'binLocation', request.binLocation)
            config.set('BasicInfo', 'recognition_time', recognition_time)

            # 添加原始图路径
            config.add_section('ImagePaths')
            config.set('ImagePaths', 'original_image_dir', str(image_dir))
            if results.get("detect_result") and results["detect_result"].get("image_path"):
                config.set('ImagePaths', 'detect_image_path',
                           results["detect_result"]["image_path"])

            # 添加Detect模块结果
            if results.get("detect_result"):
                config.add_section('DetectResult')
                detect_result = results["detect_result"]
                if detect_result.get("status"):
                    config.set('DetectResult', 'status',
                               detect_result["status"])
                if detect_result.get("total_count") is not None:
                    config.set('DetectResult', 'total_count',
                               str(detect_result["total_count"]))
                if detect_result.get("pile_id") is not None:
                    config.set('DetectResult', 'pile_id',
                               str(detect_result["pile_id"]))
                if detect_result.get("error"):
                    config.set('DetectResult', 'error', detect_result["error"])

            # 添加Barcode模块结果
            if results.get("barcode_result"):
                config.add_section('BarcodeResult')
                barcode_result = results["barcode_result"]
                if barcode_result.get("status"):
                    config.set('BarcodeResult', 'status',
                               barcode_result["status"])
                if barcode_result.get("code_type"):
                    config.set('BarcodeResult', 'code_type',
                               barcode_result["code_type"])
                if barcode_result.get("total_images") is not None:
                    config.set('BarcodeResult', 'total_images',
                               str(barcode_result["total_images"]))
                if barcode_result.get("successful") is not None:
                    config.set('BarcodeResult', 'successful',
                               str(barcode_result["successful"]))
                if barcode_result.get("failed") is not None:
                    config.set('BarcodeResult', 'failed',
                               str(barcode_result["failed"]))
                if barcode_result.get("error"):
                    config.set('BarcodeResult', 'error',
                               barcode_result["error"])
                if barcode_result.get("message"):
                    config.set('BarcodeResult', 'message',
                               barcode_result["message"])

            # 保存ini文件到原始图路径
            ini_file_path = image_dir / "recognition_result.ini"
            with open(ini_file_path, 'w', encoding='utf-8') as configfile:
                config.write(configfile)

            logger.info(f"识别结果已保存到ini文件: {ini_file_path}")
        except Exception as e:
            logger.error(f"保存识别结果到ini文件失败: {str(e)}", exc_info=True)
            # 不中断流程，继续返回结果

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "扫码+识别执行完成",
                "data": results
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"扫码+识别失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"扫码+识别失败: {str(e)}"
        )


@app.get("/api/inventory/recognition-result")
async def get_recognition_result(taskNo: str, binLocation: str):
    """
    读取识别结果接口
    获取指定任务和库位的Detect和Barcode识别结果

    Args:
        taskNo: 任务编号
        binLocation: 库位号
    """
    try:
        # 获取任务状态存储
        storage = get_task_state_storage()
        inventory_task_bins = storage["bins"]
        inventory_task_details = storage["details"]

        result_data = {
            "taskNo": taskNo,
            "binLocation": binLocation,
            "detect_result": None,
            "barcode_result": None,
            "recognition_time": None
        }

        # 从任务状态中获取（使用存储函数）
        storage = get_task_state_storage()
        task_bins = storage["bins"]
        task_details = storage["details"]

        if taskNo in task_bins:
            for bin_status in task_bins[taskNo]:
                if bin_status.bin_location == binLocation:
                    result_data["detect_result"] = bin_status.detect_result
                    result_data["barcode_result"] = bin_status.barcode_result
                    result_data["recognition_time"] = bin_status.recognition_time
                    break

        # 如果状态中没有，尝试从任务详情中获取
        if not result_data["detect_result"] and taskNo in task_details:
            if binLocation in task_details[taskNo]:
                recognition_data = task_details[taskNo][binLocation].get(
                    "recognition")
                if recognition_data:
                    result_data["detect_result"] = recognition_data.get(
                        "detect_result")
                    result_data["barcode_result"] = recognition_data.get(
                        "barcode_result")
                    result_data["recognition_time"] = recognition_data.get(
                        "recognition_time")

        # 检查是否有结果
        if not result_data["detect_result"] and not result_data["barcode_result"]:
            return JSONResponse(
                status_code=status.HTTP_404_NOT_FOUND,
                content={
                    "code": 404,
                    "message": "识别结果不存在",
                    "data": None
                }
            )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取识别结果成功",
                "data": result_data
            }
        )

    except Exception as e:
        error_msg = f"获取识别结果失败: {str(e)}"
        logger.error(error_msg, exc_info=True)  # exc_info=True 会输出完整的堆栈信息
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=error_msg
        )


######################################### LMS #########################################


@app.get("/auth/token")
async def auth_token(token: str):
    """处理前端获取用户信息请求，调用LMS的authToken接口"""
    try:
        # 调用LMS的authToken接口
        lms_auth_url = f"{LMS_BASE_URL}/auth/token?token={token}"
        response = requests.get(lms_auth_url)

        if response.status_code == 200:
            return response.json()
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS获取用户信息失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"获取用户信息请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="获取用户信息请求处理失败"
        )


@app.get("/lms/getLmsBin")
async def get_lms_bin(authToken: str):
    """获取库位信息，调用LMS的getLmsBin接口"""
    try:
        # 调用LMS的getLmsBin接口
        lms_bin_url = f"{LMS_BASE_URL}/third/api/v1/lmsToRcsService/getLmsBin"
        headers = {
            "authToken": authToken
        }
        response = requests.get(lms_bin_url, headers=headers)

        if response.status_code == 200:
            # 关键修复：处理LMS返回的压缩编码字符串
            try:
                uncompressed_data = custom_utils.decompress_and_decode(
                    response.text)

                logger.info("成功解压缩并解析库位数据")
                return JSONResponse(uncompressed_data)
            except Exception as e:
                logger.error(f"解压缩库位数据失败: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="库位数据解压缩失败"
                )
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS获取库位信息失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"获取库位信息请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="获取库位信息请求处理失败"
        )


@app.get("/lms/getCountTasks")
async def get_count_tasks(authToken: str):
    """获取盘点任务，调用LMS的getCountTasks接口"""
    try:
        logger.info(f"收到获取盘点任务请求，authToken: {authToken[:20]}...")

        lms_tasks_url = f"{LMS_BASE_URL}/third/api/v1/lmsToRcsService/getCountTasks"
        logger.info(f"准备调用LMS接口: {lms_tasks_url}")

        headers = {"authToken": authToken}
        logger.info("发送请求到LMS服务...")
        response = requests.get(lms_tasks_url, headers=headers, timeout=30)
        logger.info(f"LMS响应状态码: {response.status_code}")

        if response.status_code == 200:
            # 关键修复：处理LMS返回的压缩编码字符串
            try:
                uncompressed_data = custom_utils.decompress_and_decode(
                    response.text)

                logger.info("成功解压缩并解析盘点任务数据")
                return JSONResponse(uncompressed_data)
            except Exception as e:
                logger.error(f"解压缩盘点任务数据失败: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="盘点任务数据解压缩失败"
                )
        else:
            logger.error(
                f"LMS获取盘点任务失败: {response.status_code} - {response.text}")
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS获取盘点任务失败: {response.text}"
            )
    except requests.exceptions.Timeout:
        logger.error("LMS服务请求超时")
        raise HTTPException(
            status_code=status.HTTP_504_GATEWAY_TIMEOUT,
            detail="LMS服务响应超时"
        )
    except requests.exceptions.ConnectionError:
        logger.error("无法连接到LMS服务")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="无法连接到LMS服务"
        )
    except Exception as e:
        logger.error(f"获取盘点任务请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="获取盘点任务请求处理失败"
        )


@app.post("/lms/setTaskResults")
async def set_task_results(request: Request):
    """提交盘点任务结果，调用LMS的setTaskResults接口"""
    try:
        # 1. 从请求头获取authToken
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Unauthorized"
            )

        # 2. 从请求体获取JSON数据（前端发送的是标准JSON）
        data = await request.json()
        encoded_data = custom_utils.compress_and_encode(data)

        # 6. 调用LMS接口（使用压缩后的数据）
        lms_results_url = f"{LMS_BASE_URL}/third/api/v1/RcsToLmsService/setTaskResults"
        headers = {
            "authToken": auth_token,  # 传递给LMS的认证令牌
            "Content-Type": "text/plain"  # 关键：必须是text/plain
        }

        # 发送压缩后的base64字符串
        response = requests.post(
            lms_results_url, data=encoded_data, headers=headers)

        if response.status_code == 200:
            return {"success": True, "message": "盘点结果已提交"}
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS提交盘点结果失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"提交盘点结果请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="提交盘点结果请求处理失败"
        )


######################################### RCS #########################################

@app.post("/api/robot/reporter/task")
async def task_status(request: Request):
    try:
        # 获取请求数据
        request_data = await request.json()

        logger.info("反馈任务状态")
        logger.info(
            f"请求数据: {json.dumps(request_data, indent=2, ensure_ascii=False)}")

        # 提取任务信息
        robot_task_code = request_data.get("robotTaskCode")
        single_robot_code = request_data.get("singleRobotCode")
        extra = request_data.get("extra", "")

        # 解析extra字段
        if extra:
            try:
                extra_list = json.loads(extra)
                if isinstance(extra_list, list):
                    for item in extra_list:
                        method = item.get("method", "")
                        logger.info(f"处理method: {method}")
                        await update_robot_status(method, item)

                        if method == "start":
                            logger.info("任务开始")

                        elif method == "outbin":
                            logger.info("走出储位")

                        elif method == "end":
                            logger.info("任务完成")

                        # 根据不同的method更新您的任务状态...
            except json.JSONDecodeError:
                logger.error(f"无法解析extra字段: {extra}")

        # 返回响应
        return {
            "code": "SUCCESS",
            "message": "成功",
            "data": {
                "robotTaskCode": "ctu001"
            }
        }

    except Exception as e:
        logger.error(f"处理状态反馈失败: {str(e)}")
        raise HTTPException(status_code=500, detail=f"处理状态反馈失败: {str(e)}")


# 前端日志收集接口
@app.post("/api/log/frontend")
async def collect_frontend_log(request: FrontendLogRequest = Body(...)):
    """
    收集前端日志并保存到 debug 目录

    Args:
        request: 前端日志请求对象
    """
    try:
        # 创建前端日志文件路径（按日期命名）
        frontend_log_filename = _debug_log_dir / \
            f"frontend_{datetime.now().strftime('%Y%m%d')}.log"

        # 创建前端日志记录器
        frontend_logger = logging.getLogger("frontend")
        frontend_logger.setLevel(logging.DEBUG)

        # 如果还没有文件处理器，添加一个
        if not any(isinstance(h, logging.FileHandler) and h.baseFilename == str(frontend_log_filename)
                   for h in frontend_logger.handlers):
            file_handler = logging.FileHandler(
                str(frontend_log_filename), encoding='utf-8')
            file_handler.setLevel(logging.DEBUG)
            formatter = logging.Formatter(
                '%(asctime)s - [FRONTEND] - %(levelname)s - %(message)s',
                datefmt='%Y-%m-%d %H:%M:%S'
            )
            file_handler.setFormatter(formatter)
            frontend_logger.addHandler(file_handler)

        # 记录日志
        timestamp = request.timestamp or datetime.now().isoformat()
        source = request.source or "unknown"
        extra_info = request.extra or {}

        log_message = f"[{source}] {request.message}"
        if extra_info:
            log_message += f" | Extra: {json.dumps(extra_info, ensure_ascii=False)}"

        # 根据日志级别记录
        log_level = request.level.lower()
        if log_level == "error":
            frontend_logger.error(log_message)
        elif log_level == "warn":
            frontend_logger.warning(log_message)
        elif log_level == "info":
            frontend_logger.info(log_message)
        else:
            frontend_logger.debug(log_message)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "日志已保存",
                "data": {
                    "log_file": str(frontend_log_filename)
                }
            }
        )

    except Exception as e:
        logger.error(f"保存前端日志失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"保存前端日志失败: {str(e)}"
        )


######################################### 用户管理接口 #########################################

@app.get("/lms/getUsers")
async def get_users(request: Request):
    """获取所有用户信息，调用LMS的getUsers接口"""
    try:
        # 从查询参数或请求头获取authToken
        auth_token = request.query_params.get(
            'authToken') or request.headers.get('authToken')

        if not auth_token:
            logger.error("未提供认证令牌")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="未提供认证令牌"
            )

        logger.info(f"接收到的authToken: {auth_token[:20]}...")

        # 调用LMS的getUsers接口
        lms_users_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/getUsers"
        headers = {
            "authToken": auth_token
        }

        logger.info(f"调用LMS用户接口: {lms_users_url}")

        try:
            response = requests.get(lms_users_url, headers=headers, timeout=10)
            logger.info(f"LMS响应状态码: {response.status_code}")

            # 检查响应状态
            if response.status_code == 200:
                # 尝试解析JSON响应
                try:
                    result = response.json()
                    logger.info(f"成功获取 {len(result.get('data', []))} 条用户数据")
                    return result
                except json.JSONDecodeError as e:
                    logger.error(f"LMS返回的不是有效JSON: {response.text[:200]}")
                    raise HTTPException(
                        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                        detail=f"LMS返回了无效的JSON数据: {str(e)}"
                    )
            else:
                # 记录详细的错误信息
                error_detail = response.text[:500]  # 只取前500字符避免日志过大
                logger.error(f"LMS接口错误 {response.status_code}: {error_detail}")

                if response.status_code == 401:
                    return JSONResponse(
                        status_code=status.HTTP_401_UNAUTHORIZED,
                        content={
                            "code": 401,
                            "message": "认证失败，请重新登录",
                            "data": None
                        }
                    )
                elif response.status_code == 403:
                    return JSONResponse(
                        status_code=status.HTTP_403_FORBIDDEN,
                        content={
                            "code": 403,
                            "message": "权限不足，只有管理员可以查看用户列表",
                            "data": None
                        }
                    )
                else:
                    return JSONResponse(
                        status_code=response.status_code,
                        content={
                            "code": response.status_code,
                            "message": f"LMS获取用户信息失败: {response.text[:200]}",
                            "data": None
                        }
                    )

        except requests.exceptions.ConnectionError:
            logger.error(f"无法连接到LMS服务: {LMS_BASE_URL}")
            return JSONResponse(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                content={
                    "code": 503,
                    "message": f"无法连接到LMS服务，请确保LMS服务正在运行（{LMS_BASE_URL}）",
                    "data": None
                }
            )
        except requests.exceptions.Timeout:
            logger.error(f"连接LMS服务超时")
            return JSONResponse(
                status_code=status.HTTP_504_GATEWAY_TIMEOUT,
                content={
                    "code": 504,
                    "message": "LMS服务响应超时",
                    "data": None
                }
            )

    except HTTPException as he:
        # 重新抛出 HTTPException
        raise he
    except Exception as e:
        logger.error(f"获取用户信息请求失败: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取用户信息请求处理失败: {str(e)}"
        )


@app.post("/lms/registerUser")
async def register_user(request: Request):
    """注册新用户，调用LMS的registerUser接口"""
    try:
        # 1. 从请求头获取authToken
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Unauthorized"
            )

        # 获取操作用户信息
        user_info = await get_user_info_from_token(auth_token)

        # 2. 获取请求体数据
        data = await request.json()
        new_user_name = data.get("userName", "未知用户")

        # 3. 调用LMS接口
        lms_register_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/registerUser"
        headers = {
            "authToken": auth_token,
            "Content-Type": "application/json"
        }

        response = requests.post(lms_register_url, json=data, headers=headers)

        if response.status_code == 200:
            # 记录用户添加操作
            log_operation(
                operation_type="user_management",
                action="添加用户",
                user_id=user_info.get("userId"),
                user_name=user_info.get("userName"),
                target=new_user_name,
                status="success",
                details={
                    "new_user_data": data,
                    "response": response.json()
                }
            )

            return response.json()
        else:
            # 记录操作失败
            log_operation(
                operation_type="user_management",
                action="添加用户",
                user_id=user_info.get("userId"),
                user_name=user_info.get("userName"),
                target=new_user_name,
                status="failed",
                details={
                    "new_user_data": data,
                    "error": response.text[:200],
                    "status_code": response.status_code
                }
            )

            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS注册用户失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"注册用户请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="注册用户请求处理失败"
        )


@app.post("/lms/deleteUser")
async def delete_user(request: Request):
    """删除用户，调用LMS的deleteUser接口"""
    try:
        # 1. 从请求头获取authToken
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="Unauthorized"
            )

        # 2. 获取请求体数据
        data = await request.json()

        # 3. 调用LMS接口
        lms_delete_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/deleteUser"
        headers = {
            "authToken": auth_token,
            "Content-Type": "application/json"
        }

        response = requests.post(lms_delete_url, json=data, headers=headers)

        if response.status_code == 200:
            return response.json()
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS删除用户失败: {response.text}"
            )
    except Exception as e:
        logger.error(f"删除用户请求失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="删除用户请求处理失败"
        )


@app.get("/api/adminContact")
async def get_admin_contact():
    """
    获取管理员联系方式
    返回运维人员的姓名和电话信息
    """
    try:
        # 这里返回固定的管理员信息，也可以从数据库或配置文件中读取
        contact_info = {
            "adminName": "运维管理员",
            "adminPhone": "13800138000",
            "adminEmail": "admin@example.com",
            "note": "如忘记密码，请联系管理员协助重置"
        }

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取管理员联系方式成功",
                "data": contact_info
            }
        )
    except Exception as e:
        logger.error(f"获取管理员联系方式失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取管理员联系方式失败: {str(e)}"
        )


######################################### 操作日志 #########################################
# 操作日志文件路径
IMG_ROOT = Path(__file__).parent.parent.parent / "capture_img"

OUTPUT_ROOT = Path(__file__).parent.parent.parent / "output"
OPERATION_LOG_FILE = OUTPUT_ROOT / "operation_logs.json"
MAX_LOGS = 100  # 最多保留100条记录


def get_operation_logs_from_file() -> List[Dict[str, Any]]:
    """从文件读取操作日志"""
    try:
        if not OPERATION_LOG_FILE.exists():
            return []
        with open(OPERATION_LOG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        logger.error(f"读取操作日志文件失败: {str(e)}")
        return []


def save_operation_logs_to_file(logs: List[Dict[str, Any]]) -> bool:
    """保存操作日志到文件"""
    try:
        # 确保目录存在
        OPERATION_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
        with open(OPERATION_LOG_FILE, 'w', encoding='utf-8') as f:
            json.dump(logs, f, ensure_ascii=False, indent=2)
        return True
    except Exception as e:
        logger.error(f"保存操作日志文件失败: {str(e)}")
        return False


@app.get("/api/operationLogs")
async def get_operation_logs(limit: Optional[int] = None):
    """
    获取操作日志列表
    Args:
        limit: 可选，限制返回的记录数量
    """
    try:
        logs = get_operation_logs_from_file()

        # 如果指定了限制数量，只返回前N条
        if limit is not None and limit > 0:
            logs = logs[:limit]

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取操作日志成功",
                "data": {
                    "logs": logs,
                    "total": len(logs)
                }
            }
        )
    except Exception as e:
        logger.error(f"获取操作日志失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取操作日志失败: {str(e)}"
        )


@app.post("/api/operationLogs")
async def add_operation_log(request: Request):
    """
    添加操作日志
    """
    try:
        data = await request.json()

        # 生成ID和时间戳
        import uuid
        log_id = f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:9]}"
        timestamp = datetime.now().isoformat()

        # 构建日志对象
        new_log = {
            "id": log_id,
            "timestamp": timestamp,
            "operation_type": data.get("operation_type", "unknown"),
            "user_id": data.get("user_id"),
            "user_name": data.get("user_name"),
            "action": data.get("action", ""),
            "target": data.get("target"),
            "status": data.get("status", "success"),
            "details": data.get("details", {}),
            "ip_address": data.get("ip_address")
        }

        # 读取现有日志
        logs = get_operation_logs_from_file()

        # 添加新日志到开头
        logs.insert(0, new_log)

        # 限制日志数量
        logs = logs[:MAX_LOGS]

        # 保存到文件
        if save_operation_logs_to_file(logs):
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "添加操作日志成功",
                    "data": new_log
                }
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="保存操作日志失败"
            )
    except Exception as e:
        logger.error(f"添加操作日志失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"添加操作日志失败: {str(e)}"
        )


@app.delete("/api/operationLogs")
async def clear_operation_logs():
    """
    清空所有操作日志
    """
    try:
        if save_operation_logs_to_file([]):
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "清空操作日志成功"
                }
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="清空操作日志失败"
            )
    except Exception as e:
        logger.error(f"清空操作日志失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"清空操作日志失败: {str(e)}"
        )


@app.post("/api/operationLogs/autoCleanup")
async def auto_cleanup_logs():
    """
    自动清理操作日志（系统自动调用）
    每月第一天自动清理六个月前的数据
    也可以手动调用此接口触发
    """
    try:
        # 计算6个月前的日期
        cutoff_date = datetime.now() - timedelta(days=180)
        cutoff_date_str = cutoff_date.strftime("%Y-%m-%d")

        # 读取现有日志
        logs = get_operation_logs_from_file()

        # 过滤掉超过6个月的日志
        filtered_logs = []
        cleaned_count = 0

        for log in logs:
            try:
                log_date = datetime.fromisoformat(log.get("timestamp", ""))
                if log_date >= cutoff_date:
                    filtered_logs.append(log)
                else:
                    cleaned_count += 1
            except (ValueError, TypeError):
                # 如果无法解析日期，保留该日志
                filtered_logs.append(log)

        # 保存过滤后的日志
        if save_operation_logs_to_file(filtered_logs):
            # 添加自动清理日志记录
            auto_cleanup_log = {
                "id": f"{int(time.time() * 1000)}_{uuid.uuid4().hex[:9]}",
                "timestamp": datetime.now().isoformat(),
                "operation_type": "system_cleanup",
                "user_id": "system",
                "user_name": "系统自动",
                "action": "自动清理",
                "target": "操作日志",
                "status": "success",
                "details": {
                    "cleaned_count": cleaned_count,
                    "cutoff_date": cutoff_date_str,
                    "cleanup_type": "auto",
                    "retention_days": 180
                }
            }

            # 将自动清理日志添加到开头
            filtered_logs.insert(0, auto_cleanup_log)
            save_operation_logs_to_file(filtered_logs)

            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": f"自动清理完成，共清理 {cleaned_count} 条操作日志",
                    "data": {
                        "cleaned_count": cleaned_count,
                        "cutoff_date": cutoff_date_str,
                        "cleanup_type": "auto"
                    }
                }
            )
        else:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="保存操作日志失败"
            )
    except Exception as e:
        logger.error(f"自动清理操作日志失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"自动清理操作日志失败: {str(e)}"
        )


######################################### 历史数据 #########################################
# OUTPUT_ROOT = Path("/home/ubuntu/Projects/LeafDepot/output")
# 改用相对路径
OUTPUT_ROOT = Path(__file__).parent.parent.parent / "output"


def parse_task_date_from_filename(filename: str) -> Optional[datetime]:
    """从任务ID解析日期"""
    try:
        # 提取数字部分
        import re
        numbers = re.findall(r'\d+', filename)

        if numbers:
            # 取最长的数字串（可能是日期）
            longest_num = max(numbers, key=len)

            # 尝试解析为日期
            if len(longest_num) >= 8:
                date_str = longest_num[:8]
                year = int(date_str[:4])
                month = int(date_str[4:6])
                day = int(date_str[6:8])

                # 验证日期有效性
                if 2000 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
                    return datetime(year, month, day)

        # 如果无法解析，尝试从文件名中提取其他格式
        # 例如：HS2026011817 -> 2026-01-18
        pattern = r'(\d{4})(\d{2})(\d{2})'
        match = re.search(pattern, filename)
        if match:
            year, month, day = map(int, match.groups())
            return datetime(year, month, day)

        return None

    except Exception:
        return None


@app.delete("/api/history/task/{task_id}")
async def delete_history_task(task_id: str, request: Request):
    """
    按任务ID删除历史盘点文件
    权限要求：前端在请求头中传递 X-User-Level: admin
    """
    try:
        # 1. 从前端获取 userLevel
        user_level = request.headers.get("X-User-Level")
        if user_level != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="权限不足，仅管理员可删除历史数据"
            )

        # 2. 路径安全检查
        import re
        if not re.match(r'^[A-Za-z0-9_-]+$', task_id):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="任务ID包含非法字符"
            )

        history_tasks_dir = OUTPUT_ROOT / "history_data"
        xlsx_file = history_tasks_dir / f"{task_id}.xlsx"
        xls_file = history_tasks_dir / f"{task_id}.xls"
        target_file = None
        if xlsx_file.exists():
            target_file = xlsx_file
        elif xls_file.exists():
            target_file = xls_file
        else:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"任务 {task_id} 的历史文件不存在"
            )

        # 3. 删除文件
        target_file.unlink()
        logger.info(f"已删除历史任务文件: {target_file.name}")

        # 4. 记录操作日志（可选，记录 user_level）
        client_host = request.client.host if request.client else "unknown"
        log_operation(
            operation_type="system_cleanup",
            action="删除历史任务",
            user_name="前端传递",
            target=task_id,
            status="success",
            ip_address=client_host,
            details={
                "task_id": task_id,
                "filename": target_file.name,
                "file_path": str(target_file),
                "user_level": user_level
            }
        )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": f"任务 {task_id} 已成功删除",
                "data": {
                    "task_id": task_id,
                    "filename": target_file.name
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"删除历史任务失败 {task_id}: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"删除历史任务失败: {str(e)}"
        )


def is_task_expired(task_date: datetime) -> bool:
    """检查任务是否过期（超过6个月）"""
    now = datetime.now()
    six_months_ago = now.replace(year=now.year - (1 if now.month <= 6 else 0),
                                 month=(now.month - 6) % 12 or 12)

    if six_months_ago.month == 12 and now.month <= 6:
        six_months_ago = six_months_ago.replace(year=six_months_ago.year - 1)

    return task_date < six_months_ago


@app.get("/api/history/tasks")
async def get_history_tasks():
    """
    获取历史任务列表
    读取指定路径下的所有xlsx文件，解析为历史任务列表
    """
    try:
        print(f"文件位置: {OUTPUT_ROOT}")
        history_tasks_dir = OUTPUT_ROOT / "history_data"

        # 确保目录存在
        history_tasks_dir.mkdir(parents=True, exist_ok=True)

        # 查找所有xlsx文件
        xlsx_files = list(history_tasks_dir.glob("*.xlsx"))

        tasks = []

        for xlsx_file in xlsx_files:
            try:
                # 从文件名解析任务ID（去掉扩展名）
                task_id = xlsx_file.stem

                # 尝试从任务ID解析日期
                task_date = parse_task_date_from_filename(task_id)

                # 检查是否过期（超过6个月）
                is_expired = is_task_expired(task_date) if task_date else False

                tasks.append({
                    "taskId": task_id,
                    "taskDate": task_date.isoformat() if task_date else None,
                    "fileName": xlsx_file.name,
                    "isExpired": is_expired,
                    "filePath": str(xlsx_file.relative_to(OUTPUT_ROOT))
                })
            except Exception as e:
                logger.error(f"解析历史任务文件失败 {xlsx_file.name}: {str(e)}")
                continue

        # 按日期倒序排序（最新的在前）
        tasks.sort(key=lambda x: x.get("taskDate", ""), reverse=True)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取历史任务成功",
                "data": {
                    "tasks": tasks,
                    "total": len(tasks)
                }
            }
        )

    except Exception as e:
        logger.error(f"获取历史任务列表失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取历史任务列表失败: {str(e)}"
        )


@app.get("/api/history/task/{task_id}")
async def get_history_task_details(task_id: str):
    """
    获取历史任务详情
    读取与任务编号同名的xlsx文件，解析其中数据

    Args:
        task_id: 任务编号，如 HS2026011817
    """
    try:
        history_tasks_dir = OUTPUT_ROOT / "history_data"

        # 查找对应的xlsx文件
        xlsx_file = history_tasks_dir / f"{task_id}.xlsx"

        if not xlsx_file.exists():
            # 尝试查找其他可能的扩展名
            possible_files = list(history_tasks_dir.glob(f"{task_id}.*"))
            if possible_files:
                xlsx_file = possible_files[0]
            else:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"历史任务文件不存在: {task_id}.xlsx"
                )

        # 检查文件是否存在
        if not xlsx_file.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"文件不存在: {xlsx_file}"
            )

        logger.info(f"读取Excel文件: {xlsx_file}")

        # 方法1: 使用openpyxl读取Excel文件
        try:
            import openpyxl
            logger.info("使用openpyxl读取Excel文件")

            # 加载工作簿
            workbook = openpyxl.load_workbook(str(xlsx_file), data_only=True)

            # 获取第一个工作表
            sheet_name = workbook.sheetnames[0]
            worksheet = workbook[sheet_name]

            # 读取表头
            headers = []
            for cell in worksheet[1]:
                if cell.value:
                    headers.append(str(cell.value))

            logger.info(f"Excel表头: {headers}")

            # 读取数据行
            details = []
            row_index = 0

            for row in worksheet.iter_rows(min_row=2, values_only=True):
                row_index += 1

                # 跳过空行
                if all(cell is None for cell in row):
                    continue

                # 创建行字典
                row_dict = {}
                for i, header in enumerate(headers):
                    if i < len(row):
                        row_dict[header] = row[i]

                # 规范化数据格式
                detail = {
                    "序号": row_dict.get("序号") or row_index,
                    "品规名称": str(row_dict.get("品规名称", "")),
                    "储位名称": str(row_dict.get("储位名称", "")),
                    "实际品规": str(row_dict.get("实际品规", row_dict.get("品规名称", ""))),
                    "库存数量": int(row_dict.get("库存数量", 0) or 0),
                    "实际数量": int(row_dict.get("实际数量", 0) or 0),
                    "差异": str(row_dict.get("差异", "一致")),
                    "照片1路径": str(row_dict.get("照片1路径", "/3D_CAMERA/MAIN.JPEG")),
                    "照片2路径": str(row_dict.get("照片2路径", "/3D_CAMERA/DEPTH.JPEG")),
                    "照片3路径": str(row_dict.get("照片3路径", "/SCAN_CAMERA_1/1.JPEG")),
                    "照片4路径": str(row_dict.get("照片4路径", "/SCAN_CAMERA_2/1.JPEG")),
                }

                # 计算差异（如果未提供）
                if detail["差异"] == "一致" and detail["库存数量"] != detail["实际数量"]:
                    diff = detail["实际数量"] - detail["库存数量"]
                    if diff > 0:
                        detail["差异"] = f"多{diff}件"
                    else:
                        detail["差异"] = f"少{abs(diff)}件"

                # 如果差异为空，计算差异
                elif not detail["差异"] or detail["差异"].strip() == "":
                    if detail["库存数量"] == detail["实际数量"]:
                        detail["差异"] = "一致"
                    else:
                        diff = detail["实际数量"] - detail["库存数量"]
                        if diff > 0:
                            detail["差异"] = f"多{diff}件"
                        else:
                            detail["差异"] = f"少{abs(diff)}件"

                details.append(detail)

            logger.info(f"成功读取 {len(details)} 条记录")

        except ImportError:
            logger.warning("openpyxl未安装，尝试使用pandas")

            # 方法2: 使用pandas读取Excel文件
            try:
                import pandas as pd
                logger.info("使用pandas读取Excel文件")

                # 读取Excel文件
                df = pd.read_excel(str(xlsx_file))

                details = []
                for index, row in df.iterrows():
                    detail = {
                        "序号": row.get("序号") or (index + 1),
                        "品规名称": str(row.get("品规名称", "")),
                        "储位名称": str(row.get("储位名称", "")),
                        "实际品规": str(row.get("实际品规", row.get("品规名称", ""))),
                        "库存数量": int(row.get("库存数量", 0) or 0),
                        "实际数量": int(row.get("实际数量", 0) or 0),
                        "差异": str(row.get("差异", "一致")),
                        "照片1路径": str(row.get("照片1路径", "/3D_CAMERA/MAIN.JPEG")),
                        "照片2路径": str(row.get("照片2路径", "/3D_CAMERA/DEPTH.JPEG")),
                        "照片3路径": str(row.get("照片3路径", "/SCAN_CAMERA_1/1.JPEG")),
                        "照片4路径": str(row.get("照片4路径", "/SCAN_CAMERA_2/1.JPEG")),
                    }

                    # 计算差异（如果未提供）
                    if detail["差异"] == "一致" and detail["库存数量"] != detail["实际数量"]:
                        diff = detail["实际数量"] - detail["库存数量"]
                        if diff > 0:
                            detail["差异"] = f"多{diff}件"
                        else:
                            detail["差异"] = f"少{abs(diff)}件"

                    # 如果差异为空，计算差异
                    elif not detail["差异"] or pd.isna(detail["差异"]):
                        if detail["库存数量"] == detail["实际数量"]:
                            detail["差异"] = "一致"
                        else:
                            diff = detail["实际数量"] - detail["库存数量"]
                            if diff > 0:
                                detail["差异"] = f"多{diff}件"
                            else:
                                detail["差异"] = f"少{abs(diff)}件"

                    details.append(detail)

                logger.info(f"成功读取 {len(details)} 条记录")

            except ImportError:
                logger.error("未找到Excel解析库，请安装openpyxl或pandas")
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="Excel解析库未安装，请安装openpyxl或pandas"
                )

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取历史任务详情成功",
                "data": {
                    "taskId": task_id,
                    "fileName": xlsx_file.name,
                    "details": details,
                    "totalBins": len(details)
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取历史任务详情失败 {task_id}: {str(e)}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取历史任务详情失败: {str(e)}"
        )


@app.get("/api/history/task/{task_id}/bin/{bin_location}")
async def get_history_bin_detail(task_id: str, bin_location: str):
    """
    获取历史任务中特定储位的详情

    Args:
        task_id: 任务编号
        bin_location: 储位名称
    """
    try:
        # 先获取整个任务的详情
        from fastapi import BackgroundTasks
        import asyncio

        # 创建模拟请求来调用上面的接口
        task_details_response = await get_history_task_details(task_id)

        if task_details_response.status_code != 200:
            raise HTTPException(
                status_code=task_details_response.status_code,
                detail=task_details_response.body.decode(
                ) if task_details_response.body else "获取任务详情失败"
            )

        # 解析响应数据
        response_data = json.loads(task_details_response.body.decode())
        if response_data["code"] != 200:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=response_data["message"]
            )

        # 查找特定储位
        details = response_data["data"]["details"]
        bin_detail = None
        for detail in details:
            if detail["储位名称"] == bin_location:
                bin_detail = detail
                break

        if not bin_detail:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"储位 {bin_location} 不存在于任务 {task_id} 中"
            )

        # 获取该储位的图片
        image_urls = []
        photo_fields = ["照片1路径", "照片2路径", "照片3路径", "照片4路径"]

        for i, photo_field in enumerate(photo_fields, 1):
            photo_path = bin_detail.get(photo_field, "")
            if photo_path:
                # 构建完整的图片路径
                image_url = f"/api/history/image?taskNo={task_id}&binLocation={bin_location}&cameraType={photo_path.split('/')[1]}&filename={Path(photo_path).name}"
                image_urls.append({
                    "index": i,
                    "path": photo_path,
                    "url": image_url,
                    "cameraType": photo_path.split('/')[1] if '/' in photo_path else "未知",
                    "filename": Path(photo_path).name
                })

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取储位详情成功",
                "data": {
                    "taskId": task_id,
                    "binLocation": bin_location,
                    "detail": bin_detail,
                    "imageUrls": image_urls,
                    "hasImages": len(image_urls) > 0
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取历史任务储位详情失败 {task_id}/{bin_location}: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取储位详情失败: {str(e)}"
        )


@app.get("/api/history/image")
async def get_history_image(
    taskNo: str,
    binLocation: str,
    cameraType: str,
    filename: str,
    source: str = "capture_img"  # 新增参数：source可以是"output"、"capture_img"或"history"
):
    """
    获取盘点任务中的图片

    Args:
        taskNo: 任务编号
        binLocation: 储位名称
        cameraType: 相机类型
        filename: 文件名
    """
    try:
        img_root = IMG_ROOT

        # 构建基本路径
        base_path = img_root / taskNo / binLocation / cameraType

        # 定义可能的图片扩展名
        image_extensions = ['.jpg', '.jpeg', '.png', '.bmp', '.gif', '.webp']

        image_path = None

        # 尝试查找不同扩展名的文件
        for ext in image_extensions:
            test_path = base_path / f"{filename}{ext}"
            if test_path.exists():
                image_path = test_path
                break

        # 如果还没找到，尝试查找大写扩展名的文件
        if not image_path:
            for ext in image_extensions:
                test_path = base_path / f"{filename}{ext.upper()}"
                if test_path.exists():
                    image_path = test_path
                    break

        if not image_path:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"图片不存在: {filename} (路径: {base_path})"
            )

        # 读取图片文件
        with open(image_path, "rb") as f:
            image_data = f.read()

        # 根据文件扩展名确定媒体类型
        image_ext = image_path.suffix.lower()
        if image_ext in ['.jpg', '.jpeg']:
            media_type = "image/jpeg"
        elif image_ext == '.png':
            media_type = "image/png"
        elif image_ext == '.bmp':
            media_type = "image/bmp"
        elif image_ext == '.gif':
            media_type = "image/gif"
        elif image_ext == '.webp':
            media_type = "image/webp"
        else:
            # 默认使用jpeg
            media_type = "image/jpeg"
            logger.warning(f"未知的图片扩展名 {image_ext}，使用默认媒体类型")

        # 记录成功信息
        logger.info(
            f"成功返回图片: {image_path.name}, 大小: {len(image_data)} bytes, 类型: {media_type}")

        # 添加 CORS 头
        headers = {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*",
            "Access-Control-Expose-Headers": "*",
            "Cache-Control": "public, max-age=3600"
        }

        return Response(content=image_data, media_type=media_type, headers=headers)

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"获取图片失败: {str(e)}")
        # 返回 404 时也要添加 CORS 头
        headers = {
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET, OPTIONS",
            "Access-Control-Allow-Headers": "*"
        }
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"图片不存在: {filename}",
            headers=headers
        )


# 操作记录接口
@app.get("/api/operation/logs/recent")
async def get_recent_operation_logs(limit: int = 5):
    """
    获取最近的操作记录（默认最近5条，6个月内）

    Args:
        limit: 返回记录数量，默认5条
    """
    try:
        operations = get_recent_operations(limit=limit, days=180)

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取操作记录成功",
                "data": {
                    "operations": operations,
                    "total": len(operations),
                    "limit": limit
                }
            }
        )

    except Exception as e:
        logger.error(f"获取操作记录失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取操作记录失败: {str(e)}"
        )


@app.get("/api/operation/logs/all")
async def get_all_operation_logs(
    days: int = 180,
    operation_type: Optional[str] = None,
    status: Optional[str] = None
):
    """
    获取所有操作记录（可筛选）

    Args:
        days: 查询天数，默认180天（6个月）
        operation_type: 操作类型筛选
        status: 状态筛选
    """
    try:
        operations = get_all_operations(days=days)

        # 应用筛选
        filtered_operations = operations

        if operation_type:
            filtered_operations = [op for op in filtered_operations
                                   if op.get("operation_type") == operation_type]

        if status:
            filtered_operations = [op for op in filtered_operations
                                   if op.get("status") == status]

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取操作记录成功",
                "data": {
                    "operations": filtered_operations,
                    "total": len(filtered_operations),
                    "filtered_total": len(operations),
                    "days": days
                }
            }
        )

    except Exception as e:
        logger.error(f"获取操作记录失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取操作记录失败: {str(e)}"
        )


# 5. 历史数据清理接口
class HistoryCleanupRequest(BaseModel):
    """历史数据清理请求模型"""
    cutoff_date: Optional[str] = None  # 截止日期，格式: YYYY-MM-DD
    days: Optional[int] = 180  # 保留天数，默认180天（6个月）


@app.post("/api/history/cleanup")
async def cleanup_history_data(request: Request):
    """
    清理历史数据
    删除指定天数之前的历史任务文件

    Request body:
        cutoff_date: 截止日期（可选），格式: YYYY-MM-DD
        days: 保留天数（可选），默认180天（6个月）

    Returns:
        code: 状态码
        message: 消息
        data: 清理结果
            cleaned_count: 清理的文件数量
            cutoff_date: 实际使用的截止日期
            retention_days: 实际保留天数
            cleaned_files: 清理的文件列表
    """
    try:
        # 解析请求参数
        data = await request.json()

        # 获取参数
        cutoff_date_str = data.get("cutoff_date")
        days = data.get("days", 180)

        # 计算截止日期
        if cutoff_date_str:
            # 使用提供的截止日期
            try:
                cutoff_date = datetime.strptime(cutoff_date_str, "%Y-%m-%d")
            except ValueError:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"截止日期格式错误，应为 YYYY-MM-DD，实际为: {cutoff_date_str}"
                )
        else:
            # 使用天数计算截止日期
            cutoff_date = datetime.now() - timedelta(days=days)

        cutoff_date_str = cutoff_date.strftime("%Y-%m-%d")
        logger.info(f"开始清理历史数据，截止日期: {cutoff_date_str}")

        # 获取历史数据目录
        history_tasks_dir = OUTPUT_ROOT / "history_data"

        if not history_tasks_dir.exists():
            logger.warning(f"历史数据目录不存在: {history_tasks_dir}")
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "历史数据目录不存在，无需清理",
                    "data": {
                        "cleaned_count": 0,
                        "cutoff_date": cutoff_date_str,
                        "retention_days": days,
                        "cleaned_files": []
                    }
                }
            )

        # 查找所有历史数据文件
        all_files = list(history_tasks_dir.glob("*.xlsx"))
        all_files.extend(history_tasks_dir.glob("*.xls"))

        cleaned_files = []
        cleaned_count = 0

        for file_path in all_files:
            try:
                # 从文件名解析任务日期
                task_date = parse_task_date_from_filename(file_path.stem)

                if task_date is None:
                    # 如果无法解析日期，跳过该文件
                    logger.warning(f"无法从文件名解析日期: {file_path.name}")
                    continue

                # 检查是否需要删除
                if task_date < cutoff_date:
                    # 删除文件
                    file_path.unlink()
                    cleaned_files.append({
                        "filename": file_path.name,
                        "task_date": task_date.isoformat(),
                        "file_path": str(file_path)
                    })
                    cleaned_count += 1
                    logger.info(
                        f"已删除历史数据文件: {file_path.name} (日期: {task_date})")

            except Exception as e:
                logger.error(f"处理文件失败 {file_path.name}: {str(e)}")
                continue

        # 记录清理操作
        client_host = request.client.host if request.client else "unknown"
        log_operation(
            operation_type="system_cleanup",
            action="清理历史数据",
            status="success",
            ip_address=client_host,
            details={
                "cleaned_count": cleaned_count,
                "cutoff_date": cutoff_date_str,
                "retention_days": days,
                "cleaned_files": cleaned_files,
                "request_ip": client_host
            }
        )

        logger.info(f"历史数据清理完成，共删除 {cleaned_count} 个文件")

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": f"历史数据清理完成，共删除 {cleaned_count} 个文件",
                "data": {
                    "cleaned_count": cleaned_count,
                    "cutoff_date": cutoff_date_str,
                    "retention_days": days,
                    "cleaned_files": cleaned_files
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"清理历史数据失败: {str(e)}")

        # 记录失败操作
        client_host = request.client.host if request.client else "unknown"
        log_operation(
            operation_type="system_cleanup",
            action="清理历史数据",
            status="failed",
            ip_address=client_host,
            details={
                "error": str(e),
                "request_ip": client_host
            }
        )

        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"清理历史数据失败: {str(e)}"
        )


# 6. 添加历史数据清理记录函数
async def log_history_cleanup(cleaned_count: int, cutoff_date: str, details: Optional[Dict] = None):
    """记录历史数据清理操作"""
    log_operation(
        operation_type="system_cleanup",
        action="清理历史数据",
        status="success",
        details={
            "cleaned_count": cleaned_count,
            "cutoff_date": cutoff_date,
            "retention_days": 180,
            ** (details or {})
        }
    )


@app.get("/api/dashboard/stats")
async def get_dashboard_stats():
    """
    获取Dashboard统计数据

    从原始位置读取Excel文件：
    - 储位数：services/sim/lms/bins_data.xlsx (总行数 - 1)
    - 品类数：archive/TobaccoCaseMultiImageReader/烟箱信息汇总完整版.xlsx (总行数 - 1)

    Returns:
        code: 状态码
        message: 消息
        data:
            bin_count: 储位数量
            category_count: 品类数量
    """
    try:
        import openpyxl

        # 原始文件路径
        bins_file = _project_root / "services" / "sim" / "lms" / "bins_data.xlsx"
        category_file = _project_root / "shared" / "data" / "烟箱信息汇总完整版.xlsx"

        bin_count = 0
        category_count = 0

        # 读取储位数
        if bins_file.exists():
            try:
                wb = openpyxl.load_workbook(bins_file, read_only=True)
                ws = wb.active
                bin_count = max(0, ws.max_row - 1)  # 总行数 - 1 (减去表头)
                wb.close()
                logger.info(f"读取储位数据: {bins_file}, 数量: {bin_count}")
            except Exception as e:
                logger.error(f"读取储位文件失败: {str(e)}")

        # 读取品类数
        if category_file.exists():
            try:
                wb = openpyxl.load_workbook(category_file, read_only=True)
                ws = wb.active
                category_count = max(0, ws.max_row - 1)  # 总行数 - 1 (减去表头)
                wb.close()
                logger.info(f"读取品类数据: {category_file}, 数量: {category_count}")
            except Exception as e:
                logger.error(f"读取品类文件失败: {str(e)}")

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "获取Dashboard统计数据成功",
                "data": {
                    "bin_count": bin_count,
                    "category_count": category_count
                }
            }
        )

    except Exception as e:
        logger.error(f"获取Dashboard统计数据失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取Dashboard统计数据失败: {str(e)}"
        )


@app.get("/api/history/monthly-count")
async def get_monthly_inventory_count():
    """
    获取本月盘点次数和准确率

    通过解析 history_data 目录下的文件名来判断本月盘点次数
    文件名格式: HS2026011917 (前缀 + YYYYMMDD + 序号)
    准确率：读取Excel第H列（差异列），"一致"的数量/总数量

    Returns:
        code: 状态码
        message: 消息
        data: 
            count: 本月盘点次数
            current_month: 当前月份 (YYYY-MM)
            accuracy: 准确率 (百分比，无盘点时为 null)
            files: 本月的盘点文件列表
    """
    try:
        import openpyxl

        # 获取历史数据目录
        history_tasks_dir = OUTPUT_ROOT / "history_data"

        if not history_tasks_dir.exists():
            logger.warning(f"历史数据目录不存在: {history_tasks_dir}")
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={
                    "code": 200,
                    "message": "历史数据目录不存在",
                    "data": {
                        "count": 0,
                        "current_month": datetime.now().strftime("%Y-%m"),
                        "accuracy": None,
                        "files": []
                    }
                }
            )

        # 获取当前年月
        current_year = datetime.now().year
        current_month = datetime.now().month
        current_month_str = f"{current_year}-{current_month:02d}"

        # 查找所有历史数据文件
        all_files = list(history_tasks_dir.glob("*.xlsx"))
        all_files.extend(history_tasks_dir.glob("*.xls"))

        monthly_files = []

        for file_path in all_files:
            try:
                # 从文件名解析日期
                # 文件名格式: HS2026011917 -> 提取 20260119
                filename = file_path.stem
                if len(filename) >= 10 and filename[:2] == "HS":
                    date_str = filename[2:10]  # 提取 YYYYMMDD
                    try:
                        file_date = datetime.strptime(date_str, "%Y%m%d")
                        file_year = file_date.year
                        file_month = file_date.month

                        # 判断是否为当前月份
                        if file_year == current_year and file_month == current_month:
                            monthly_files.append({
                                "filename": file_path.name,
                                "filepath": str(file_path),
                                "date": file_date.strftime("%Y-%m-%d"),
                                "task_no": filename
                            })
                    except ValueError:
                        # 日期解析失败，跳过
                        continue

            except Exception as e:
                logger.error(f"处理文件失败 {file_path.name}: {str(e)}")
                continue

        count = len(monthly_files)

        # 计算准确率：读取Excel第H列（差异列）
        accuracy = None
        if count > 0:
            total_count = 0
            match_count = 0

            for file_info in monthly_files:
                try:
                    wb = openpyxl.load_workbook(
                        file_info["filepath"], read_only=True)
                    ws = wb.active

                    # 从第2行开始读取（第1行是表头）
                    for row_idx in range(2, ws.max_row + 1):
                        cell_h = ws.cell(row=row_idx, column=8)  # 第H列
                        if cell_h.value is not None:
                            total_count += 1
                            if str(cell_h.value).strip() == "一致":
                                match_count += 1

                    wb.close()
                except Exception as e:
                    logger.error(f"读取文件失败 {file_info['filename']}: {str(e)}")
                    continue

            # 计算准确率
            if total_count > 0:
                accuracy = round((match_count / total_count) * 100, 1)
                logger.info(
                    f"本月盘点准确率: {accuracy}% ({match_count}/{total_count})")
            else:
                accuracy = None

        logger.info(f"本月({current_month_str})盘点次数: {count}")

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": f"获取本月盘点次数成功",
                "data": {
                    "count": count,
                    "current_month": current_month_str,
                    "accuracy": accuracy,
                    "files": monthly_files
                }
            }
        )

    except Exception as e:
        logger.error(f"获取本月盘点次数失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取本月盘点次数失败: {str(e)}"
        )


@app.post("/api/inventory/save-results")
async def save_inventory_results(request: Request):
    """
    保存盘点结果并生成Excel文件

    Args:
        request: 请求体，包含 taskNo 和 inventoryResults
    """
    try:
        data = await request.json()
        task_no = data.get("taskNo")
        inventory_results = data.get("inventoryResults", [])

        if not task_no:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="任务编号不能为空"
            )

        if not inventory_results:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="盘点结果不能为空"
            )

        logger.info(f"保存盘点结果: {task_no}, 共 {len(inventory_results)} 个储位")

        # 构建Excel数据
        excel_data = []
        for i, result in enumerate(inventory_results, 1):
            # 去掉照片路径中的 /{taskNo}/{binLocation} 前缀
            photo3dPath = result.get("photo3dPath", "")
            photoDepthPath = result.get("photoDepthPath", "")
            photoScan1Path = result.get("photoScan1Path", "")
            photoScan2Path = result.get("photoScan2Path", "")

            # 如果路径包含taskNo和binLocation，则去掉
            if photo3dPath and photo3dPath.startswith(f"/{task_no}/"):
                photo3dPath = photo3dPath[len(
                    f"/{task_no}/{result.get('binLocation', '')}/"):]
            if photoDepthPath and photoDepthPath.startswith(f"/{task_no}/"):
                photoDepthPath = photoDepthPath[len(
                    f"/{task_no}/{result.get('binLocation', '')}/"):]
            if photoScan1Path and photoScan1Path.startswith(f"/{task_no}/"):
                photoScan1Path = photoScan1Path[len(
                    f"/{task_no}/{result.get('binLocation', '')}/"):]
            if photoScan2Path and photoScan2Path.startswith(f"/{task_no}/"):
                photoScan2Path = photoScan2Path[len(
                    f"/{task_no}/{result.get('binLocation', '')}/"):]

            # 获取品规名称（优先使用specName，如果不存在则使用actualSpec）
            spec_name = result.get(
                "specName", "") or result.get("actualSpec", "")

            excel_data.append({
                "任务编号": task_no,
                "序号": i,
                "品规名称": spec_name,  # 品规名称
                "储位名称": result.get("binLocation", ""),
                "实际品规": result.get("actualSpec", ""),
                "库存数量": result.get("systemQuantity", 0),  # 库存数量
                "实际数量": result.get("actualQuantity", 0),
                "差异": result.get("difference", 0),
                "照片1路径": photo3dPath,
                "照片2路径": photoDepthPath,
                "照片3路径": photoScan1Path,
                "照片4路径": photoScan2Path,
            })

        # 创建DataFrame
        df = pd.DataFrame(excel_data)

        # 创建输出目录
        output_dir = Path(__file__).parent.parent.parent / \
            "output" / "history_data"
        output_dir.mkdir(parents=True, exist_ok=True)

        # 生成Excel文件路径
        xlsx_file = output_dir / f"{task_no}.xlsx"

        # 保存Excel文件
        with pd.ExcelWriter(xlsx_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='盘点结果')

            # 获取工作表对象，设置样式
            workbook = writer.book
            worksheet = writer.sheets['盘点结果']

            # 设置列宽
            for idx, col in enumerate(df.columns, 1):
                max_length = max(
                    df[col].astype(str).apply(len).max(),
                    len(col)
                )
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[chr(
                    64 + idx)].width = adjusted_width

            # 设置表头样式
            from openpyxl.styles import Font, PatternFill, Alignment
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4CAF50", end_color="4CAF50", fill_type="solid")
            header_alignment = Alignment(
                horizontal="center", vertical="center")

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 设置数据行样式
            data_alignment = Alignment(horizontal="center", vertical="center")
            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = data_alignment

        logger.info(f"成功生成Excel文件: {xlsx_file}")

        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={
                "code": 200,
                "message": "保存成功",
                "data": {
                    "taskNo": task_no,
                    "xlsxFile": str(xlsx_file),
                    "xlsxUrl": f"/api/inventory/download-xlsx?taskNo={task_no}",
                    "count": len(inventory_results)
                }
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"保存盘点结果失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"保存盘点结果失败: {str(e)}"
        )


@app.get("/api/inventory/download-xlsx")
async def download_xlsx(taskNo: str):
    """
    下载Excel文件

    Args:
        taskNo: 任务编号
    """
    try:
        output_dir = Path(__file__).parent.parent.parent / \
            "output" / "history_data"
        xlsx_file = output_dir / f"{taskNo}.xlsx"

        if not xlsx_file.exists():
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Excel文件不存在: {taskNo}.xlsx"
            )

        # 读取Excel文件
        with open(xlsx_file, "rb") as f:
            excel_data = f.read()

        return Response(
            content=excel_data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={taskNo}.xlsx"
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"下载Excel文件失败: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"下载Excel文件失败: {str(e)}"
        )


if __name__ == "__main__":
    # 确保日志配置正确（已经在文件开头配置，这里不再重复配置）
    logger.info("🚀 Gateway服务启动")
    logger.info("📡 API地址: http://0.0.0.0:8000")
    logger.info("📚 API文档: http://localhost:8000/docs")
    logger.info("🔍 测试页面: http://localhost:8080/test_detect.html")
    logger.info(f"📝 日志文件: {_log_filename}")
    # 修复后的代码
    logger.info(
        f"📝 前端日志文件: {_debug_log_dir}/frontend_{datetime.now().strftime('%Y%m%d')}.log")
    logger.info("\n按 Ctrl+C 停止服务\n")

    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")
