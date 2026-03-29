"""
LMS 系统对接路由
"""
import json
import logging
import requests
from fastapi import APIRouter, Request, HTTPException, status
from fastapi.responses import JSONResponse

from services.api.shared.config import LMS_BASE_URL, logger
import services.api.custom_utils as custom_utils

router = APIRouter(prefix="/lms", tags=["lms"])


@router.get("/getLmsBin")
async def get_lms_bin(authToken: str, useLocal: bool = False):
    """获取库位信息

    Args:
        authToken: 认证令牌
        useLocal: 是否直接从本地 bins_data.xlsx 读取（默认 False）

    当 useLocal=True 或 LMS 服务不可用时，从本地 bins_data.xlsx 读取数据
    """
    # 如果指定使用本地数据，直接读取
    if useLocal:
        logger.info("使用本地文件读取库位数据 (useLocal=True)")
        return _get_local_bins_data()

    from pathlib import Path
    from fastapi.responses import JSONResponse
    from fastapi import status
    from services.api.shared.config import project_root

    try:
        lms_bin_url = f"{LMS_BASE_URL}/third/api/v1/lmsToRcsService/getLmsBin"
        headers = {"authToken": authToken}
        response = requests.get(lms_bin_url, headers=headers, timeout=10)

        if response.status_code == 200:
            try:
                uncompressed_data = custom_utils.decompress_and_decode(response.text)
                logger.info("成功解压缩并解析库位数据")
                return JSONResponse(uncompressed_data)
            except Exception as e:
                logger.error(f"解压缩库位数据失败: {str(e)}")
                return _get_local_bins_data()
        else:
            logger.warning(f"LMS返回错误状态码 {response.status_code}，从本地文件读取库位数据")
            return _get_local_bins_data()
    except requests.exceptions.Timeout:
        logger.warning("LMS服务响应超时，从本地文件读取库位数据")
        return _get_local_bins_data()
    except requests.exceptions.ConnectionError:
        logger.warning("无法连接到LMS服务,从本地文件读取库位数据")
        return _get_local_bins_data()
    except Exception as e:
        logger.error(f"获取库位信息请求失败: {str(e)}")
        return _get_local_bins_data()


def _get_local_bins_data():
    """从本地 bins_data.xlsx 读取库位数据作为备用"""
    from pathlib import Path
    from fastapi.responses import JSONResponse
    from fastapi import status

    try:
        from services.api.shared.config import project_root
        bins_data_path = project_root / "services" / "sim" / "lms" / "bins_data.xlsx"

        if not bins_data_path.exists():
            logger.error(f"本地库位文件不存在: {bins_data_path}")
            return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={"code": 404, "message": "本地库位文件不存在", "data": []}
            )

        from openpyxl import load_workbook
        wb = load_workbook(bins_data_path, read_only=True)
        ws = wb.active

        # 读取表头
        headers = [cell.value for cell in ws[1]]

        bins_list = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):  # 跳过空行
                continue

            row_dict = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}

            # 构建与 LMS 接口兼容的数据格式
            bin_item = {
                "whCode": row_dict.get("仓库编码", ""),
                "areaCode": row_dict.get("区域编码", ""),
                "areaName": row_dict.get("区域名称", ""),
                "binCode": row_dict.get("储位编码", ""),
                "binDesc": row_dict.get("储位名称", ""),
                "maxQty": row_dict.get("最大数量", 0),
                "binStatus": row_dict.get("储位状态", ""),
                "tobaccoQty": row_dict.get("数量(万支)", 0) or 0,
                "tobaccoCode": row_dict.get("烟草编码", ""),
                "tobaccoName": row_dict.get("品规名称", ""),
                "rcsCode": row_dict.get("RCS编码", ""),
            }
            bins_list.append(bin_item)

        wb.close()

        logger.info(f"从本地文件读取 {len(bins_list)} 条库位数据作为备用")
        return JSONResponse(
            status_code=status.HTTP_200_OK,
            content={"code": 200, "message": "从本地文件获取库位数据成功", "data": bins_list}
        )

    except Exception as e:
        logger.error(f"读取本地库位文件失败: {str(e)}")
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"code": 500, "message": f"读取本地库位文件失败: {str(e)}", "data": []}
        )


@router.get("/getCountTasks")
async def get_count_tasks(authToken: str):
    """获取盘点任务，调用LMS的getCountTasks接口"""
    try:
        logger.info(f"收到获取盘点任务请求，authToken: {authToken[:20]}...")
        lms_tasks_url = f"{LMS_BASE_URL}/third/api/v1/lmsToRcsService/getCountTasks"
        headers = {"authToken": authToken}
        response = requests.get(lms_tasks_url, headers=headers, timeout=30)

        if response.status_code == 200:
            try:
                uncompressed_data = custom_utils.decompress_and_decode(response.text)
                logger.info("成功解压缩并解析盘点任务数据")
                return JSONResponse(uncompressed_data)
            except Exception as e:
                logger.error(f"解压缩盘点任务数据失败: {str(e)}")
                raise HTTPException(
                    status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                    detail="盘点任务数据解压缩失败"
                )
        else:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"LMS获取盘点任务失败: {response.text}"
            )
    except requests.exceptions.Timeout:
        raise HTTPException(status_code=status.HTTP_504_GATEWAY_TIMEOUT, detail="LMS服务响应超时")
    except requests.exceptions.ConnectionError:
        raise HTTPException(status_code=status.HTTP_503_SERVICE_UNAVAILABLE, detail="无法连接到LMS服务")
    except Exception as e:
        logger.error(f"获取盘点任务请求失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="获取盘点任务请求处理失败")


@router.post("/setTaskResults")
async def set_task_results(request: Request):
    """提交盘点任务结果，调用LMS的setTaskResults接口"""
    try:
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Unauthorized")

        data = await request.json()
        encoded_data = custom_utils.compress_and_encode(data)

        lms_results_url = f"{LMS_BASE_URL}/third/api/v1/RcsToLmsService/setTaskResults"
        headers = {"authToken": auth_token, "Content-Type": "text/plain"}
        response = requests.post(lms_results_url, data=encoded_data, headers=headers)

        if response.status_code == 200:
            return {"success": True, "message": "盘点结果已提交"}
        else:
            raise HTTPException(status_code=response.status_code, detail=f"LMS提交盘点结果失败: {response.text}")
    except Exception as e:
        logger.error(f"提交盘点结果请求失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="提交盘点结果请求处理失败")


@router.get("/getUsers")
async def get_users(request: Request):
    """获取所有用户信息，调用LMS的getUsers接口"""
    try:
        auth_token = request.query_params.get('authToken') or request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未提供认证令牌")

        lms_users_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/getUsers"
        headers = {"authToken": auth_token}
        response = requests.get(lms_users_url, headers=headers, timeout=10)

        if response.status_code == 200:
            result = response.json()
            logger.info(f"成功获取 {len(result.get('data', []))} 条用户数据")
            return result
        else:
            raise HTTPException(status_code=response.status_code, detail=f"LMS获取用户信息失败: {response.text}")
    except Exception as e:
        logger.error(f"获取用户信息请求失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="获取用户信息请求处理失败")


@router.post("/registerUser")
async def register_user(request: Request):
    """注册用户，调用LMS的registerUser接口"""
    try:
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未提供认证令牌")

        data = await request.json()
        lms_register_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/registerUser"
        headers = {"authToken": auth_token, "Content-Type": "application/json"}
        response = requests.post(lms_register_url, json=data, headers=headers, timeout=10)

        if response.status_code == 200:
            result = response.json()
            logger.info(f"用户注册成功: {data.get('userCode')}")
            return result
        else:
            raise HTTPException(status_code=response.status_code, detail=f"LMS注册用户失败: {response.text}")
    except Exception as e:
        logger.error(f"注册用户请求失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="注册用户请求处理失败")


@router.post("/deleteUser")
async def delete_user(request: Request):
    """删除用户，调用LMS的deleteUser接口"""
    try:
        auth_token = request.headers.get('authToken')
        if not auth_token:
            raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="未提供认证令牌")

        data = await request.json()
        lms_delete_url = f"{LMS_BASE_URL}/third/api/v1/userManagement/deleteUser"
        headers = {"authToken": auth_token, "Content-Type": "application/json"}
        response = requests.post(lms_delete_url, json=data, headers=headers, timeout=10)

        if response.status_code == 200:
            result = response.json()
            logger.info(f"用户删除成功: {data.get('userCode')}")
            return result
        else:
            raise HTTPException(status_code=response.status_code, detail=f"LMS删除用户失败: {response.text}")
    except Exception as e:
        logger.error(f"删除用户请求失败: {str(e)}")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="删除用户请求处理失败")
